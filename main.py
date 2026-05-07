import sys, os, io, re

if sys.platform == "win32":
    try:
        import ctypes
        ctypes.windll.kernel32.SetConsoleOutputCP(65001)
        ctypes.windll.kernel32.SetConsoleCP(65001)
    except Exception: pass
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")
    except Exception: pass
    os.environ.setdefault("PYTHONIOENCODING", "utf-8")

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3, datetime, webbrowser, urllib.parse
from collections import defaultdict

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


SOCIETY_NAME    = "My Apartment"
SOCIETY_ADDRESS = ""
MONTHLY_FEE      = 500
SOCIETY_START_FY = 2023

EXPENDITURE_CATEGORIES = [
    "Pump Maintenance", "Electricity",
    "Cleaning / Housekeeping", "Others/Misc Expense",
]

UNIT_LABEL = "Flat"
OWNER_LABEL = "Owner"

_DEFAULT_RESIDENTS = {}   # No hardcoded residents — add via Manage Units

RESIDENTS = {}

MONTHS = ["January","February","March","April","May","June",
          "July","August","September","October","November","December"]
FY_MONTHS = ["April","May","June","July","August","September",
             "October","November","December","January","February","March"]
_FY_APR_DEC = set(FY_MONTHS[:9])
_FY_JAN_MAR = set(FY_MONTHS[9:])

def fy_label(y):
    return f"FY {y}-{str(y+1)[-2:]}"

BG        = "#07080F"
SIDEBAR   = "#05060C"
SIDEBAR_H = "#0C1020"
SIDEBAR_S = "#131A2E"
ACCENT    = "#7C3AED"
ACCENT_H  = "#6D28D9"
ACCENT2   = "#C4B5FD"
SUCCESS   = "#10B981"
SUCCESS_H = "#059669"
DANGER    = "#F87171"
DANGER_H  = "#EF4444"
WARNING   = "#F59E0B"
WARNING_H = "#D97706"
CARD      = "#0E1220"
BORDER    = "#1C2540"
BORDER2   = "#283055"
TEXT      = "#EDF2FF"
TEXT2     = "#8B9DC3"
TEXT3     = "#3D4F72"
HEAD_I    = "#1E3A6A"
HEAD_E    = "#7C1D1D"
MUTED     = "#3D4F72"
WA_GRN    = "#25D366"
WA_GRN_H  = "#1DA851"
PILL_BG   = "#1A2652"
PILL_FG   = "#A78BFA"
WARN_BG   = "#2D1F07"
WARN_FG   = "#FCD34D"
ERR_BG    = "#2D0F0F"
ERR_FG    = "#FCA5A5"
OK_BG     = "#052E16"
OK_FG     = "#6EE7B7"
PURPLE    = "#A78BFA"
PURPLE_H  = "#8B5CF6"
TEAL      = "#0EA5E9"
TEAL_H    = "#0284C7"
ROW_ODD   = "#0E1220"
ROW_EVEN  = "#0A0D18"
ROW_WARN  = "#2D1F07"
ROW_SEL   = "#1A2652"

import json as _json

def _base_dir():
    return (os.path.dirname(sys.executable)
            if getattr(sys, "frozen", False)
            else os.path.dirname(os.path.abspath(__file__)))

_ACTIVE_DB = None

def _db_path():
    return _ACTIVE_DB or os.path.join(_base_dir(), "maintenance.db")

def _reg_path():
    return os.path.join(_base_dir(), "societies.json")

def _reg_load():
    try:
        with open(_reg_path(), "r", encoding="utf-8") as f:
            return _json.load(f)
    except Exception:
        return []

def _reg_save(data):
    with open(_reg_path(), "w", encoding="utf-8") as f:
        _json.dump(data, f, indent=2)

def society_list():
    return _reg_load()

def society_add(name, db_path):
    reg = _reg_load()
    if any(r["db"] == db_path for r in reg):
        return
    reg.append({"name": name, "db": db_path})
    _reg_save(reg)

def society_rename(db_path, new_name):
    reg = _reg_load()
    for r in reg:
        if r["db"] == db_path:
            r["name"] = new_name
    _reg_save(reg)

def society_remove(db_path):
    reg = [r for r in _reg_load() if r["db"] != db_path]
    _reg_save(reg)

def society_switch(db_path):
    global _ACTIVE_DB
    _ACTIVE_DB = db_path

def init_db():
    con = sqlite3.connect(_db_path())
    con.executescript("""
        CREATE TABLE IF NOT EXISTS flats (
            flat_no TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            mobile TEXT DEFAULT '',
            sort_order INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT, receipt_no TEXT UNIQUE NOT NULL,
            date TEXT NOT NULL, year_from INTEGER NOT NULL, year_to INTEGER NOT NULL,
            flat_no TEXT NOT NULL, owner_name TEXT NOT NULL, amount REAL NOT NULL,
            month_from TEXT DEFAULT '', month_to TEXT DEFAULT '',
            monthly_fee REAL DEFAULT 0, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS expenditure_accounts (
            id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT UNIQUE NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS expenditures (
            id INTEGER PRIMARY KEY AUTOINCREMENT, date TEXT NOT NULL,
            description TEXT NOT NULL, amount REAL NOT NULL,
            category TEXT DEFAULT 'Others/Misc Expense', account_id INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS manual_journals (
            id INTEGER PRIMARY KEY AUTOINCREMENT, date TEXT NOT NULL,
            account_type TEXT NOT NULL, account_id TEXT NOT NULL,
            entry_type TEXT NOT NULL, amount REAL NOT NULL,
            narration TEXT NOT NULL, fy_year INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS admission_fees (
            id INTEGER PRIMARY KEY AUTOINCREMENT, receipt_no TEXT UNIQUE NOT NULL,
            date TEXT NOT NULL, flat_no TEXT NOT NULL, owner_name TEXT NOT NULL,
            amount REAL NOT NULL, mobile TEXT DEFAULT '', paid INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL);
    """)
    for alter in [
        "ALTER TABLE payments ADD COLUMN monthly_fee REAL DEFAULT 0",
        "ALTER TABLE payments ADD COLUMN late_fee REAL DEFAULT 0",
        "ALTER TABLE expenditures ADD COLUMN category TEXT DEFAULT 'Others/Misc Expense'",
        "ALTER TABLE expenditures ADD COLUMN account_id INTEGER DEFAULT 1",
    ]:
        try: con.execute(alter); con.commit()
        except Exception: pass
    cur = con.cursor()
    cur.execute("SELECT COUNT(*) FROM flats")
    if cur.fetchone()[0] == 0:
        for i, (flat_no, info) in enumerate(_DEFAULT_RESIDENTS.items()):
            con.execute("INSERT INTO flats (flat_no, name, mobile, sort_order) VALUES (?,?,?,?)",
                        (flat_no, info["name"], info["mobile"], i))
        con.commit()
    cur.execute("SELECT COUNT(*) FROM expenditure_accounts")
    if cur.fetchone()[0] == 0:
        for i, cat in enumerate(EXPENDITURE_CATEGORIES, start=1):
            con.execute("INSERT INTO expenditure_accounts (id, name) VALUES (?,?)", (i, cat))
        con.commit()
    # Seed default settings if not present
    import json as _json
    _defaults = {
        "society_name":    SOCIETY_NAME,
        "society_address": SOCIETY_ADDRESS,
        "monthly_fee":     str(MONTHLY_FEE),
        "society_start_fy":str(SOCIETY_START_FY),
        "exp_categories":  _json.dumps(EXPENDITURE_CATEGORIES),
        "unit_label":      UNIT_LABEL,
    }
    for k, v in _defaults.items():
        con.execute("INSERT OR IGNORE INTO settings (key, value) VALUES (?, ?)", (k, v))
    con.commit()
    con.close()

def db_get_setting(key, default=""):
    try:
        con = sqlite3.connect(_db_path()); cur = con.cursor()
        cur.execute("SELECT value FROM settings WHERE key=?", (key,))
        row = cur.fetchone(); con.close()
        return row[0] if row else default
    except Exception: return default

def db_set_setting(key, value):
    con = sqlite3.connect(_db_path())
    con.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, str(value)))
    con.commit(); con.close()

def load_settings():
    import json as _json
    global SOCIETY_NAME, SOCIETY_ADDRESS, MONTHLY_FEE, SOCIETY_START_FY, EXPENDITURE_CATEGORIES, UNIT_LABEL, OWNER_LABEL
    SOCIETY_NAME    = db_get_setting("society_name",    SOCIETY_NAME)
    SOCIETY_ADDRESS = db_get_setting("society_address", SOCIETY_ADDRESS)
    try: MONTHLY_FEE = float(db_get_setting("monthly_fee", str(MONTHLY_FEE)))
    except Exception: pass
    try: SOCIETY_START_FY = int(db_get_setting("society_start_fy", str(SOCIETY_START_FY)))
    except Exception: pass
    try:
        cats = _json.loads(db_get_setting("exp_categories", "[]"))
        if cats: EXPENDITURE_CATEGORIES = cats
    except Exception: pass
    UNIT_LABEL = db_get_setting("unit_label", "Flat")
    OWNER_LABEL = "Member" if UNIT_LABEL == "Member" else ("Tenant" if UNIT_LABEL in ("Shop","Office") else "Owner")


    global RESIDENTS
    con = sqlite3.connect(_db_path()); cur = con.cursor()
    cur.execute("SELECT flat_no, name, mobile FROM flats ORDER BY sort_order, flat_no")
    rows = cur.fetchall(); con.close()
    RESIDENTS.clear()
    for flat_no, name, mobile in rows:
        RESIDENTS[flat_no] = {"name": name, "mobile": mobile or ""}

def load_residents():
    global RESIDENTS
    con = sqlite3.connect(_db_path()); cur = con.cursor()
    cur.execute("SELECT flat_no, name, mobile FROM flats ORDER BY sort_order, flat_no")
    rows = cur.fetchall(); con.close()
    RESIDENTS.clear()
    for flat_no, name, mobile in rows:
        RESIDENTS[flat_no] = {"name": name, "mobile": mobile or ""}

def db_get_flats():
    con = sqlite3.connect(_db_path()); cur = con.cursor()
    cur.execute("SELECT flat_no, name, mobile, sort_order FROM flats ORDER BY sort_order, flat_no")
    rows = cur.fetchall(); con.close(); return rows

def db_add_flat(flat_no, name, mobile, sort_order=9999):
    flat_no = flat_no.strip().upper()
    try:
        con = sqlite3.connect(_db_path())
        con.execute("INSERT INTO flats (flat_no, name, mobile, sort_order) VALUES (?,?,?,?)",
                    (flat_no, name.strip().upper(), mobile.strip(), sort_order))
        con.commit(); con.close(); load_residents(); return True
    except sqlite3.IntegrityError:
        messagebox.showerror("Duplicate", f"Flat '{flat_no}' already exists."); return False
    except Exception as exc:
        messagebox.showerror("Error", str(exc)); return False

def db_update_flat(flat_no, name, mobile):
    con = sqlite3.connect(_db_path())
    con.execute("UPDATE flats SET name=?, mobile=? WHERE flat_no=?",
                (name.strip().upper(), mobile.strip(), flat_no))
    con.commit(); con.close(); load_residents()

def db_delete_flat(flat_no):
    con = sqlite3.connect(_db_path()); cur = con.cursor()
    cur.execute("SELECT COUNT(*) FROM payments WHERE flat_no=?", (flat_no,))
    count = cur.fetchone()[0]; con.close()
    if count > 0:
        messagebox.showerror("Cannot Delete",
                             f"Flat {flat_no} has {count} payment record(s).\nDelete those first.")
        return False
    con = sqlite3.connect(_db_path())
    con.execute("DELETE FROM flats WHERE flat_no=?", (flat_no,))
    con.commit(); con.close(); load_residents(); return True

def db_reorder_flat(flat_no, direction):
    rows = db_get_flats()
    ids = [r[0] for r in rows]
    if flat_no not in ids: return
    idx = ids.index(flat_no)
    new_idx = idx - 1 if direction == "up" else idx + 1
    if new_idx < 0 or new_idx >= len(ids): return
    ids[idx], ids[new_idx] = ids[new_idx], ids[idx]
    con = sqlite3.connect(_db_path())
    for i, fno in enumerate(ids):
        con.execute("UPDATE flats SET sort_order=? WHERE flat_no=?", (i, fno))
    con.commit(); con.close(); load_residents()

def db_get_expenditure_accounts():
    con = sqlite3.connect(_db_path()); cur = con.cursor()
    cur.execute("SELECT id, name FROM expenditure_accounts ORDER BY id")
    rows = cur.fetchall(); con.close(); return rows

def db_add_expenditure_account(name):
    try:
        con = sqlite3.connect(_db_path())
        con.execute("INSERT INTO expenditure_accounts (name) VALUES (?)", (name,))
        con.commit(); con.close(); return True
    except sqlite3.IntegrityError:
        messagebox.showerror("Duplicate", f"Account '{name}' already exists."); return False
    except Exception as exc:
        messagebox.showerror("Error", str(exc)); return False

def db_delete_expenditure_account(aid):
    con = sqlite3.connect(_db_path()); cur = con.cursor()
    cur.execute("SELECT COUNT(*) FROM expenditures WHERE account_id=?", (aid,))
    count = cur.fetchone()[0]; con.close()
    if count > 0:
        messagebox.showerror("Cannot Delete", f"This account has {count} expenditure record(s).\nDelete those first.")
        return False
    con = sqlite3.connect(_db_path())
    con.execute("DELETE FROM expenditure_accounts WHERE id=?", (aid,))
    con.commit(); con.close(); return True

def _build_receipt_no(month_from, month_to, year_from):
    con = sqlite3.connect(_db_path()); cur = con.cursor()
    cur.execute("SELECT COALESCE(MAX(id),0) FROM payments")
    seq = cur.fetchone()[0] + 1; con.close()
    yr = str(year_from)
    if month_from and month_from in MONTHS:
        mf_tag = month_from[:3].upper()
        if month_to and month_to in MONTHS and month_to != month_from:
            prefix = f"{mf_tag}-{month_to[:3].upper()}-{yr}"
        else:
            prefix = f"{mf_tag}-{yr}"
    else:
        prefix = f"GEN-{yr}"
    return f"{prefix}-{seq:04d}"

def db_save_payment(data):
    try:
        con = sqlite3.connect(_db_path())
        con.execute("""INSERT INTO payments
            (receipt_no,date,year_from,year_to,flat_no,owner_name,amount,month_from,month_to,monthly_fee,late_fee)
            VALUES (:receipt_no,:date,:year_from,:year_to,:flat_no,:owner_name,:amount,:month_from,:month_to,:monthly_fee,:late_fee)""", data)
        con.commit(); con.close(); return True
    except sqlite3.IntegrityError:
        messagebox.showerror("Duplicate Receipt", "Receipt number already exists."); return False
    except Exception as exc:
        messagebox.showerror("Database Error", str(exc)); return False

def db_fetch_payments(flat_filter=None):
    con = sqlite3.connect(_db_path()); cur = con.cursor()
    q = ("SELECT id,receipt_no,date,year_from,year_to,flat_no,owner_name,"
         "amount,month_from,month_to,COALESCE(monthly_fee,0) FROM payments")
    if flat_filter: cur.execute(q + " WHERE flat_no=? ORDER BY id", (flat_filter,))
    else: cur.execute(q + " ORDER BY id DESC")
    rows = cur.fetchall(); con.close(); return rows

def db_fetch_payments_filtered(flat_filter=None, month_filter=None, year_filter=None):
    con = sqlite3.connect(_db_path()); cur = con.cursor()
    q = ("SELECT id,receipt_no,date,year_from,year_to,flat_no,owner_name,"
         "amount,month_from,month_to,COALESCE(monthly_fee,0) FROM payments WHERE 1=1")
    params = []
    if flat_filter: q += " AND flat_no=?"; params.append(flat_filter)
    if month_filter: q += " AND (month_from=? OR month_to=?)"; params += [month_filter]*2
    if year_filter: q += " AND year_from=?"; params.append(year_filter)
    q += " ORDER BY id DESC"; cur.execute(q, params)
    rows = cur.fetchall(); con.close(); return rows

def db_delete_payment(pid):
    con = sqlite3.connect(_db_path())
    con.execute("DELETE FROM payments WHERE id=?", (pid,))
    con.commit(); con.close()

def db_update_payment(pid, data):
    try:
        con = sqlite3.connect(_db_path())
        con.execute("""UPDATE payments SET date=:date,year_from=:year_from,year_to=:year_to,
            flat_no=:flat_no,owner_name=:owner_name,amount=:amount,month_from=:month_from,
            month_to=:month_to,monthly_fee=:monthly_fee,late_fee=:late_fee WHERE id=:id""",
            {**data, "id": pid})
        con.commit(); con.close(); return True
    except Exception as exc:
        messagebox.showerror("Update Error", str(exc)); return False

def _build_admission_receipt_no():
    con = sqlite3.connect(_db_path()); cur = con.cursor()
    cur.execute("SELECT COALESCE(MAX(id),0) FROM admission_fees")
    seq = cur.fetchone()[0] + 1; con.close()
    return f"ADM-{seq:04d}"

def db_save_admission_fee(data):
    try:
        con = sqlite3.connect(_db_path())
        con.execute("""INSERT INTO admission_fees
            (receipt_no,date,flat_no,owner_name,amount,mobile,paid)
            VALUES (:receipt_no,:date,:flat_no,:owner_name,:amount,:mobile,:paid)""", data)
        con.commit(); con.close(); return True
    except Exception as exc:
        messagebox.showerror("Error", str(exc)); return False

def db_fetch_admission_fees(flat_filter=None):
    con = sqlite3.connect(_db_path()); cur = con.cursor()
    q = "SELECT id,receipt_no,date,flat_no,owner_name,amount,mobile,paid FROM admission_fees"
    if flat_filter: cur.execute(q + " WHERE flat_no=? ORDER BY id DESC", (flat_filter,))
    else: cur.execute(q + " ORDER BY id DESC")
    rows = cur.fetchall(); con.close(); return rows

def db_delete_admission_fee(aid):
    con = sqlite3.connect(_db_path())
    con.execute("DELETE FROM admission_fees WHERE id=?", (aid,))
    con.commit(); con.close()

def db_fetch_expenditures(account_id=None):
    con = sqlite3.connect(_db_path()); cur = con.cursor()
    if account_id is not None:
        cur.execute("SELECT id,date,description,amount,COALESCE(category,'Others/Misc Expense') "
                    "FROM expenditures WHERE account_id=? ORDER BY date, id", (account_id,))
    else:
        cur.execute("SELECT id,date,description,amount,COALESCE(category,'Others/Misc Expense') "
                    "FROM expenditures ORDER BY date, id")
    rows = cur.fetchall(); con.close(); return rows

def db_save_expenditure(date, desc, amount, category="Others/Misc Expense", account_id=1):
    con = sqlite3.connect(_db_path())
    con.execute("INSERT INTO expenditures (date,description,amount,category,account_id) VALUES (?,?,?,?,?)",
                (date, desc, amount, category, account_id))
    con.commit(); con.close()

def db_delete_expenditure(eid):
    con = sqlite3.connect(_db_path())
    con.execute("DELETE FROM expenditures WHERE id=?", (eid,))
    con.commit(); con.close()

def db_delete_entire_ledger():
    con = sqlite3.connect(_db_path())
    con.execute("DELETE FROM payments"); con.execute("DELETE FROM expenditures")
    con.execute("DELETE FROM manual_journals")
    con.executescript("""
        DELETE FROM sqlite_sequence WHERE name='payments';
        DELETE FROM sqlite_sequence WHERE name='expenditures';
        DELETE FROM sqlite_sequence WHERE name='manual_journals';""")
    con.commit(); con.close()

def db_save_manual_journal(date, account_type, account_id, entry_type, amount, narration, fy_year):
    con = sqlite3.connect(_db_path())
    con.execute("INSERT INTO manual_journals (date,account_type,account_id,entry_type,amount,narration,fy_year) "
                "VALUES (?,?,?,?,?,?,?)", (date, account_type, account_id, entry_type, amount, narration, fy_year))
    con.commit(); con.close()

def db_delete_manual_journal(jid):
    con = sqlite3.connect(_db_path())
    con.execute("DELETE FROM manual_journals WHERE id=?", (jid,))
    con.commit(); con.close()

def _fetch_manual_journals(account_type, account_id, fy_year):
    con = sqlite3.connect(_db_path()); cur = con.cursor()
    cur.execute("SELECT id,date,entry_type,amount,narration FROM manual_journals "
                "WHERE account_type=? AND account_id=? AND fy_year=? ORDER BY id",
                (account_type, str(account_id), fy_year))
    rows = cur.fetchall(); con.close(); return rows

def calc_months_from_amount(amount, monthly_fee):
    if monthly_fee <= 0: return 1
    return max(1, int(amount // monthly_fee))

def advance_fy_months(start_month, n):
    if start_month not in FY_MONTHS: return start_month
    idx = FY_MONTHS.index(start_month)
    return FY_MONTHS[min(idx + n - 1, 11)]

def months_in_range(mf, mt):
    if mf not in FY_MONTHS: return 1
    if not mt or mt not in FY_MONTHS: return 1
    i1 = FY_MONTHS.index(mf); i2 = FY_MONTHS.index(mt)
    return max(1, i2 - i1 + 1)

def get_current_fy():
    today = datetime.date.today()
    return today.year if today.month >= 4 else today.year - 1

def check_fy_overflow(start_month, n_months):
    if start_month not in FY_MONTHS or n_months <= 0:
        return n_months, 0, False, start_month, ""
    idx = FY_MONTHS.index(start_month); available = 12 - idx
    if n_months <= available:
        return n_months, 0, False, FY_MONTHS[idx + n_months - 1], ""
    this_fy_count = available; next_fy_count = n_months - available
    return this_fy_count, next_fy_count, True, "March", FY_MONTHS[next_fy_count - 1]

def build_split_records(base, this_fy_count, next_fy_count, this_fy_end, next_fy_end):
    fee = float(base.get("monthly_fee", 0) or 0); total_n = this_fy_count + next_fy_count
    amt_a = fee * this_fy_count if fee > 0 else round(base["amount"] * this_fy_count / total_n, 2)
    amt_b = fee * next_fy_count if fee > 0 else round(base["amount"] * next_fy_count / total_n, 2)
    yf_a = base["year_from"]; yt_a = base["year_to"]
    yf_b = yf_a + 1; yt_b = yt_a + 1
    rno_a = _build_receipt_no(base["month_from"], this_fy_end, yf_a)
    rno_b = _build_receipt_no("April", next_fy_end, yf_b)
    rec_a = {**base, "receipt_no": rno_a, "year_from": yf_a, "year_to": yt_a,
             "month_from": base["month_from"], "month_to": this_fy_end, "amount": amt_a, "monthly_fee": fee}
    rec_b = {**base, "receipt_no": rno_b, "year_from": yf_b, "year_to": yt_b,
             "month_from": "April", "month_to": next_fy_end, "amount": amt_b, "monthly_fee": fee}
    return rec_a, rec_b

def get_arrears_for_flat(flat_no, monthly_fee=MONTHLY_FEE):
    today = datetime.date.today(); current_fy = get_current_fy()
    unpaid_list = []; by_fy = {}
    for fy in range(SOCIETY_START_FY, current_fy + 1):
        matrix = get_payment_matrix_with_fees(fy)
        if flat_no not in matrix: continue
        fy_unpaid = []
        for month in FY_MONTHS:
            cal_year = fy if month in _FY_APR_DEC else fy + 1
            month_num = MONTHS.index(month) + 1
            if datetime.date(cal_year, month_num, 1) >= datetime.date(today.year, today.month, 1): continue
            if matrix[flat_no][month] == 0:
                fy_unpaid.append(month); unpaid_list.append((fy, month))
        if fy_unpaid: by_fy[fy] = fy_unpaid
    return {"count": len(unpaid_list), "total_owed": len(unpaid_list) * monthly_fee,
            "by_fy": by_fy, "unpaid_list": unpaid_list}

def get_all_arrears(monthly_fee=MONTHLY_FEE):
    return {fid: get_arrears_for_flat(fid, monthly_fee) for fid in RESIDENTS}

def get_payment_matrix(fy_year):
    matrix = {fid: {m: False for m in FY_MONTHS} for fid in RESIDENTS}
    for row in db_fetch_payments():
        _, rno, date, year_from, year_to, flat_no, owner, amt, mf, mt, fee = row
        if flat_no not in matrix: continue
        if mf and mf in FY_MONTHS:
            if year_from != fy_year: continue
            mf_idx = FY_MONTHS.index(mf)
            mt_idx = FY_MONTHS.index(mt) if (mt and mt in FY_MONTHS) else mf_idx
            for idx in range(mf_idx, mt_idx + 1): matrix[flat_no][FY_MONTHS[idx]] = True
        else:
            try:
                parts = date.split("/"); p_cal_month = int(parts[1]); p_cal_year = int(parts[2])
                month_name = MONTHS[p_cal_month - 1]
                if month_name not in FY_MONTHS: continue
                cal_fy = p_cal_year if month_name in _FY_APR_DEC else p_cal_year - 1
                if cal_fy == fy_year: matrix[flat_no][month_name] = True
            except (IndexError, ValueError): pass
    return matrix

def get_payment_matrix_with_fees(fy_year):
    matrix = {fid: {m: 0.0 for m in FY_MONTHS} for fid in RESIDENTS}
    for row in db_fetch_payments():
        _, rno, date, year_from, year_to, flat_no, owner, amt, mf, mt, fee = row
        if flat_no not in matrix: continue
        if mf and mf in FY_MONTHS:
            if year_from != fy_year: continue
            mf_idx = FY_MONTHS.index(mf)
            mt_idx = FY_MONTHS.index(mt) if (mt and mt in FY_MONTHS) else mf_idx
            n = mt_idx - mf_idx + 1; per_m = fee if fee > 0 else (amt / n if n > 0 else amt)
            for idx in range(mf_idx, mt_idx + 1): matrix[flat_no][FY_MONTHS[idx]] = per_m
        else:
            try:
                parts = date.split("/"); p_cal_month = int(parts[1]); p_cal_year = int(parts[2])
                month_name = MONTHS[p_cal_month - 1]
                if month_name not in FY_MONTHS: continue
                cal_fy = p_cal_year if month_name in _FY_APR_DEC else p_cal_year - 1
                if cal_fy == fy_year: matrix[flat_no][month_name] = fee if fee > 0 else amt
            except (IndexError, ValueError): pass
    return matrix

def _build_flat_acct(flat_no, fy_year, opening_balance):
    entries = []; balance = opening_balance; total_dr = 0.0; total_cr = 0.0
    month_payments = defaultdict(list); unlinked_cr = []
    for row in db_fetch_payments(flat_filter=flat_no):
        _, rno, date, yf, yt, flat, owner, amt, mf, mt, fee = row
        if yf != fy_year: continue
        if mf and mf in FY_MONTHS:
            mf_idx = FY_MONTHS.index(mf)
            mt_idx = FY_MONTHS.index(mt) if (mt and mt in FY_MONTHS) else mf_idx
            n = mt_idx - mf_idx + 1; per_m = fee if fee > 0 else (amt / n if n > 0 else amt)
            for idx in range(mf_idx, mt_idx + 1): month_payments[FY_MONTHS[idx]].append((date, rno, per_m, row[0]))
        else: unlinked_cr.append((date, rno, amt, row[0]))
    for month in FY_MONTHS:
        cal_year = fy_year if month in _FY_APR_DEC else fy_year + 1
        for pay_date, rno, per_m, pid in month_payments[month]:
            balance -= per_m; total_cr += per_m
            entries.append({"date": pay_date, "narration": f"Maintenance — {month} {cal_year}",
                            "vch_type": "Receipt", "vch_no": rno, "dr_amt": 0, "cr_amt": per_m,
                            "balance": balance, "type": "Cr", "source": "payment", "source_id": str(pid)})
    for pay_date, rno, amt, pid in unlinked_cr:
        balance -= amt; total_cr += amt
        entries.append({"date": pay_date, "narration": "Maintenance payment", "vch_type": "Receipt",
                        "vch_no": rno, "dr_amt": 0, "cr_amt": amt, "balance": balance,
                        "type": "Cr", "source": "payment", "source_id": str(pid)})
    for jid, jdate, jtype, jamt, jnarr in _fetch_manual_journals("flat", flat_no, fy_year):
        if jtype == "Dr":
            balance += jamt; total_dr += jamt
            entries.append({"date": jdate, "narration": jnarr, "vch_type": "Journal",
                            "vch_no": f"J-{jid}", "dr_amt": jamt, "cr_amt": 0, "balance": balance,
                            "type": "Dr", "source": "manual", "source_id": str(jid)})
        else:
            balance -= jamt; total_cr += jamt
            entries.append({"date": jdate, "narration": jnarr, "vch_type": "Journal",
                            "vch_no": f"J-{jid}", "dr_amt": 0, "cr_amt": jamt, "balance": balance,
                            "type": "Cr", "source": "manual", "source_id": str(jid)})
    def _sort_key(e):
        try:
            p = e["date"].split("/")
            return datetime.date(int(p[2]), int(p[1]), int(p[0]))
        except Exception: return datetime.date(1900, 1, 1)
    entries.sort(key=_sort_key)
    balance = opening_balance; total_dr = 0.0; total_cr = 0.0
    for e in entries:
        if e["type"] == "Dr": balance += e["dr_amt"]; total_dr += e["dr_amt"]
        else: balance -= e["cr_amt"]; total_cr += e["cr_amt"]
        e["balance"] = balance
    return {"fy_year": fy_year, "opening_balance": opening_balance, "entries": entries,
            "total_dr": total_dr, "total_cr": total_cr, "closing_balance": balance}

def get_flat_account_entries(flat_no, fy_year):
    if fy_year <= SOCIETY_START_FY: opening = 0.0
    else:
        prev = get_flat_account_entries(flat_no, fy_year - 1)
        opening = prev["closing_balance"]
    return _build_flat_acct(flat_no, fy_year, opening)

def get_expenditure_account_entries(fy_year, account_id=1):
    if fy_year <= SOCIETY_START_FY: opening = 0.0
    else:
        prev = get_expenditure_account_entries(fy_year - 1, account_id)
        opening = prev["closing_balance"]
    fy_start = datetime.date(fy_year, 4, 1); fy_end = datetime.date(fy_year + 1, 3, 31)
    entries = []; balance = opening; total_dr = 0.0; total_cr = 0.0
    for eid, date, desc, amt, cat in db_fetch_expenditures(account_id=account_id):
        try:
            parts = date.split("/")
            exp_date = datetime.date(int(parts[2]), int(parts[1]), int(parts[0]))
        except Exception: continue
        if not (fy_start <= exp_date <= fy_end): continue
        balance += amt; total_dr += amt
        entries.append({"date": date, "narration": desc, "vch_type": cat, "vch_no": f"EXP-{eid}",
                        "dr_amt": amt, "cr_amt": 0, "balance": balance, "type": "Dr",
                        "source": "expenditure", "source_id": str(eid)})
    for jid, jdate, jtype, jamt, jnarr in _fetch_manual_journals("expenditure", str(account_id), fy_year):
        if jtype == "Dr":
            balance += jamt; total_dr += jamt
            entries.append({"date": jdate, "narration": jnarr, "vch_type": "Journal",
                            "vch_no": f"J-{jid}", "dr_amt": jamt, "cr_amt": 0, "balance": balance,
                            "type": "Dr", "source": "manual", "source_id": str(jid)})
        else:
            balance -= jamt; total_cr += jamt
            entries.append({"date": jdate, "narration": jnarr, "vch_type": "Journal",
                            "vch_no": f"J-{jid}", "dr_amt": 0, "cr_amt": jamt, "balance": balance,
                            "type": "Cr", "source": "manual", "source_id": str(jid)})
    return {"fy_year": fy_year, "opening_balance": opening, "entries": entries,
            "total_dr": total_dr, "total_cr": total_cr, "closing_balance": balance}

def build_receipt(d):
    line = "-" * 30; mf = d.get("month_from", ""); mt = d.get("month_to", "")
    fee = float(d.get("monthly_fee", 0) or 0); amt = float(d.get("amount", 0))
    late = float(d.get("late_fee", 0) or 0)
    period_line = breakdown_line = late_line = ""
    if mf:
        n_months = months_in_range(mf, mt) if (mt and mt != mf) else 1
        period_line = f"\nPeriod      : {mf}" + (f" - {mt}" if mt and mt != mf else "")
        if fee > 0:
            breakdown_line = (f"\nBreakdown   : {n_months} month{'s' if n_months > 1 else ''}"
                              f" x Rs.{fee:,.0f} = Rs.{fee * n_months:,.0f}")
    if late > 0:
        late_line = f"\nLate Fee    : Rs.{late:,.0f}"
    arrears_note = ""
    current_fy = get_current_fy()
    if d.get("year_from", current_fy) < current_fy:
        arrears_note = f"\n[ARREARS SETTLEMENT — {fy_label(d['year_from'])}]"
    return (f"*{SOCIETY_NAME}*\n{line}\n       *PAYMENT RECEIPT*\n{line}\n"
            f"Receipt No  : {d['receipt_no']}\nDate        : {d['date']}\n"
            f"Fin. Year   : {d['year_from']} - {d['year_to']}{arrears_note}\n{line}\n"
            f"{UNIT_LABEL} No     : *{d['flat_no']}*\n{OWNER_LABEL}       : {d['owner_name']}\n{line}\n"
            f"Amount Paid : *Rs. {amt:,.0f}*{period_line}{breakdown_line}{late_line}\n{line}\n"
            f"*Thank you for your payment!*")

def build_reminder(flat_no, owner_name, month, year):
    first_name = owner_name.split()[0].capitalize(); line = "-" * 30
    return (f"*{SOCIETY_NAME}*\n{line}\n   *MAINTENANCE REMINDER*\n{line}\n"
            f"{UNIT_LABEL} No  : *{flat_no}*\nDear {first_name},\n\n"
            f"Your maintenance payment for\n*{month} {year}* is still pending.\n\n"
            f"Kindly pay at the earliest convenience.\n{line}\nThank you.")

def generate_receipt_pdf(d, filepath):
    if not PDF_AVAILABLE:
        messagebox.showerror("PDF Error", "reportlab not installed.\npip install reportlab"); return False
    try:
        doc = SimpleDocTemplate(filepath, pagesize=A4,
                                leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
        story = []; W = 17*cm
        hdr = Table([[Paragraph(f"<b>{SOCIETY_NAME}</b>",
                               ParagraphStyle("H", fontSize=16, fontName="Helvetica-Bold",
                                              textColor=colors.white, alignment=TA_CENTER))],
                     [Paragraph(SOCIETY_ADDRESS,
                               ParagraphStyle("HS", fontSize=9, alignment=TA_CENTER,
                                              textColor=colors.HexColor("#AED6F1")))]], colWidths=[W])
        hdr.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#1E3A5F")),
                                 ("ALIGN",(0,0),(-1,-1),"CENTER"),
                                 ("TOPPADDING",(0,0),(0,0),14), ("BOTTOMPADDING",(0,1),(0,1),14)]))
        story.append(hdr); story.append(Spacer(1, 8))
        current_fy = get_current_fy(); is_arrears = d.get("year_from", current_fy) < current_fy
        badge_text = "ARREARS SETTLEMENT" if is_arrears else "PAYMENT RECEIPT"
        badge = Table([[Paragraph(badge_text, ParagraphStyle("Badge", fontSize=14,
                                                              fontName="Helvetica-Bold",
                                                              textColor=colors.HexColor("#1E3A5F"),
                                                              alignment=TA_CENTER))]], colWidths=[W])
        badge.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#EBF5FB")),
                                   ("TOPPADDING",(0,0),(-1,-1),8), ("BOTTOMPADDING",(0,0),(-1,-1),8),
                                   ("BOX",(0,0),(-1,-1),1,colors.HexColor("#1A5276"))]))
        story.append(badge); story.append(Spacer(1, 14))
        amt = float(d.get("amount", 0))
        flat_tbl = Table([[Paragraph(f"{UNIT_LABEL.upper()} DETAILS", ParagraphStyle("FH", fontSize=9,
                                                                     fontName="Helvetica-Bold", textColor=colors.white))],
                          [Paragraph(f"Flat No :  <b>{d['flat_no']}</b>",
                                    ParagraphStyle("FV", fontSize=12, textColor=colors.HexColor("#1E3A5F")))],
                          [Paragraph(f"{OWNER_LABEL}   :  <b>{d['owner_name']}</b>",
                                    ParagraphStyle("FV2", fontSize=12, textColor=colors.HexColor("#1E3A5F")))]], colWidths=[W])
        flat_tbl.setStyle(TableStyle([("BACKGROUND",(0,0),(0,0),colors.HexColor("#1A5276")),
                                      ("BACKGROUND",(0,1),(0,-1),colors.HexColor("#EBF5FB")),
                                      ("TOPPADDING",(0,0),(-1,-1),6), ("BOTTOMPADDING",(0,0),(-1,-1),6),
                                      ("LEFTPADDING",(0,0),(-1,-1),12),
                                      ("BOX",(0,0),(-1,-1),0.8,colors.HexColor("#1A5276"))]))
        story.append(flat_tbl); story.append(Spacer(1, 14))
        amt_tbl = Table([[Paragraph("AMOUNT PAID", ParagraphStyle("AL", fontSize=9,
                                                                   textColor=colors.HexColor("#7F8C8D"), alignment=TA_CENTER))],
                         [Paragraph(f"Rs. {amt:,.0f}", ParagraphStyle("AV", fontSize=22,
                                                                       fontName="Helvetica-Bold",
                                                                       textColor=colors.HexColor("#1E8449"), alignment=TA_CENTER))]], colWidths=[W])
        amt_tbl.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#EAFAF1")),
                                     ("TOPPADDING",(0,0),(-1,-1),8), ("BOTTOMPADDING",(0,0),(-1,-1),8),
                                     ("BOX",(0,0),(-1,-1),1,colors.HexColor("#1E8449"))]))
        story.append(amt_tbl); story.append(Spacer(1, 20))
        story.append(Paragraph("Thank you for your payment!",
                               ParagraphStyle("TY", fontSize=11, fontName="Helvetica-Bold",
                                              textColor=colors.HexColor("#1E8449"), alignment=TA_CENTER)))
        doc.build(story); return True
    except Exception as exc:
        messagebox.showerror("PDF Error", str(exc)); return False

def generate_admission_fee_pdf(data, filepath):
    if not PDF_AVAILABLE:
        messagebox.showerror("PDF Error", "reportlab not installed.\npip install reportlab"); return False
    try:
        doc = SimpleDocTemplate(filepath, pagesize=A4,
                                leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
        story = []; W = 17*cm
        hdr = Table([[Paragraph(f"<b>{SOCIETY_NAME}</b>",
                               ParagraphStyle("H", fontSize=16, fontName="Helvetica-Bold",
                                              textColor=colors.white, alignment=TA_CENTER))],
                     [Paragraph(SOCIETY_ADDRESS or "&nbsp;",
                               ParagraphStyle("HS", fontSize=9, alignment=TA_CENTER,
                                              textColor=colors.HexColor("#AED6F1")))]], colWidths=[W])
        hdr.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#1E3A5F")),
                                 ("TOPPADDING",(0,0),(0,0),14),("BOTTOMPADDING",(0,1),(0,1),14)]))
        story.append(hdr); story.append(Spacer(1, 8))
        badge = Table([[Paragraph("ADMISSION FEE RECEIPT",
                                  ParagraphStyle("B", fontSize=14, fontName="Helvetica-Bold",
                                                 textColor=colors.HexColor("#1E3A5F"), alignment=TA_CENTER))]],
                      colWidths=[W])
        badge.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#EBF5FB")),
                                   ("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8),
                                   ("BOX",(0,0),(-1,-1),1,colors.HexColor("#1A5276"))]))
        story.append(badge); story.append(Spacer(1, 14))
        amt = float(data.get("amount", 0))
        details = Table([
            [Paragraph("Receipt No :", ParagraphStyle("DL", fontSize=10, textColor=colors.HexColor("#7F8C8D"))),
             Paragraph(f"<b>{data['receipt_no']}</b>", ParagraphStyle("DV", fontSize=10))],
            [Paragraph("Date :", ParagraphStyle("DL2", fontSize=10, textColor=colors.HexColor("#7F8C8D"))),
             Paragraph(f"<b>{data['date']}</b>", ParagraphStyle("DV2", fontSize=10))],
            [Paragraph(f"{UNIT_LABEL} No :", ParagraphStyle("DL3", fontSize=10, textColor=colors.HexColor("#7F8C8D"))),
             Paragraph(f"<b>{data['flat_no']}</b>", ParagraphStyle("DV3", fontSize=12, fontName="Helvetica-Bold", textColor=colors.HexColor("#1E3A5F")))],
            [Paragraph(f"{OWNER_LABEL} :", ParagraphStyle("DL4", fontSize=10, textColor=colors.HexColor("#7F8C8D"))),
             Paragraph(f"<b>{data['owner_name']}</b>", ParagraphStyle("DV4", fontSize=10))],
        ], colWidths=[W*0.3, W*0.7])
        details.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#EBF5FB")),
                                     ("TOPPADDING",(0,0),(-1,-1),7),("BOTTOMPADDING",(0,0),(-1,-1),7),
                                     ("LEFTPADDING",(0,0),(0,-1),14),
                                     ("BOX",(0,0),(-1,-1),0.8,colors.HexColor("#1A5276"))]))
        story.append(details); story.append(Spacer(1, 14))
        amt_tbl = Table([[Paragraph("ADMISSION FEE PAID",
                                    ParagraphStyle("AL", fontSize=9, textColor=colors.HexColor("#7F8C8D"), alignment=TA_CENTER))],
                         [Paragraph(f"Rs. {amt:,.0f}",
                                    ParagraphStyle("AV", fontSize=24, fontName="Helvetica-Bold",
                                                   textColor=colors.HexColor("#1E8449"), alignment=TA_CENTER))]],
                        colWidths=[W])
        amt_tbl.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#EAFAF1")),
                                     ("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),10),
                                     ("BOX",(0,0),(-1,-1),1,colors.HexColor("#1E8449"))]))
        story.append(amt_tbl); story.append(Spacer(1, 20))
        story.append(Paragraph("Thank you! This receipt acknowledges your Admission Fee payment.",
                               ParagraphStyle("TY", fontSize=10, fontName="Helvetica-Bold",
                                              textColor=colors.HexColor("#1E8449"), alignment=TA_CENTER)))
        story.append(Spacer(1, 8))
        story.append(Paragraph(f"Generated on {datetime.date.today().strftime('%d %B %Y')}  |  {SOCIETY_NAME}",
                               ParagraphStyle("GN", fontSize=7, textColor=colors.HexColor("#7F8C8D"), alignment=TA_CENTER)))
        doc.build(story); return True
    except Exception as exc:
        messagebox.showerror("PDF Error", str(exc)); return False

def generate_flat_account_pdf(flat_no, fy_year, filepath):
    if not PDF_AVAILABLE:
        messagebox.showerror("PDF Error", "reportlab not installed."); return False
    try:
        acct = get_flat_account_entries(flat_no, fy_year)
        doc = SimpleDocTemplate(filepath, pagesize=A4, leftMargin=1.8*cm, rightMargin=1.8*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        W = 17.4*cm; story = []
        cover = Table([[Paragraph(f"<b>{SOCIETY_NAME}</b>",
                                  ParagraphStyle("CT", fontSize=16, fontName="Helvetica-Bold",
                                                 textColor=colors.white, alignment=TA_CENTER))],
                       [Paragraph(SOCIETY_ADDRESS, ParagraphStyle("CS", fontSize=8, alignment=TA_CENTER,
                                                                   textColor=colors.HexColor("#AED6F1")))],
                       [Paragraph("<b>MAINTENANCE ACCOUNT STATEMENT</b>",
                                  ParagraphStyle("CTi", fontSize=12, fontName="Helvetica-Bold",
                                                 textColor=colors.HexColor("#F9E79F"), alignment=TA_CENTER))],
                       [Paragraph(f"Flat {flat_no}  —  {RESIDENTS.get(flat_no, {}).get('name', flat_no)}  |  {fy_label(fy_year)}",
                                  ParagraphStyle("CTi2", fontSize=10, textColor=colors.HexColor("#AED6F1"), alignment=TA_CENTER))]], colWidths=[W])
        cover.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#1E3A5F")),
                                   ("TOPPADDING",(0,0),(0,0),14), ("BOTTOMPADDING",(0,-1),(0,-1),12),
                                   ("TOPPADDING",(0,1),(-1,-2),3)]))
        story.append(cover); story.append(Spacer(1, 12))
        th = ParagraphStyle("TH", fontSize=8, fontName="Helvetica-Bold", textColor=colors.white, alignment=TA_CENTER)
        tl = ParagraphStyle("TL", fontSize=8, alignment=TA_LEFT)
        tc = ParagraphStyle("TC", fontSize=8, alignment=TA_CENTER)
        tr = ParagraphStyle("TR", fontSize=8, alignment=TA_RIGHT)
        headers = [Paragraph(x, th) for x in ["Date","Particulars","Vch Type","Vch No.","Debit","Credit","Balance"]]
        table_rows = [headers]; cell_styles = []; row_idx = 1
        ob = acct["opening_balance"]
        ob_label = f"Rs.{abs(ob):,.2f} {'Dr' if ob > 0 else 'Cr'}" if ob != 0 else "Rs.0.00"
        table_rows.append([Paragraph("—", tc),
                           Paragraph("<b>Opening Balance  (b/f from previous year)</b>", tl),
                           Paragraph("", tc), Paragraph("", tc),
                           Paragraph(f"{ob:,.2f}" if ob > 0 else "", tr),
                           Paragraph(f"{abs(ob):,.2f}" if ob < 0 else "", tr),
                           Paragraph(ob_label, ParagraphStyle("OB", fontSize=8, fontName="Helvetica-Bold",
                                                               textColor=colors.HexColor("#1A5276"), alignment=TA_CENTER))])
        cell_styles.append(("BACKGROUND",(0,row_idx),(-1,row_idx),colors.HexColor("#EBF5FB")))
        cell_styles.append(("FONT",(0,row_idx),(-1,row_idx),"Helvetica-Bold",8)); row_idx += 1
        for entry in acct["entries"]:
            bal = entry["balance"]
            bal_label = f"Rs.{abs(bal):,.2f} {'Dr' if bal > 0 else 'Cr'}" if bal != 0 else "Rs.0.00"
            bg = colors.HexColor("#FFF5F5") if entry["type"] == "Dr" else colors.HexColor("#F0FFF4")
            if entry.get("source") == "manual": bg = colors.HexColor("#FEFAE7")
            table_rows.append([Paragraph(entry["date"], tc), Paragraph(entry["narration"], tl),
                               Paragraph(entry.get("vch_type",""), tc), Paragraph(entry.get("vch_no",""), tc),
                               Paragraph(f"{entry['dr_amt']:,.2f}" if entry["dr_amt"] else "", tr),
                               Paragraph(f"{entry['cr_amt']:,.2f}" if entry["cr_amt"] else "", tr),
                               Paragraph(bal_label, ParagraphStyle("BL", fontSize=8, textColor=colors.HexColor("#1A5276"), alignment=TA_CENTER))])
            cell_styles.append(("BACKGROUND",(0,row_idx),(-1,row_idx),bg)); row_idx += 1
        table_rows.append([Paragraph("", tc),
                           Paragraph("<b>TOTAL</b>", ParagraphStyle("TOT", fontSize=9, fontName="Helvetica-Bold", alignment=TA_LEFT)),
                           Paragraph("", tc), Paragraph("", tc),
                           Paragraph(f"<b>{acct['total_dr']:,.2f}</b>", ParagraphStyle("TD", fontSize=9, fontName="Helvetica-Bold", alignment=TA_RIGHT)),
                           Paragraph(f"<b>{acct['total_cr']:,.2f}</b>", ParagraphStyle("TC2", fontSize=9, fontName="Helvetica-Bold", alignment=TA_RIGHT)),
                           Paragraph("", tc)])
        cell_styles.append(("BACKGROUND",(0,row_idx),(-1,row_idx),colors.HexColor("#1E3A5F")))
        cell_styles.append(("TEXTCOLOR",(0,row_idx),(-1,row_idx),colors.white))
        tbl = Table(table_rows, colWidths=[2.0*cm, 6.2*cm, 2.4*cm, 2.4*cm, 1.8*cm, 1.8*cm, 2.0*cm])
        base = [("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1E3A5F")),
                ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#D5D8DC")),
                ("TOPPADDING",(0,0),(-1,-1),4), ("BOTTOMPADDING",(0,0),(-1,-1),4), ("VALIGN",(0,0),(-1,-1),"MIDDLE")]
        tbl.setStyle(TableStyle(base + cell_styles)); story.append(tbl); story.append(Spacer(1, 8))
        cb = acct["closing_balance"]; cb_color = "#C0392B" if cb > 0 else "#145A32"
        cb_label = (f"Rs.{abs(cb):,.2f} "
                    f"{'Dr (Amount due from owner)' if cb > 0 else 'Cr (Advance / credit with society)'}")
        final = Table([[Paragraph("Closing Balance  (c/f to next year)",
                                  ParagraphStyle("CBL", fontSize=10, fontName="Helvetica-Bold", textColor=colors.HexColor(cb_color))),
                        Paragraph(cb_label, ParagraphStyle("CBR", fontSize=11, fontName="Helvetica-Bold",
                                                            textColor=colors.HexColor(cb_color), alignment=TA_RIGHT))]],
                      colWidths=[W*0.5, W*0.5])
        final.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#FEF9E7" if cb > 0 else "#EAFAF1")),
                                   ("TOPPADDING",(0,0),(-1,-1),10), ("BOTTOMPADDING",(0,0),(-1,-1),10),
                                   ("LEFTPADDING",(0,0),(0,0),12), ("BOX",(0,0),(-1,-1),1.5,colors.HexColor(cb_color))]))
        story.append(final)
        story.append(Paragraph(f"Generated on {datetime.date.today().strftime('%d %B %Y')}  |  {SOCIETY_NAME}",
                               ParagraphStyle("GEN", fontSize=7, textColor=colors.HexColor("#7F8C8D"), alignment=TA_CENTER)))
        doc.build(story); return True
    except Exception as exc:
        messagebox.showerror("PDF Error", str(exc)); return False

def generate_expenditure_account_pdf(fy_year, filepath, account_id=1, account_name="Society Expenditure"):
    if not PDF_AVAILABLE:
        messagebox.showerror("PDF Error", "reportlab not installed."); return False
    try:
        acct = get_expenditure_account_entries(fy_year, account_id)
        doc = SimpleDocTemplate(filepath, pagesize=A4, leftMargin=1.8*cm, rightMargin=1.8*cm,
                                topMargin=2*cm, bottomMargin=2*cm)
        W = 17.4*cm; story = []
        cover = Table([[Paragraph(f"<b>{SOCIETY_NAME}</b>",
                                  ParagraphStyle("CT", fontSize=16, fontName="Helvetica-Bold",
                                                 textColor=colors.white, alignment=TA_CENTER))],
                       [Paragraph(SOCIETY_ADDRESS, ParagraphStyle("CS", fontSize=8, alignment=TA_CENTER,
                                                                   textColor=colors.HexColor("#AED6F1")))],
                       [Paragraph(f"<b>EXPENDITURE ACCOUNT — {account_name.upper()}</b>",
                                  ParagraphStyle("CTi", fontSize=12, fontName="Helvetica-Bold",
                                                 textColor=colors.HexColor("#F9E79F"), alignment=TA_CENTER))],
                       [Paragraph(f"{fy_label(fy_year)}  (April {fy_year} – March {fy_year+1})",
                                  ParagraphStyle("CTi2", fontSize=10, textColor=colors.HexColor("#AED6F1"), alignment=TA_CENTER))]], colWidths=[W])
        cover.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#922B21")),
                                   ("TOPPADDING",(0,0),(0,0),14), ("BOTTOMPADDING",(0,-1),(0,-1),12),
                                   ("TOPPADDING",(0,1),(-1,-2),3)]))
        story.append(cover); story.append(Spacer(1, 12))
        th = ParagraphStyle("TH", fontSize=8, fontName="Helvetica-Bold", textColor=colors.white, alignment=TA_CENTER)
        tl = ParagraphStyle("TL", fontSize=8, alignment=TA_LEFT)
        tc = ParagraphStyle("TC", fontSize=8, alignment=TA_CENTER)
        tr = ParagraphStyle("TR", fontSize=8, alignment=TA_RIGHT)
        headers = [Paragraph(x, th) for x in ["Date","Particulars","Category","Vch No.","Debit","Credit","Balance"]]
        table_rows = [headers]; cell_styles = []; row_idx = 1
        ob = acct["opening_balance"]
        ob_label = f"Rs.{abs(ob):,.2f} {'Dr' if ob > 0 else 'Cr'}" if ob != 0 else "Rs.0.00"
        table_rows.append([Paragraph("—", tc),
                           Paragraph("<b>Opening Balance  (b/f from previous year)</b>", tl),
                           Paragraph("", tc), Paragraph("", tc),
                           Paragraph(f"{ob:,.2f}" if ob > 0 else "", tr),
                           Paragraph(f"{abs(ob):,.2f}" if ob < 0 else "", tr),
                           Paragraph(ob_label, ParagraphStyle("OB", fontSize=8, fontName="Helvetica-Bold",
                                                               textColor=colors.HexColor("#922B21"), alignment=TA_CENTER))])
        cell_styles.append(("BACKGROUND",(0,row_idx),(-1,row_idx),colors.HexColor("#FADBD8")))
        cell_styles.append(("FONT",(0,row_idx),(-1,row_idx),"Helvetica-Bold",8)); row_idx += 1
        for entry in acct["entries"]:
            bal = entry["balance"]
            bal_label = f"Rs.{abs(bal):,.2f} {'Dr' if bal > 0 else 'Cr'}" if bal != 0 else "Rs.0.00"
            bg = colors.HexColor("#FFF5F5") if entry["type"] == "Dr" else colors.HexColor("#F0FFF4")
            table_rows.append([Paragraph(entry["date"], tc), Paragraph(entry["narration"], tl),
                               Paragraph(entry.get("vch_type",""), tc), Paragraph(entry.get("vch_no",""), tc),
                               Paragraph(f"{entry['dr_amt']:,.2f}" if entry["dr_amt"] else "", tr),
                               Paragraph(f"{entry['cr_amt']:,.2f}" if entry["cr_amt"] else "", tr),
                               Paragraph(bal_label, ParagraphStyle("BL", fontSize=8, textColor=colors.HexColor("#922B21"), alignment=TA_CENTER))])
            cell_styles.append(("BACKGROUND",(0,row_idx),(-1,row_idx),bg)); row_idx += 1
        table_rows.append([Paragraph("", tc),
                           Paragraph("<b>TOTAL</b>", ParagraphStyle("TOT", fontSize=9, fontName="Helvetica-Bold", alignment=TA_LEFT)),
                           Paragraph("", tc), Paragraph("", tc),
                           Paragraph(f"<b>{acct['total_dr']:,.2f}</b>", ParagraphStyle("TD", fontSize=9, fontName="Helvetica-Bold", alignment=TA_RIGHT)),
                           Paragraph(f"<b>{acct['total_cr']:,.2f}</b>", ParagraphStyle("TC2", fontSize=9, fontName="Helvetica-Bold", alignment=TA_RIGHT)),
                           Paragraph("", tc)])
        cell_styles.append(("BACKGROUND",(0,row_idx),(-1,row_idx),colors.HexColor("#922B21")))
        cell_styles.append(("TEXTCOLOR",(0,row_idx),(-1,row_idx),colors.white))
        tbl = Table(table_rows, colWidths=[2.0*cm, 5.8*cm, 2.6*cm, 2.2*cm, 1.8*cm, 1.8*cm, 2.0*cm])
        base = [("BACKGROUND",(0,0),(-1,0),colors.HexColor("#922B21")),
                ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#D5D8DC")),
                ("TOPPADDING",(0,0),(-1,-1),4), ("BOTTOMPADDING",(0,0),(-1,-1),4), ("VALIGN",(0,0),(-1,-1),"MIDDLE")]
        tbl.setStyle(TableStyle(base + cell_styles)); story.append(tbl); story.append(Spacer(1, 8))
        cb = acct["closing_balance"]; cb_color = "#C0392B" if cb > 0 else "#145A32"
        cb_label = f"Rs.{abs(cb):,.2f} {'Dr (Net expenditure)' if cb > 0 else 'Cr (Net surplus)'}"
        final = Table([[Paragraph("Closing Balance  (c/f to next year)",
                                  ParagraphStyle("CBL", fontSize=10, fontName="Helvetica-Bold", textColor=colors.HexColor(cb_color))),
                        Paragraph(cb_label, ParagraphStyle("CBR", fontSize=11, fontName="Helvetica-Bold",
                                                            textColor=colors.HexColor(cb_color), alignment=TA_RIGHT))]],
                      colWidths=[W*0.5, W*0.5])
        final.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#FADBD8" if cb > 0 else "#EAFAF1")),
                                   ("TOPPADDING",(0,0),(-1,-1),10), ("BOTTOMPADDING",(0,0),(-1,-1),10),
                                   ("LEFTPADDING",(0,0),(0,0),12), ("BOX",(0,0),(-1,-1),1.5,colors.HexColor(cb_color))]))
        story.append(final); doc.build(story); return True
    except Exception as exc:
        messagebox.showerror("PDF Error", str(exc)); return False


# ─────────────────────────── EXCEL EXPORT HELPERS ────────────────────────────

def _xl_styles():
    """Return a dict of openpyxl style objects for reuse across sheets."""
    thin  = Side(style="thin",   color="D5D8DC")
    thick = Side(style="medium", color="1A5276")
    def border(t=None, b=None, l=None, r=None):
        return Border(top=t or thin, bottom=b or thin, left=l or thin, right=r or thin)

    return {
        # fills
        "fill_hdr_blue":  PatternFill("solid", fgColor="1E3A5F"),
        "fill_hdr_red":   PatternFill("solid", fgColor="922B21"),
        "fill_ob":        PatternFill("solid", fgColor="EBF5FB"),
        "fill_ob_red":    PatternFill("solid", fgColor="FADBD8"),
        "fill_cr":        PatternFill("solid", fgColor="F0FFF4"),
        "fill_dr":        PatternFill("solid", fgColor="FFF5F5"),
        "fill_manual":    PatternFill("solid", fgColor="FEFAE7"),
        "fill_total_blue":PatternFill("solid", fgColor="1E3A5F"),
        "fill_total_red": PatternFill("solid", fgColor="922B21"),
        "fill_cb_due":    PatternFill("solid", fgColor="FEF9E7"),
        "fill_cb_adv":    PatternFill("solid", fgColor="EAFAF1"),
        "fill_sum_hdr":   PatternFill("solid", fgColor="1E3A5F"),
        "fill_sub_blue":  PatternFill("solid", fgColor="D6EAF8"),
        "fill_sub_red":   PatternFill("solid", fgColor="FADBD8"),
        "fill_alt":       PatternFill("solid", fgColor="F2F4F8"),
        # fonts
        "font_hdr":   Font(name="Arial", bold=True, color="FFFFFF", size=10),
        "font_title": Font(name="Arial", bold=True, color="FFFFFF", size=12),
        "font_body":  Font(name="Arial", size=9),
        "font_bold":  Font(name="Arial", bold=True, size=9),
        "font_cb_due":Font(name="Arial", bold=True, color="C0392B", size=10),
        "font_cb_adv":Font(name="Arial", bold=True, color="145A32", size=10),
        "font_total": Font(name="Arial", bold=True, color="FFFFFF", size=10),
        "font_ob":    Font(name="Arial", bold=True, color="1A5276", size=9),
        # alignment
        "al_ctr": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "al_lft": Alignment(horizontal="left",   vertical="center", wrap_text=True),
        "al_rgt": Alignment(horizontal="right",  vertical="center"),
        # border
        "border_thin": border(),
    }


def _xl_write_ledger_sheet(ws, acct, title_label, style_theme="blue"):
    """Write a full Tally-style ledger into an openpyxl worksheet."""
    S = _xl_styles()
    hdr_fill  = S["fill_hdr_blue"]  if style_theme == "blue" else S["fill_hdr_red"]
    tot_fill  = S["fill_total_blue"] if style_theme == "blue" else S["fill_total_red"]
    ob_fill   = S["fill_ob"]        if style_theme == "blue" else S["fill_ob_red"]

    # ── Title rows ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    ws["A1"] = SOCIETY_NAME
    ws["A1"].font = S["font_title"]; ws["A1"].fill = hdr_fill
    ws["A1"].alignment = S["al_ctr"]

    ws.merge_cells("A2:G2")
    ws["A2"] = title_label
    ws["A2"].font = Font(name="Arial", bold=True, color="F9E79F", size=10)
    ws["A2"].fill = hdr_fill; ws["A2"].alignment = S["al_ctr"]

    ws.merge_cells("A3:G3")
    ws["A3"] = f"Generated: {datetime.date.today().strftime('%d %B %Y')}"
    ws["A3"].font = Font(name="Arial", italic=True, color="94A3B8", size=8)
    ws["A3"].fill = hdr_fill; ws["A3"].alignment = S["al_ctr"]

    # ── Column headers ───────────────────────────────────────────────────────
    headers = ["Date", "Particulars / Narration", "Vch Type", "Vch No.",
               "Debit (Rs.)", "Credit (Rs.)", "Balance"]
    col_widths = [13, 42, 16, 18, 14, 14, 20]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=ci, value=h)
        cell.font = S["font_hdr"]; cell.fill = hdr_fill
        cell.alignment = S["al_ctr"]; cell.border = S["border_thin"]
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[4].height = 20

    row = 5

    # ── Opening balance ──────────────────────────────────────────────────────
    ob = acct["opening_balance"]
    ob_str = f"Rs.{abs(ob):,.2f} {'Dr' if ob > 0 else 'Cr'}" if ob != 0 else "Rs. 0.00"
    ob_data = ["—", "Opening Balance  (b/f from previous year)", "b/f", "—",
               ob if ob > 0 else None, abs(ob) if ob < 0 else None, ob_str]
    for ci, val in enumerate(ob_data, 1):
        cell = ws.cell(row=row, column=ci, value=val)
        cell.fill = ob_fill; cell.font = S["font_ob"]
        cell.border = S["border_thin"]
        cell.alignment = S["al_rgt"] if ci in (5, 6) else (S["al_lft"] if ci == 2 else S["al_ctr"])
    ws.row_dimensions[row].height = 16; row += 1

    # ── Entry rows ───────────────────────────────────────────────────────────
    for idx, e in enumerate(acct["entries"]):
        bal = e["balance"]
        bal_str = f"Rs.{abs(bal):,.2f} {'Dr' if bal > 0 else 'Cr'}" if bal != 0 else "Rs. 0.00"

        if e.get("source") == "manual":
            fill = S["fill_manual"]
        elif e["type"] == "Dr":
            fill = S["fill_dr"]
        else:
            fill = S["fill_cr"]

        row_data = [e["date"], e["narration"], e.get("vch_type", ""), e.get("vch_no", ""),
                    e["dr_amt"] if e["dr_amt"] else None,
                    e["cr_amt"] if e["cr_amt"] else None,
                    bal_str]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.fill = fill; cell.font = S["font_body"]
            cell.border = S["border_thin"]
            cell.alignment = S["al_rgt"] if ci in (5, 6) else (S["al_lft"] if ci == 2 else S["al_ctr"])
            if ci in (5, 6) and val is not None:
                cell.number_format = '#,##0.00'
        ws.row_dimensions[row].height = 15; row += 1

    # ── Total row ────────────────────────────────────────────────────────────
    for ci in range(1, 8):
        cell = ws.cell(row=row, column=ci)
        cell.fill = tot_fill; cell.font = S["font_total"]
        cell.border = S["border_thin"]; cell.alignment = S["al_ctr"]
    ws.cell(row=row, column=2, value="TOTAL").alignment = S["al_lft"]
    ws.cell(row=row, column=5, value=acct["total_dr"]).number_format = '#,##0.00'
    ws.cell(row=row, column=5).alignment = S["al_rgt"]
    ws.cell(row=row, column=6, value=acct["total_cr"]).number_format = '#,##0.00'
    ws.cell(row=row, column=6).alignment = S["al_rgt"]
    ws.row_dimensions[row].height = 18; row += 1

    # ── Closing balance ───────────────────────────────────────────────────────
    cb = acct["closing_balance"]
    cb_str = f"Rs.{abs(cb):,.2f} {'Dr (amount due)' if cb > 0 else 'Cr (advance)'}" if cb != 0 else "Rs. 0.00"
    cb_fill = S["fill_cb_due"] if cb > 0 else S["fill_cb_adv"]
    cb_font = S["font_cb_due"] if cb > 0 else S["font_cb_adv"]
    for ci in range(1, 8):
        cell = ws.cell(row=row, column=ci)
        cell.fill = cb_fill; cell.border = S["border_thin"]; cell.font = cb_font
        cell.alignment = S["al_ctr"]
    ws.cell(row=row, column=2, value="Closing Balance  (c/f to next year)").alignment = S["al_lft"]
    ws.cell(row=row, column=7, value=cb_str)
    ws.row_dimensions[row].height = 18

    ws.freeze_panes = "A5"


def generate_flat_account_excel(flat_no, fy_year, filepath):
    if not EXCEL_AVAILABLE:
        messagebox.showerror("Excel Error", "openpyxl not installed.\npip install openpyxl")
        return False
    try:
        acct = get_flat_account_entries(flat_no, fy_year)
        wb = Workbook(); ws = wb.active
        owner = RESIDENTS.get(flat_no, {}).get("name", flat_no)
        ws.title = f"Flat {flat_no}"
        title = f"Flat {flat_no} — {owner} | {fy_label(fy_year)} (April {fy_year} – March {fy_year+1})"
        _xl_write_ledger_sheet(ws, acct, title, style_theme="blue")
        wb.save(filepath); return True
    except Exception as exc:
        messagebox.showerror("Excel Error", str(exc)); return False


def generate_expenditure_account_excel(fy_year, filepath, account_id=1, account_name="Society Expenditure"):
    if not EXCEL_AVAILABLE:
        messagebox.showerror("Excel Error", "openpyxl not installed.\npip install openpyxl")
        return False
    try:
        acct = get_expenditure_account_entries(fy_year, account_id)
        wb = Workbook(); ws = wb.active
        ws.title = re.sub(r'[\\/*?:\[\]]', '-', account_name)[:30]
        title = f"Expenditure: {account_name} | {fy_label(fy_year)} (April {fy_year} – March {fy_year+1})"
        _xl_write_ledger_sheet(ws, acct, title, style_theme="red")
        wb.save(filepath); return True
    except Exception as exc:
        messagebox.showerror("Excel Error", str(exc)); return False


def generate_all_ledger_excel(fy_year, filepath):
    """
    Master Excel export: one sheet per flat + one per expenditure account
    + a Summary sheet with all flat accounts side by side.
    """
    if not EXCEL_AVAILABLE:
        messagebox.showerror("Excel Error", "openpyxl not installed.\npip install openpyxl")
        return False
    try:
        wb = Workbook()
        S = _xl_styles()

        # ── Summary sheet ─────────────────────────────────────────────────
        ws_sum = wb.active; ws_sum.title = "Summary"

        # Title
        ws_sum.merge_cells("A1:G1")
        ws_sum["A1"] = f"{SOCIETY_NAME} — All Flat Accounts — {fy_label(fy_year)}"
        ws_sum["A1"].font = S["font_title"]; ws_sum["A1"].fill = S["fill_hdr_blue"]
        ws_sum["A1"].alignment = S["al_ctr"]
        ws_sum.merge_cells("A2:G2")
        ws_sum["A2"] = f"April {fy_year} – March {fy_year+1}  |  Generated: {datetime.date.today().strftime('%d %B %Y')}"
        ws_sum["A2"].font = Font(name="Arial", italic=True, color="94A3B8", size=8)
        ws_sum["A2"].fill = S["fill_hdr_blue"]; ws_sum["A2"].alignment = S["al_ctr"]

        sum_headers = [UNIT_LABEL, OWNER_LABEL, "Opening Balance", "Total Dr (Bills)", "Total Cr (Payments)", "Closing Balance", "Status"]
        sum_widths   = [8, 30, 18, 18, 18, 18, 14]
        for ci, (h, w) in enumerate(zip(sum_headers, sum_widths), 1):
            cell = ws_sum.cell(row=3, column=ci, value=h)
            cell.font = S["font_hdr"]; cell.fill = S["fill_hdr_blue"]
            cell.alignment = S["al_ctr"]; cell.border = S["border_thin"]
            ws_sum.column_dimensions[get_column_letter(ci)].width = w

        sum_row = 4
        t_ob = t_dr = t_cr = t_cb = 0.0

        def _sum_section_hdr(label, fill_obj):
            nonlocal sum_row
            ws_sum.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=7)
            cell = ws_sum.cell(row=sum_row, column=1, value=label)
            cell.fill = fill_obj; cell.font = S["font_hdr"]; cell.border = S["border_thin"]
            cell.alignment = S["al_lft"]
            for ci in range(2, 8):
                ws_sum.cell(row=sum_row, column=ci).fill = fill_obj
                ws_sum.cell(row=sum_row, column=ci).border = S["border_thin"]
            sum_row += 1

        def _sum_total_row(label, ob_s, dr, cr, cb_s, fill_obj):
            nonlocal sum_row
            vals = [label, "", ob_s, dr, cr, cb_s, ""]
            for ci, val in enumerate(vals, 1):
                cell = ws_sum.cell(row=sum_row, column=ci, value=val)
                cell.fill = fill_obj; cell.font = S["font_total"]; cell.border = S["border_thin"]
                cell.alignment = S["al_rgt"] if ci in (4, 5) else S["al_ctr"]
                if ci in (4, 5) and isinstance(val, float): cell.number_format = '#,##0.00'
            sum_row += 1

        _sum_section_hdr(f"── {UNIT_LABEL.upper()} ACCOUNTS ──", S["fill_hdr_blue"])
        for idx, (flat_no, res) in enumerate(RESIDENTS.items()):
            acct = get_flat_account_entries(flat_no, fy_year)
            ob, dr, cr, cb = acct["opening_balance"], acct["total_dr"], acct["total_cr"], acct["closing_balance"]
            t_ob += ob; t_dr += dr; t_cr += cr; t_cb += cb
            fill = S["fill_dr"] if cb > 0 else (S["fill_cr"] if cb < 0 else S["fill_alt"])
            status = "DUE" if cb > 0 else ("ADVANCE" if cb < 0 else "CLEAR")
            ob_s = f"Rs.{abs(ob):,.2f} {'Dr' if ob>0 else 'Cr'}" if ob != 0 else "Nil"
            cb_s = f"Rs.{abs(cb):,.2f} {'Dr' if cb>0 else 'Cr'}" if cb != 0 else "Nil"
            row_vals = [flat_no, res["name"], ob_s, dr, cr, cb_s, status]
            for ci, val in enumerate(row_vals, 1):
                cell = ws_sum.cell(row=sum_row, column=ci, value=val)
                cell.fill = fill; cell.font = S["font_body"]; cell.border = S["border_thin"]
                cell.alignment = S["al_rgt"] if ci in (4, 5) else (S["al_lft"] if ci == 2 else S["al_ctr"])
                if ci in (4, 5): cell.number_format = '#,##0.00'
            ws_sum.row_dimensions[sum_row].height = 15; sum_row += 1

        _sum_total_row("Sub-Total (Flat Accounts)", f"Rs.{abs(t_ob):,.2f}", t_dr, t_cr,
                       f"Rs.{abs(t_cb):,.2f} {'Dr' if t_cb>0 else 'Cr'}", S["fill_total_blue"])
        sum_row += 1

        _sum_section_hdr("── EXPENDITURE ACCOUNTS ──", S["fill_hdr_red"])
        te_ob = te_dr = te_cr = te_cb = 0.0
        exp_accounts_data = []
        for aid, aname in db_get_expenditure_accounts():
            ea = get_expenditure_account_entries(fy_year, aid)
            e_ob = ea["opening_balance"]; e_dr = ea["total_dr"]; e_cr = ea["total_cr"]; e_cb = ea["closing_balance"]
            te_ob += e_ob; te_dr += e_dr; te_cr += e_cr; te_cb += e_cb
            exp_accounts_data.append((aname, e_dr, e_cr))
            eob_s = f"Rs.{abs(e_ob):,.2f} {'Dr' if e_ob>0 else 'Cr'}" if e_ob != 0 else "Nil"
            ecb_s = f"Rs.{abs(e_cb):,.2f} {'Dr' if e_cb>0 else 'Cr'}" if e_cb != 0 else "Nil"
            status_e = "Net Expense" if e_cb > 0 else ("Net Surplus" if e_cb < 0 else "Balanced")
            row_vals = [aid, aname, eob_s, e_dr, e_cr, ecb_s, status_e]
            for ci, val in enumerate(row_vals, 1):
                cell = ws_sum.cell(row=sum_row, column=ci, value=val)
                cell.fill = S["fill_dr"]; cell.font = S["font_body"]; cell.border = S["border_thin"]
                cell.alignment = S["al_rgt"] if ci in (4, 5) else (S["al_lft"] if ci == 2 else S["al_ctr"])
                if ci in (4, 5): cell.number_format = '#,##0.00'
            ws_sum.row_dimensions[sum_row].height = 15; sum_row += 1

        _sum_total_row("Sub-Total (Expenditure Accounts)", f"Rs.{abs(te_ob):,.2f}", te_dr, te_cr,
                       f"Rs.{abs(te_cb):,.2f} {'Dr' if te_cb>0 else 'Cr'}", S["fill_hdr_red"])
        ws_sum.freeze_panes = "A4"

        # ── Income & Expenditure Statement sheet ──────────────────────────
        ws_ie = wb.create_sheet(title="I&E Statement", index=1)
        ws_ie.merge_cells("A1:C1")
        ws_ie["A1"] = f"{SOCIETY_NAME} — Income & Expenditure Statement — {fy_label(fy_year)}"
        ws_ie["A1"].font = S["font_title"]; ws_ie["A1"].fill = S["fill_hdr_blue"]; ws_ie["A1"].alignment = S["al_ctr"]
        ws_ie.merge_cells("A2:C2")
        ws_ie["A2"] = f"April {fy_year} – March {fy_year+1}  |  Generated: {datetime.date.today().strftime('%d %B %Y')}"
        ws_ie["A2"].font = Font(name="Arial", italic=True, color="94A3B8", size=8)
        ws_ie["A2"].fill = S["fill_hdr_blue"]; ws_ie["A2"].alignment = S["al_ctr"]
        ws_ie.column_dimensions["A"].width = 38
        ws_ie.column_dimensions["B"].width = 22
        ws_ie.column_dimensions["C"].width = 22

        def _ie_hdr(label, fill_obj):
            for ci, txt in enumerate(["Particulars", "Amount (Rs.)", "Total (Rs.)"], 1):
                c = ws_ie.cell(row=label, column=ci, value=txt)
                c.font = S["font_hdr"]; c.fill = fill_obj; c.border = S["border_thin"]; c.alignment = S["al_ctr"]

        def _ie_section(r, label, fill_obj):
            ws_ie.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            c = ws_ie.cell(row=r, column=1, value=label)
            c.font = S["font_hdr"]; c.fill = fill_obj; c.border = S["border_thin"]; c.alignment = S["al_lft"]
            for ci in range(2, 4):
                ws_ie.cell(row=r, column=ci).fill = fill_obj
                ws_ie.cell(row=r, column=ci).border = S["border_thin"]

        def _ie_row(r, label, amt, total=None):
            c1 = ws_ie.cell(row=r, column=1, value=label)
            c1.font = S["font_body"]; c1.border = S["border_thin"]; c1.alignment = S["al_lft"]
            c2 = ws_ie.cell(row=r, column=2, value=amt if total is None else "")
            c2.font = S["font_body"]; c2.border = S["border_thin"]; c2.alignment = S["al_rgt"]
            if isinstance(amt, float) and total is None: c2.number_format = '#,##0.00'
            c3 = ws_ie.cell(row=r, column=3, value=total if total is not None else "")
            c3.font = S["font_body"]; c3.border = S["border_thin"]; c3.alignment = S["al_rgt"]
            if isinstance(total, float): c3.number_format = '#,##0.00'

        def _ie_total_row(r, label, val, bold=True):
            c1 = ws_ie.cell(row=r, column=1, value=label)
            c1.fill = S["fill_total_blue"]; c1.font = S["font_total"]; c1.border = S["border_thin"]
            ws_ie.cell(row=r, column=2).fill = S["fill_total_blue"]; ws_ie.cell(row=r, column=2).border = S["border_thin"]
            c3 = ws_ie.cell(row=r, column=3, value=val)
            c3.fill = S["fill_total_blue"]; c3.font = S["font_total"]; c3.border = S["border_thin"]; c3.alignment = S["al_rgt"]
            if isinstance(val, float): c3.number_format = '#,##0.00'

        adm_rows = db_fetch_admission_fees()
        adm_total = sum(r[5] for r in adm_rows)
        total_income = t_cr + adm_total

        ie_r = 3
        _ie_hdr(ie_r, S["fill_hdr_blue"]); ie_r += 1
        _ie_section(ie_r, "INCOME", S["fill_hdr_blue"]); ie_r += 1
        _ie_row(ie_r, "Maintenance Collections (from Flats)", t_cr); ie_r += 1
        _ie_row(ie_r, "Admission Fees Received", adm_total); ie_r += 1
        for aid, aname in db_get_expenditure_accounts():
            ea = get_expenditure_account_entries(fy_year, aid)
            if ea["total_cr"] > 0:
                _ie_row(ie_r, f"  {aname} (receipts)", ea["total_cr"]); ie_r += 1
        _ie_total_row(ie_r, "TOTAL INCOME", total_income); ie_r += 2

        _ie_section(ie_r, "EXPENDITURE", S["fill_hdr_red"]); ie_r += 1
        for aname, e_dr, e_cr in exp_accounts_data:
            net_exp = e_dr - e_cr
            if net_exp != 0:
                _ie_row(ie_r, f"  {aname}", net_exp); ie_r += 1
        _ie_total_row(ie_r, "TOTAL EXPENDITURE", te_dr - te_cr); ie_r += 2

        net = total_income - (te_dr - te_cr)
        surplus_label = "NET SURPLUS FOR THE YEAR" if net >= 0 else "NET DEFICIT FOR THE YEAR"
        for ci in range(1, 4):
            c = ws_ie.cell(row=ie_r, column=ci)
            c.fill = PatternFill("solid", fgColor="166534" if net >= 0 else "991B1B")
            c.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
            c.border = S["border_thin"]; c.alignment = S["al_ctr"]
        ws_ie.cell(row=ie_r, column=1, value=surplus_label).alignment = S["al_lft"]
        ws_ie.cell(row=ie_r, column=3, value=abs(net)).number_format = '#,##0.00'
        ws_ie.freeze_panes = "A4"

        # ── One sheet per flat ────────────────────────────────────────────
        for flat_no, res in RESIDENTS.items():
            ws = wb.create_sheet(title=f"Flat {flat_no}")
            acct = get_flat_account_entries(flat_no, fy_year)
            owner = res["name"]
            title = f"Flat {flat_no} — {owner} | {fy_label(fy_year)}"
            _xl_write_ledger_sheet(ws, acct, title, style_theme="blue")

        # ── One sheet per expenditure account ─────────────────────────────
        for aid, aname in db_get_expenditure_accounts():
            safe_name = re.sub(r'[\\/*?:\[\]]', '-', aname)[:28]
            ws = wb.create_sheet(title=f"Exp-{safe_name}")
            acct = get_expenditure_account_entries(fy_year, aid)
            title = f"Expenditure: {aname} | {fy_label(fy_year)}"
            _xl_write_ledger_sheet(ws, acct, title, style_theme="red")

        wb.save(filepath); return True
    except Exception as exc:
        messagebox.showerror("Excel Error", str(exc)); return False


def generate_master_ledger_excel(fy_year, filepath):
    if not EXCEL_AVAILABLE:
        messagebox.showerror("Excel Error", "openpyxl not installed.\npip install openpyxl")
        return False
    try:
        wb = Workbook(); S = _xl_styles()
        ws = wb.active; ws.title = "Master Ledger"
        cols = ["Account", "Type", "Opening Bal", "Total Dr (Rs.)", "Total Cr (Rs.)", "Closing Bal", "Net Position"]
        widths = [36, 16, 18, 18, 18, 20, 16]
        ws.merge_cells("A1:G1")
        ws["A1"] = f"{SOCIETY_NAME} — Master Ledger — {fy_label(fy_year)}"
        ws["A1"].font = S["font_title"]; ws["A1"].fill = S["fill_hdr_blue"]; ws["A1"].alignment = S["al_ctr"]
        ws.merge_cells("A2:G2")
        ws["A2"] = f"April {fy_year} – March {fy_year+1}  |  Generated: {datetime.date.today().strftime('%d %B %Y')}"
        ws["A2"].font = Font(name="Arial", italic=True, color="94A3B8", size=8)
        ws["A2"].fill = S["fill_hdr_blue"]; ws["A2"].alignment = S["al_ctr"]
        for ci, (h, w) in enumerate(zip(cols, widths), 1):
            c = ws.cell(row=3, column=ci, value=h)
            c.font = S["font_hdr"]; c.fill = S["fill_hdr_blue"]
            c.alignment = S["al_ctr"]; c.border = S["border_thin"]
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "A4"
        row = 4
        def _section_hdr(label, fill):
            for ci in range(1, 8):
                c = ws.cell(row=row, column=ci)
                c.fill = fill; c.font = S["font_hdr"]; c.border = S["border_thin"]
            ws.cell(row=row, column=1, value=label).alignment = S["al_lft"]
        def _data_row(vals, fill):
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.fill = fill; c.font = S["font_body"]; c.border = S["border_thin"]
                c.alignment = S["al_rgt"] if ci in (4,5) else S["al_ctr"] if ci != 2 else S["al_lft"]
                if ci in (4,5) and isinstance(v, (int,float)): c.number_format = '#,##0.00'
            ws.row_dimensions[row].height = 15
        def _total_row(vals):
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.fill = S["fill_total_blue"]; c.font = S["font_total"]; c.border = S["border_thin"]
                c.alignment = S["al_rgt"] if ci in (4,5) else S["al_ctr"]
                if ci in (4,5) and isinstance(v, (int,float)): c.number_format = '#,##0.00'
            ws.row_dimensions[row].height = 16

        _section_hdr(f"── {UNIT_LABEL.upper()} ACCOUNTS ──", S["fill_hdr_blue"]); row += 1
        f_ob=f_dr=f_cr=f_cb=0.0
        for flat_no, res in RESIDENTS.items():
            acct = get_flat_account_entries(flat_no, fy_year)
            ob=acct["opening_balance"]; dr=acct["total_dr"]; cr=acct["total_cr"]; cb=acct["closing_balance"]
            f_ob+=ob; f_dr+=dr; f_cr+=cr; f_cb+=cb
            ob_s = f"Rs.{abs(ob):,.2f} {'Dr' if ob>0 else 'Cr'}" if ob!=0 else "Nil"
            cb_s = f"Rs.{abs(cb):,.2f} {'Dr' if cb>0 else 'Cr'}" if cb!=0 else "Nil"
            net = "Due" if cb>0 else ("Advance" if cb<0 else "Clear")
            fill = S["fill_dr"] if cb>0 else (S["fill_cr"] if cb<0 else S["fill_alt"])
            _data_row([f"{flat_no}  {res['name']}", "Maintenance", ob_s, dr, cr, cb_s, net], fill); row += 1
        ftcb_s = f"Rs.{abs(f_cb):,.2f} {'Dr' if f_cb>0 else 'Cr'}" if f_cb!=0 else "Nil"
        _total_row(["Sub-Total (Flat Accounts)", "", f"Rs.{abs(f_ob):,.2f}", f_dr, f_cr, ftcb_s, ""]); row += 1

        adm_rows = db_fetch_admission_fees()
        adm_total = sum(r[5] for r in adm_rows)
        _section_hdr("── ADMISSION FEES ──", PatternFill("solid", fgColor="1A2332")); row += 1
        for r_ in adm_rows:
            _, rno, date, flat, owner, amt, mob, _ = r_
            _data_row([f"{UNIT_LABEL} {flat}  ({owner})", "Admission", date, 0, amt, f"Rs.{amt:,.2f} Cr", "Paid"], S["fill_cr"]); row += 1
        _total_row(["Sub-Total (Admission Fees)", "", "", 0, adm_total, f"Rs.{adm_total:,.2f} Cr", ""]); row += 1

        _section_hdr("── EXPENDITURE ACCOUNTS ──", S["fill_hdr_red"]); row += 1
        total_e_dr=total_e_cr=0.0
        for aid, aname in db_get_expenditure_accounts():
            ea=get_expenditure_account_entries(fy_year, aid)
            e_ob=ea["opening_balance"]; e_dr=ea["total_dr"]; e_cr=ea["total_cr"]; e_cb=ea["closing_balance"]
            total_e_dr+=e_dr; total_e_cr+=e_cr
            eob_s=f"Rs.{abs(e_ob):,.2f} {'Dr' if e_ob>0 else 'Cr'}" if e_ob!=0 else "Nil"
            ecb_s=f"Rs.{abs(e_cb):,.2f} {'Dr' if e_cb>0 else 'Cr'}" if e_cb!=0 else "Nil"
            net_e="Net Expense" if e_cb>0 else "Net Surplus" if e_cb<0 else "Balanced"
            _data_row([aname, "Expenditure", eob_s, e_dr, e_cr, ecb_s, net_e], S["fill_dr"]); row += 1

        g_dr=f_dr+total_e_dr; g_cr=f_cr+adm_total+total_e_cr; g_net=f_cr+adm_total-total_e_dr
        gcb_s=f"Rs.{abs(g_net):,.2f} {'Surplus' if g_net>=0 else 'Deficit'}"
        row += 1
        _total_row(["GRAND TOTAL", "", "", g_dr, g_cr, gcb_s, "Net Position"]); row += 1

        row += 1
        for label, val in [("Total Income (Maintenance)", f"Rs.{f_cr:,.2f}"),
                            ("Total Admission Fees", f"Rs.{adm_total:,.2f}"),
                            ("Total Expenditure", f"Rs.{total_e_dr:,.2f}"),
                            ("Net Surplus / Deficit", gcb_s)]:
            ws.cell(row=row, column=1, value=label).font = S["font_body"]
            ws.cell(row=row, column=2, value=val).font = Font(name="Arial", bold=True, size=10)
            row += 1

        wb.save(filepath); return True
    except Exception as exc:
        messagebox.showerror("Excel Error", str(exc)); return False


def generate_master_ledger_pdf(fy_year, filepath):
    if not PDF_AVAILABLE:
        messagebox.showerror("PDF Error", "reportlab not installed.\npip install reportlab")
        return False
    try:
        doc = SimpleDocTemplate(filepath, pagesize=A4,
                                leftMargin=1.5*cm, rightMargin=1.5*cm,
                                topMargin=1.5*cm, bottomMargin=1.5*cm)
        styles = getSampleStyleSheet()
        elems = []
        title_style = ParagraphStyle("T", fontName="Helvetica-Bold", fontSize=14,
                                     textColor=colors.HexColor("#E2E8F0"), spaceAfter=4, alignment=TA_CENTER)
        sub_style   = ParagraphStyle("S", fontName="Helvetica", fontSize=8,
                                     textColor=colors.HexColor("#94A3B8"), spaceAfter=12, alignment=TA_CENTER)
        elems.append(Paragraph(SOCIETY_NAME, title_style))
        elems.append(Paragraph(f"Master Ledger — {fy_label(fy_year)}", title_style))
        elems.append(Paragraph(f"April {fy_year} – March {fy_year+1}  |  Generated: {datetime.date.today().strftime('%d %B %Y')}", sub_style))
        elems.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#1E2A3A"), spaceAfter=10))
        hdr_c   = colors.HexColor("#1E3A5F"); sub_c  = colors.HexColor("#0F2238")
        red_c   = colors.HexColor("#7C1D1D"); dr_c   = colors.HexColor("#2D0F0F")
        cr_c    = colors.HexColor("#052E16"); tot_c  = colors.HexColor("#0D1117")
        txt_wht = colors.HexColor("#E2E8F0"); txt_ok = colors.HexColor("#86EFAC")
        txt_err = colors.HexColor("#FCA5A5"); txt_dim= colors.HexColor("#94A3B8")
        col_w   = [5.5*cm, 2.2*cm, 2.2*cm, 2.4*cm, 2.4*cm, 2.8*cm, 2.4*cm]
        hdrs    = ["Account", "Type", "Opening", "Dr (Rs.)", "Cr (Rs.)", "Closing", "Position"]
        def _hdr_row(label, bg):
            return [[Paragraph(f"<b>{label}</b>", ParagraphStyle("h", fontName="Helvetica-Bold", fontSize=8, textColor=txt_wht))] + [""]*6]
        def _col_hdrs():
            return [[Paragraph(f"<b>{h}</b>", ParagraphStyle("ch", fontName="Helvetica-Bold", fontSize=7, textColor=txt_wht, alignment=TA_CENTER)) for h in hdrs]]
        def _fmt(v): return f"Rs.{abs(v):,.2f}" if v else "-"
        def _mk_table(data, style_cmds):
            t = Table(data, colWidths=col_w)
            t.setStyle(TableStyle(style_cmds)); return t
        base_style = [("FONTNAME", (0,0), (-1,-1), "Helvetica"),
                      ("FONTSIZE", (0,0), (-1,-1), 7),
                      ("TEXTCOLOR", (0,0), (-1,-1), txt_dim),
                      ("ROWBACKGROUNDS", (0,0), (-1,-1), [colors.HexColor("#161B27"), colors.HexColor("#111520")]),
                      ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#1E2A3A")),
                      ("ALIGN", (2,0), (-1,-1), "RIGHT"), ("ALIGN", (0,0), (0,-1), "LEFT")]
        elems.append(_mk_table(_col_hdrs(), [("BACKGROUND",(0,0),(-1,-1),hdr_c),
                                              ("TEXTCOLOR",(0,0),(-1,-1),txt_wht),
                                              ("ALIGN",(0,0),(-1,-1),"CENTER"),
                                              ("FONTNAME",(0,0),(-1,-1),"Helvetica-Bold"),
                                              ("FONTSIZE",(0,0),(-1,-1),7),
                                              ("GRID",(0,0),(-1,-1),0.3,colors.HexColor("#1E2A3A"))]))
        elems.append(Spacer(1, 2))
        elems.append(_mk_table(_hdr_row(f"── {UNIT_LABEL.upper()} ACCOUNTS ──", hdr_c),
                                [("BACKGROUND",(0,0),(-1,-1),hdr_c),("SPAN",(0,0),(-1,0)),
                                 ("FONTNAME",(0,0),(-1,-1),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),8),
                                 ("TEXTCOLOR",(0,0),(-1,-1),txt_wht)]))
        flat_data = []; f_ob=f_dr=f_cr=f_cb=0.0
        for flat_no, res in RESIDENTS.items():
            acct = get_flat_account_entries(flat_no, fy_year)
            ob=acct["opening_balance"]; dr=acct["total_dr"]; cr=acct["total_cr"]; cb=acct["closing_balance"]
            f_ob+=ob; f_dr+=dr; f_cr+=cr; f_cb+=cb
            ob_s = _fmt(ob)+"Dr" if ob>0 else _fmt(ob)+"Cr" if ob<0 else "Nil"
            cb_s = _fmt(cb)+"Dr" if cb>0 else _fmt(cb)+"Cr" if cb<0 else "Nil"
            flat_data.append([f"{flat_no} {res['name'][:22]}", "Maint.", ob_s, _fmt(dr), _fmt(cr), cb_s,
                               "Due" if cb>0 else ("Adv" if cb<0 else "Clear")])
        ftcb_s = _fmt(f_cb)+("Dr" if f_cb>0 else "Cr") if f_cb!=0 else "Nil"
        flat_data.append(["Sub-Total","","",_fmt(f_dr),_fmt(f_cr),ftcb_s,""])
        t = Table(flat_data, colWidths=col_w); s_cmds = list(base_style)
        s_cmds += [("BACKGROUND",(0,-1),(-1,-1),sub_c),("TEXTCOLOR",(0,-1),(-1,-1),txt_wht),
                   ("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold")]
        t.setStyle(TableStyle(s_cmds)); elems.append(t); elems.append(Spacer(1,4))
        adm_rows = db_fetch_admission_fees(); adm_total = sum(r[5] for r in adm_rows)
        if adm_rows:
            elems.append(_mk_table(_hdr_row("── ADMISSION FEES ──", colors.HexColor("#1A2332")),
                                    [("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#1A2332")),("SPAN",(0,0),(-1,0)),
                                     ("FONTNAME",(0,0),(-1,-1),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),8),
                                     ("TEXTCOLOR",(0,0),(-1,-1),txt_wht)]))
            adm_data = [[f"{UNIT_LABEL} {r[3]} ({r[4][:16]})", "Adm.", r[2], "-", f"Rs.{r[5]:,.2f}", f"Rs.{r[5]:,.2f} Cr","Paid"] for r in adm_rows]
            adm_data.append(["Sub-Total","","","-",f"Rs.{adm_total:,.2f}",f"Rs.{adm_total:,.2f} Cr",""])
            ta = Table(adm_data, colWidths=col_w); ta_cmds = list(base_style)
            ta_cmds += [("BACKGROUND",(0,-1),(-1,-1),sub_c),("TEXTCOLOR",(0,-1),(-1,-1),txt_ok),
                        ("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold")]
            ta.setStyle(TableStyle(ta_cmds)); elems.append(ta); elems.append(Spacer(1,4))
        elems.append(_mk_table(_hdr_row("── EXPENDITURE ACCOUNTS ──", red_c),
                                [("BACKGROUND",(0,0),(-1,-1),red_c),("SPAN",(0,0),(-1,0)),
                                 ("FONTNAME",(0,0),(-1,-1),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),8),
                                 ("TEXTCOLOR",(0,0),(-1,-1),txt_wht)]))
        exp_data = []; total_e_dr=total_e_cr=0.0
        for aid, aname in db_get_expenditure_accounts():
            ea=get_expenditure_account_entries(fy_year, aid)
            e_dr=ea["total_dr"]; e_cr=ea["total_cr"]; e_cb=ea["closing_balance"]
            total_e_dr+=e_dr; total_e_cr+=e_cr
            ecb_s=(_fmt(e_cb)+"Dr" if e_cb>0 else _fmt(e_cb)+"Cr") if e_cb!=0 else "Nil"
            exp_data.append([aname[:28],"Exp.","Nil",_fmt(e_dr),_fmt(e_cr),ecb_s,
                              "Expense" if e_cb>0 else "Surplus"])
        te=Table(exp_data, colWidths=col_w); te_cmds=list(base_style)
        te_cmds+=[("TEXTCOLOR",(0,0),(-1,-1),txt_err)]
        te.setStyle(TableStyle(te_cmds)); elems.append(te); elems.append(Spacer(1,6))
        g_cr=f_cr+adm_total; g_net=g_cr-total_e_dr
        gcb_s=f"Rs.{abs(g_net):,.2f} {'Surplus' if g_net>=0 else 'Deficit'}"
        grand = Table([["GRAND TOTAL","","",f"Rs.{f_dr+total_e_dr:,.2f}",f"Rs.{g_cr:,.2f}",gcb_s,"Net"]],
                      colWidths=col_w)
        grand.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#0D1117")),
                                    ("TEXTCOLOR",(0,0),(-1,-1),txt_wht),
                                    ("FONTNAME",(0,0),(-1,-1),"Helvetica-Bold"),
                                    ("FONTSIZE",(0,0),(-1,-1),8),
                                    ("ALIGN",(2,0),(-1,-1),"RIGHT"),
                                    ("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#5B6AF0"))]))
        elems.append(grand)
        doc.build(elems); return True
    except Exception as exc:
        messagebox.showerror("PDF Error", str(exc)); return False


def normalize_mobile(mobile):
    mob = mobile.strip().replace(" ","").replace("-","").lstrip("0").lstrip("+")
    if len(mob) == 10: mob = "91" + mob
    elif not mob.startswith("91"): mob = "91" + mob
    return mob


def _card(parent, **kw):
    return tk.Frame(parent, bg=CARD, relief="flat", **kw)

def _section_label(parent, text):
    return tk.Label(parent, text=text, bg=CARD, fg=ACCENT2, font=("Segoe UI", 7, "bold"))

def _field_label(parent, text):
    return tk.Label(parent, text=text, bg=CARD, fg=TEXT2, font=("Segoe UI", 9, "bold"), anchor="w")

def _h_sep(parent, bg=BORDER):
    return tk.Frame(parent, bg=bg, height=1)

def _round_rect(c, x1, y1, x2, y2, r, **kw):
    pts = [x1+r,y1, x2-r,y1, x2,y1, x2,y1+r, x2,y2-r, x2,y2, x2-r,y2, x1+r,y2, x1,y2, x1,y2-r, x1,y1+r, x1,y1]
    return c.create_polygon(pts, smooth=True, **kw)

def _icon_btn(parent, text, bg, hover, fg="white", cmd=None, **kw):
    import tkinter.font as tkf
    for k in ("padx","pady","bd"): kw.pop(k, None)
    fnt = ("Segoe UI", 9, "bold")
    fm = tkf.Font(family="Segoe UI", size=9, weight="bold")
    w = fm.measure(text) + 38; h = fm.metrics("linespace") + 20
    try: pbg = parent.cget("bg")
    except: pbg = BG
    c = tk.Canvas(parent, width=w, height=h, bg=pbg, highlightthickness=0, bd=0, cursor="hand2", **kw)
    def _draw(col):
        c.delete("all")
        _round_rect(c, 1, 1, w-1, h-1, 10, fill=col, outline=col)
        c.create_text(w//2, h//2, text=text, fill=fg, font=fnt)
    _draw(bg)
    c.bind("<Enter>", lambda e: _draw(hover))
    c.bind("<Leave>", lambda e: _draw(bg))
    if cmd: c.bind("<Button-1>", lambda e: cmd())
    return c

def _pill_label(parent, text, bg=PILL_BG, fg=PILL_FG):
    return tk.Label(parent, text=text, bg=bg, fg=fg, font=("Segoe UI", 8, "bold"), padx=8, pady=3)

def _sidebar_btn(parent, text, icon, cmd, active=False):
    bg_n = SIDEBAR_S if active else SIDEBAR
    f = tk.Frame(parent, bg=bg_n, cursor="hand2")
    f.pack(fill="x", pady=0, padx=6)
    bar = tk.Frame(f, bg=ACCENT if active else bg_n, width=3)
    bar.pack(side="left", fill="y")
    inner = tk.Frame(f, bg=bg_n); inner.pack(fill="x", padx=12, pady=10)
    icon_lbl = tk.Label(inner, text=icon, bg=bg_n, fg=ACCENT if active else TEXT3, font=("Segoe UI", 11))
    icon_lbl.pack(side="left")
    text_lbl = tk.Label(inner, text=text, bg=bg_n, fg=ACCENT2 if active else TEXT2,
                        font=("Segoe UI", 9, "bold" if active else "normal"))
    text_lbl.pack(side="left", padx=10)
    all_w = [f, inner, icon_lbl, text_lbl]
    def _enter(e):
        for w in all_w: w.configure(bg=SIDEBAR_H)
        bar.configure(bg=ACCENT if active else BORDER2)
    def _leave(e):
        for w in all_w: w.configure(bg=bg_n)
        bar.configure(bg=ACCENT if active else bg_n)
    def _click(e): cmd()
    for w in all_w:
        w.bind("<Enter>", _enter); w.bind("<Leave>", _leave); w.bind("<Button-1>", _click)
    bar.bind("<Enter>", _enter); bar.bind("<Leave>", _leave); bar.bind("<Button-1>", _click)
    return f


class FlatsManagerWindow(tk.Toplevel):
    def __init__(self, parent, on_close=None):
        super().__init__(parent)
        self._on_close = on_close
        self.title(f"Manage {UNIT_LABEL}s")
        self.geometry("760x540"); self.configure(bg=BG); self.grab_set()
        self.protocol("WM_DELETE_WINDOW", self._close)
        self._build(); self._load()

    def _close(self):
        if self._on_close: self._on_close()
        self.destroy()

    def _build(self):
        hdr = tk.Frame(self, bg=TEAL, height=56); hdr.pack(fill="x"); hdr.pack_propagate(False)
        tk.Label(hdr, text=f"  🏠  Manage {UNIT_LABEL}s", bg=TEAL, fg="white",
                 font=("Segoe UI", 13, "bold")).pack(side="left", padx=20, pady=16)
        _icon_btn(hdr, f"+ Add {UNIT_LABEL}", SUCCESS, SUCCESS_H, cmd=self._add_dlg).pack(side="right", padx=12, pady=12)

        info = tk.Frame(self, bg="#0A0C18", height=24); info.pack(fill="x"); info.pack_propagate(False)
        tk.Label(info, text="  Double-click to edit  ·  Right-click to delete  ·  ↑↓ buttons to reorder",
                 bg="#0A0C18", fg="#404870", font=("Segoe UI", 8, "italic")).pack(side="left", padx=14)

        body = tk.Frame(self, bg=BG); body.pack(fill="both", expand=True, padx=10, pady=8)

        cols = ("#", f"{UNIT_LABEL} No", f"{OWNER_LABEL} Name", "Mobile", "Order")
        self._tree = ttk.Treeview(body, columns=cols, show="headings", height=16, selectmode="browse")
        for c, w, anc in [("#",35,"center"),(f"{UNIT_LABEL} No",70,"center"),(f"{OWNER_LABEL} Name",280,"w"),
                           ("Mobile",130,"center"),("Order",80,"center")]:
            self._tree.heading(c, text=c); self._tree.column(c, width=w, anchor=anc)
        vsb = ttk.Scrollbar(body, orient="vertical", command=self._tree.yview)
        self._tree.configure(yscrollcommand=vsb.set)
        self._tree.pack(side="left", fill="both", expand=True)
        vsb.pack(side="left", fill="y")
        self._tree.tag_configure("odd",  background=ROW_ODD,  foreground=TEXT)
        self._tree.tag_configure("even", background=ROW_EVEN, foreground=TEXT)
        self._tree.bind("<Double-1>", lambda e: self._edit_selected())
        self._tree.bind("<Button-3>", self._rclick)

        ctrl = tk.Frame(body, bg=BG, padx=8); ctrl.pack(side="left", fill="y")
        _icon_btn(ctrl, "✎ Edit",   ACCENT,   ACCENT_H,  cmd=self._edit_selected ).pack(fill="x", pady=3)
        _icon_btn(ctrl, "✕ Delete", DANGER,   DANGER_H,  cmd=self._delete_selected).pack(fill="x", pady=3)
        tk.Frame(ctrl, bg=BORDER2, height=1).pack(fill="x", pady=8)
        _icon_btn(ctrl, "▲ Up",   WARNING,  WARNING_H, cmd=lambda: self._move("up")  ).pack(fill="x", pady=3)
        _icon_btn(ctrl, "▼ Down", WARNING,  WARNING_H, cmd=lambda: self._move("down")).pack(fill="x", pady=3)

        self._count_var = tk.StringVar()
        tk.Label(self, textvariable=self._count_var, bg=BG, fg=TEXT3,
                 font=("Segoe UI", 8)).pack(anchor="w", padx=14, pady=(0,6))

    def _load(self):
        for i in self._tree.get_children(): self._tree.delete(i)
        rows = db_get_flats()
        for idx, (flat_no, name, mobile, sort_order) in enumerate(rows, 1):
            tag = "odd" if idx % 2 else "even"
            self._tree.insert("", "end", iid=flat_no,
                              values=(idx, flat_no, name, mobile or "—", sort_order), tags=(tag,))
        self._count_var.set(f"  {len(rows)} flat(s) registered")

    def _selected_flat(self):
        sel = self._tree.selection()
        return sel[0] if sel else None

    def _edit_selected(self):
        fno = self._selected_flat()
        if not fno: return
        info = RESIDENTS.get(fno, {})
        FlatEditDialog(self, flat_no=fno, name=info.get("name",""), mobile=info.get("mobile",""),
                       refresh_cb=self._load, edit_mode=True)

    def _delete_selected(self):
        fno = self._selected_flat()
        if not fno: return
        if messagebox.askyesno("Delete Flat",
                               f"Permanently remove Flat {fno} — {RESIDENTS.get(fno,{}).get('name','')}?\n\n"
                               "This cannot be undone.", icon="warning", parent=self):
            db_delete_flat(fno); self._load()

    def _move(self, direction):
        fno = self._selected_flat()
        if not fno: return
        db_reorder_flat(fno, direction); self._load()
        try: self._tree.selection_set(fno); self._tree.see(fno)
        except Exception: pass

    def _rclick(self, event):
        item = self._tree.identify_row(event.y)
        if not item: return
        self._tree.selection_set(item)
        m = tk.Menu(self, tearoff=0)
        m.add_command(label="Edit", command=self._edit_selected)
        m.add_command(label="Delete", command=self._delete_selected)
        m.post(event.x_root, event.y_root)

    def _add_dlg(self):
        FlatEditDialog(self, refresh_cb=self._load, edit_mode=False)


class FlatEditDialog(tk.Toplevel):
    def __init__(self, parent, flat_no="", name="", mobile="", refresh_cb=None, edit_mode=False):
        super().__init__(parent)
        self._refresh = refresh_cb; self._edit_mode = edit_mode; self._orig_flat = flat_no
        title = f"Edit Flat — {flat_no}" if edit_mode else "Add New Flat"
        self.title(title); self.geometry("400x260"); self.resizable(False, False)
        self.configure(bg=CARD); self.grab_set()

        hdr_bg = ACCENT if not edit_mode else TEAL
        tk.Label(self, text=f"  {'✎ Edit' if edit_mode else '+ Add'} Flat",
                 bg=hdr_bg, fg="white", font=("Segoe UI", 11, "bold"), pady=12).pack(fill="x")

        frm = tk.Frame(self, bg=CARD, padx=28, pady=18); frm.pack(fill="both", expand=True)

        self._fno_var  = tk.StringVar(value=flat_no)
        self._name_var = tk.StringVar(value=name)
        self._mob_var  = tk.StringVar(value=mobile)

        fields = [(f"{UNIT_LABEL} No", self._fno_var),
                  (f"{OWNER_LABEL} Name",        self._name_var),
                  ("Mobile (10 digits)",     self._mob_var)]
        for r, (lbl, var) in enumerate(fields):
            _field_label(frm, lbl).grid(row=r, column=0, sticky="w", pady=7, padx=(0,16))
            e = ttk.Entry(frm, textvariable=var, width=26)
            e.grid(row=r, column=1, sticky="w")
            if edit_mode and r == 0: e.configure(state="disabled")
            if r == 0: e.focus()

        bf = tk.Frame(frm, bg=CARD); bf.grid(row=3, column=0, columnspan=2, pady=14)
        _icon_btn(bf, "Save", SUCCESS, SUCCESS_H, cmd=self._save).pack(side="left", padx=(0,8))
        _icon_btn(bf, "Cancel", DANGER, DANGER_H, cmd=self.destroy).pack(side="left")

    def _save(self):
        fno  = self._fno_var.get().strip().upper()
        name = self._name_var.get().strip().upper()
        mob  = self._mob_var.get().strip()
        if not fno or not name:
            messagebox.showerror("Missing", f"{UNIT_LABEL} No and {OWNER_LABEL} Name are required.", parent=self); return
        if self._edit_mode:
            db_update_flat(fno, name, mob)
        else:
            if not db_add_flat(fno, name, mob): return
        if self._refresh: self._refresh()
        self.destroy()


class WhatsAppDialog(tk.Toplevel):
    def __init__(self, parent, mobile, message, title_extra=""):
        super().__init__(parent)
        self._mob = normalize_mobile(mobile); self._msg = message
        self.title("Send via WhatsApp"); self.geometry("480x500")
        self.resizable(False, False); self.configure(bg=CARD); self.grab_set(); self.focus_force()

        hdr = tk.Frame(self, bg="#1A4E3A", height=56)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        tk.Label(hdr, text="  💬  WhatsApp", bg="#1A4E3A", fg="white",
                 font=("Segoe UI", 13, "bold")).pack(side="left", padx=20, pady=16)

        info = tk.Frame(self, bg=OK_BG, height=34)
        info.pack(fill="x"); info.pack_propagate(False)
        lbl = f"  To: +{self._mob}" + (f"   ·   {title_extra}" if title_extra else "")
        tk.Label(info, text=lbl, bg=OK_BG, fg=OK_FG,
                 font=("Segoe UI", 9, "bold")).pack(side="left", padx=16, pady=9)

        pf = tk.Frame(self, bg=CARD, padx=16, pady=8); pf.pack(fill="both", expand=True)
        txt = tk.Text(pf, font=("Consolas", 10), bg=BG, fg=TEXT, relief="flat",
                      wrap="word", padx=12, pady=10, height=14, cursor="arrow",
                      highlightthickness=1, highlightbackground=BORDER2)
        txt.insert("1.0", self._msg); txt.configure(state="disabled")
        txt.pack(fill="both", expand=True)

        _h_sep(self).pack(fill="x")
        bf = tk.Frame(self, bg=CARD, padx=16, pady=12); bf.pack(fill="x")
        _icon_btn(bf, "Open App", WA_GRN, WA_GRN_H, cmd=self._open_app).pack(side="left", padx=(0,6))
        _icon_btn(bf, "Web", "#075E54", "#054438", cmd=self._open_web).pack(side="left", padx=(0,6))
        _icon_btn(bf, "Copy", ACCENT, ACCENT_H, cmd=self._copy).pack(side="left")
        _icon_btn(bf, "Close", DANGER, DANGER_H, cmd=self.destroy).pack(side="right")

    def _open_app(self):
        try:
            enc = urllib.parse.quote(self._msg, safe="")
            uri = f"whatsapp://send?phone={self._mob}&text={enc}"
            if sys.platform == "win32": os.startfile(uri)
            else: webbrowser.open(uri)
        except Exception: self._open_web()

    def _open_web(self):
        try:
            enc = urllib.parse.quote(self._msg, safe="")
            webbrowser.open(f"https://wa.me/{self._mob}?text={enc}")
        except Exception as e: messagebox.showerror("Error", str(e), parent=self)

    def _copy(self):
        self.clipboard_clear(); self.clipboard_append(self._msg)
        messagebox.showinfo("Copied", "Message copied to clipboard.", parent=self)


class CrossFYSplitDialog(tk.Toplevel):
    def __init__(self, parent, rec_a, rec_b, this_fy_count, next_fy_count):
        super().__init__(parent)
        self.confirmed = False; self.rec_a = rec_a; self.rec_b = rec_b
        self.title("Cross-Year Payment"); self.geometry("500x300")
        self.resizable(False, False); self.configure(bg=CARD); self.grab_set(); self.focus_force()

        hdr = tk.Frame(self, bg=WARNING, height=56)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        tk.Label(hdr, text="  ⚠  Cross Financial Year Payment",
                 bg=WARNING, fg="#0A0800", font=("Segoe UI", 11, "bold")).pack(side="left", padx=20, pady=18)

        body = tk.Frame(self, bg=CARD, padx=28, pady=20); body.pack(fill="both", expand=True)
        total_amt = float(rec_a["amount"]) + float(rec_b["amount"])
        tk.Label(body, text=f"This payment of Rs.{total_amt:,.0f} spans two financial years.",
                 bg=CARD, fg=TEXT, font=("Segoe UI", 11, "bold")).pack(anchor="w")
        tk.Label(body, text="It will be split and saved as two linked records automatically.",
                 bg=CARD, fg=TEXT2, font=("Segoe UI", 9)).pack(anchor="w", pady=(4, 16))

        for rec, label in [(rec_a, f"Record 1 — {fy_label(rec_a['year_from'])}"),
                           (rec_b, f"Record 2 — {fy_label(rec_b['year_from'])}")]:
            card = tk.Frame(body, bg=BG, highlightbackground=BORDER2, highlightthickness=1)
            card.pack(fill="x", pady=3)
            inner = tk.Frame(card, bg=BG, padx=14, pady=8); inner.pack(fill="x")
            tk.Label(inner, text=label, bg=BG, fg=TEXT3, font=("Segoe UI", 8, "bold")).pack(anchor="w")
            tk.Label(inner, text=f"{rec['month_from']} → {rec['month_to']}   ·   Rs.{rec['amount']:,.0f}",
                     bg=BG, fg=TEXT, font=("Segoe UI", 10, "bold")).pack(anchor="w")

        _h_sep(self).pack(fill="x")
        bf = tk.Frame(self, bg=CARD, padx=20, pady=12); bf.pack(fill="x")
        _icon_btn(bf, "✓  Confirm & Save Both", SUCCESS, SUCCESS_H, cmd=self._confirm).pack(side="left", padx=(0,8))
        _icon_btn(bf, "Cancel", DANGER, DANGER_H, cmd=self.destroy).pack(side="left")

    def _confirm(self): self.confirmed = True; self.destroy()


class ManualJournalDialog(tk.Toplevel):
    def __init__(self, parent, account_type, account_id, account_label, fy_year, refresh_cb=None):
        super().__init__(parent)
        self._account_type = account_type; self._account_id = account_id
        self._fy_year = fy_year; self._refresh = refresh_cb
        self.title(f"Manual Journal — {account_label}")
        self.geometry("480x380"); self.resizable(False, False)
        self.configure(bg=CARD); self.grab_set()

        hdr_bg = HEAD_E if account_type == "expenditure" else HEAD_I
        tk.Label(self, text=f"  Manual Journal Entry — {account_label}",
                 bg=hdr_bg, fg="white", font=("Segoe UI", 11, "bold"), pady=12).pack(fill="x")
        tk.Label(self, text=f"  {fy_label(fy_year)}  ·  Tally-style entry",
                 bg=BG, fg=TEXT2, font=("Segoe UI", 8), pady=5).pack(fill="x")

        frm = tk.Frame(self, bg=CARD, padx=28, pady=18); frm.pack(fill="both", expand=True)
        today_str = datetime.date.today().strftime("%d/%m/%Y")

        self._date_var = tk.StringVar(value=today_str)
        self._amt_var  = tk.StringVar()
        self._narr_var = tk.StringVar()
        self._type_var = tk.StringVar(value="Dr")

        fields = [("Date (DD/MM/YYYY)", self._date_var), ("Amount (Rs.)", self._amt_var), ("Narration", self._narr_var)]
        for r, (label, var) in enumerate(fields):
            _field_label(frm, label).grid(row=r, column=0, sticky="w", pady=7, padx=(0, 16))
            ttk.Entry(frm, textvariable=var, width=26).grid(row=r, column=1, sticky="w")

        _field_label(frm, "Entry Type").grid(row=3, column=0, sticky="w", pady=7)
        tf = tk.Frame(frm, bg=CARD); tf.grid(row=3, column=1, sticky="w")
        ttk.Radiobutton(tf, text="Dr  (Debit / bill raised)", variable=self._type_var, value="Dr").pack(anchor="w")
        ttk.Radiobutton(tf, text="Cr  (Credit / adjustment)", variable=self._type_var, value="Cr").pack(anchor="w")

        tk.Label(frm, text="Dr = raise a charge   ·   Cr = payment or correction",
                 bg=CARD, fg=TEXT3, font=("Segoe UI", 8, "italic")).grid(row=4, column=0, columnspan=2, sticky="w", pady=(4, 0))

        bf = tk.Frame(frm, bg=CARD); bf.grid(row=5, column=0, columnspan=2, pady=16)
        _icon_btn(bf, "Save Entry", SUCCESS, SUCCESS_H, cmd=self._save).pack(side="left", padx=(0, 8))
        _icon_btn(bf, "Cancel", DANGER, DANGER_H, cmd=self.destroy).pack(side="left")

    def _save(self):
        date = self._date_var.get().strip(); etype = self._type_var.get()
        narr = self._narr_var.get().strip(); amt_s = self._amt_var.get().strip()
        if not narr:
            messagebox.showerror("Missing", "Enter a narration.", parent=self); return
        try:
            amt = float(amt_s)
            if amt <= 0: raise ValueError
        except ValueError:
            messagebox.showerror("Invalid", "Enter a valid positive amount.", parent=self); return
        db_save_manual_journal(date, self._account_type, self._account_id, etype, amt, narr, self._fy_year)
        if self._refresh: self._refresh()
        self.destroy()


class NewExpAccountDialog(tk.Toplevel):
    def __init__(self, parent, refresh_cb=None):
        super().__init__(parent)
        self._refresh = refresh_cb
        self.title("New Expenditure Account")
        self.geometry("380x190"); self.resizable(False, False)
        self.configure(bg=CARD); self.grab_set()

        tk.Label(self, text="  Create New Expenditure Account",
                 bg=HEAD_E, fg="white", font=("Segoe UI", 11, "bold"), pady=12).pack(fill="x")
        frm = tk.Frame(self, bg=CARD, padx=28, pady=20); frm.pack(fill="both", expand=True)
        _field_label(frm, "Account Name").grid(row=0, column=0, sticky="w", pady=8)
        self._name_var = tk.StringVar()
        e = ttk.Entry(frm, textvariable=self._name_var, width=26)
        e.grid(row=0, column=1, sticky="w", padx=10); e.focus()
        tk.Label(frm, text="e.g. Water Charges, Lift Maintenance",
                 bg=CARD, fg=TEXT3, font=("Segoe UI", 8, "italic")).grid(row=1, column=0, columnspan=2, sticky="w")
        bf = tk.Frame(frm, bg=CARD); bf.grid(row=2, column=0, columnspan=2, pady=12)
        _icon_btn(bf, "Create", SUCCESS, SUCCESS_H, cmd=self._create).pack(side="left", padx=(0, 8))
        _icon_btn(bf, "Cancel", DANGER, DANGER_H, cmd=self.destroy).pack(side="left")

    def _create(self):
        name = self._name_var.get().strip()
        if not name:
            messagebox.showerror("Missing", "Enter an account name.", parent=self); return
        if db_add_expenditure_account(name):
            if self._refresh: self._refresh()
            self.destroy()


class LedgerWindow(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Accounts Ledger"); self.geometry("1300x740")
        self.configure(bg=BG); self.grab_set()
        self._selected_account = None
        self._fy_var = tk.StringVar(value=str(get_current_fy()))
        self._build(); self._load_account_list()

    def _build(self):
        top = tk.Frame(self, bg=SIDEBAR, height=58)
        top.pack(fill="x"); top.pack_propagate(False)
        tk.Label(top, text="  📊  Accounts Ledger", bg=SIDEBAR, fg="white",
                 font=("Segoe UI", 13, "bold")).pack(side="left", padx=18, pady=18)
        tk.Label(top, text="FY:", bg=SIDEBAR, fg=TEXT3, font=("Segoe UI", 10)).pack(side="left", padx=(24,4))
        ttk.Spinbox(top, from_=2020, to=2100, width=6, textvariable=self._fy_var,
                    font=("Segoe UI", 10)).pack(side="left")
        _icon_btn(top, "Load", ACCENT, ACCENT_H, cmd=self._reload_current).pack(side="left", padx=8)
        _icon_btn(top, "Master Ledger", TEAL, TEAL_H, cmd=self._open_master_ledger).pack(side="right", padx=6, pady=10)
        _icon_btn(top, "Export All Excel", SUCCESS, SUCCESS_H, cmd=self._export_all_excel).pack(side="right", padx=6, pady=10)
        _icon_btn(top, "Export All PDFs", PURPLE, PURPLE_H, cmd=self._export_all_pdfs).pack(side="right", padx=6, pady=10)
        _icon_btn(top, "Delete All", DANGER, DANGER_H, cmd=self._delete_entire).pack(side="right", padx=6, pady=10)

        info = tk.Frame(self, bg="#0A0C18", height=26)
        info.pack(fill="x"); info.pack_propagate(False)
        tk.Label(info, text="  Tally-style: only recorded entries shown  ·  Use '+ Manual Entry' to raise a bill  ·  Click Flat Accounts header for combined summary",
                 bg="#0A0C18", fg="#404870", font=("Segoe UI", 8, "italic")).pack(side="left", padx=14)

        body = tk.Frame(self, bg=BG); body.pack(fill="both", expand=True, padx=8, pady=8)
        self._build_account_list(body)
        tk.Frame(body, bg=BORDER2, width=1).pack(side="left", fill="y")
        self._build_right_panel(body)

    def _build_account_list(self, parent):
        lp = tk.Frame(parent, bg=CARD, highlightbackground=BORDER2, highlightthickness=1)
        lp.pack(side="left", fill="y")
        lh = tk.Frame(lp, bg=SIDEBAR, height=38)
        lh.pack(fill="x"); lh.pack_propagate(False)
        tk.Label(lh, text="  ACCOUNTS", bg=SIDEBAR, fg=TEXT3,
                 font=("Segoe UI", 8, "bold")).pack(side="left", pady=11, padx=12)

        chips = tk.Frame(lp, bg=BG); chips.pack(fill="x", padx=6, pady=6)
        self._chip_dr  = self._small_chip(chips, "Dr",  ERR_BG, ERR_FG)
        self._chip_cr  = self._small_chip(chips, "Cr",  OK_BG, OK_FG)
        self._chip_bal = self._small_chip(chips, "Net", PILL_BG, PILL_FG)

        cols = ("Account", "Balance")
        self._acct_tree = ttk.Treeview(lp, columns=cols, show="headings", height=28, selectmode="browse")
        self._acct_tree.heading("Account", text="Account")
        self._acct_tree.heading("Balance", text="Balance")
        self._acct_tree.column("Account", width=210, anchor="w")
        self._acct_tree.column("Balance", width=110, anchor="center")
        sb = ttk.Scrollbar(lp, orient="vertical", command=self._acct_tree.yview)
        self._acct_tree.configure(yscrollcommand=sb.set)
        self._acct_tree.pack(side="left", fill="both", padx=(4,0), pady=4)
        sb.pack(side="right", fill="y", pady=4)
        self._acct_tree.tag_configure("flat_hdr", background=HEAD_I, foreground="white", font=("Segoe UI",8,"bold"))
        self._acct_tree.tag_configure("exp_hdr",  background=HEAD_E, foreground="white", font=("Segoe UI",8,"bold"))
        self._acct_tree.tag_configure("dr_bal",  background=ERR_BG, foreground=ERR_FG)
        self._acct_tree.tag_configure("cr_bal",  background=OK_BG, foreground=OK_FG)
        self._acct_tree.tag_configure("nil_bal", background=BG, foreground=MUTED)
        self._acct_tree.bind("<<TreeviewSelect>>", self._on_account_select)

    def _small_chip(self, parent, label, bg, fg):
        f = tk.Frame(parent, bg=bg, padx=6, pady=3, highlightbackground=BORDER2, highlightthickness=1)
        f.pack(side="left", fill="x", expand=True, padx=2)
        tk.Label(f, text=label, bg=bg, fg=fg, font=("Segoe UI",7,"bold")).pack()
        lbl = tk.Label(f, text="-", bg=bg, fg=fg, font=("Segoe UI",9,"bold")); lbl.pack()
        return lbl

    def _build_right_panel(self, parent):
        rp = tk.Frame(parent, bg=CARD, highlightbackground=BORDER2, highlightthickness=1)
        rp.pack(side="left", fill="both", expand=True)
        self._detail_hdr = tk.Frame(rp, bg=HEAD_I, height=48)
        self._detail_hdr.pack(fill="x"); self._detail_hdr.pack_propagate(False)
        self._detail_title = tk.Label(self._detail_hdr, text="  Select an account",
                                       bg=HEAD_I, fg="white", font=("Segoe UI", 11, "bold"))
        self._detail_title.pack(side="left", padx=14, pady=14)
        self._detail_btns = tk.Frame(self._detail_hdr, bg=HEAD_I)
        self._detail_btns.pack(side="right", padx=8)
        self._bal_strip = tk.Frame(rp, bg=BG, height=34)
        self._bal_strip.pack(fill="x"); self._bal_strip.pack_propagate(False)
        self._bal_label = tk.Label(self._bal_strip, text="", bg=BG, fg=TEXT, font=("Segoe UI", 9))
        self._bal_label.pack(side="left", padx=14, pady=9)
        foot = tk.Frame(rp, bg=BG, height=28)
        foot.pack(fill="x", side="bottom"); foot.pack_propagate(False)
        tk.Label(foot, text="  Right-click manual entries to delete  ·  Double-click summary rows to drill in",
                 bg=BG, fg=MUTED, font=("Segoe UI", 8, "italic")).pack(side="left", pady=7, padx=12)
        self._zone = tk.Frame(rp, bg=CARD); self._zone.pack(fill="both", expand=True)
        self._ledger_view  = self._build_ledger_view(self._zone)
        self._summary_view = self._build_summary_view(self._zone)
        self._master_view  = self._build_master_view(self._zone)
        self._active_view  = None; self._switch_view("ledger")

    def _build_ledger_view(self, zone):
        f = tk.Frame(zone, bg=CARD)
        ir = tk.Frame(f, bg=BG, padx=10, pady=5); ir.pack(fill="x")
        tk.Label(ir, text=f"Dr = Debit (bills/charges raised)   ·   Cr = Credit (payments/receipts)   ·   Balance Dr = {OWNER_LABEL.lower()} owes   ·   Balance Cr = advance",
                 bg=BG, fg=TEXT2, font=("Segoe UI", 8)).pack(side="left")
        tw = tk.Frame(f, bg=CARD); tw.pack(fill="both", expand=True)
        cols = ("Date","Particulars / Narration","Vch Type","Vch No.","Debit (Rs.)","Credit (Rs.)","Balance")
        self._ledger_tree = ttk.Treeview(tw, columns=cols, show="headings", height=20)
        for c, w, anc in [("Date",88,"center"),("Particulars / Narration",360,"w"),("Vch Type",92,"center"),
                           ("Vch No.",112,"center"),("Debit (Rs.)",96,"e"),("Credit (Rs.)",96,"e"),("Balance",122,"center")]:
            self._ledger_tree.heading(c, text=c); self._ledger_tree.column(c, width=w, anchor=anc)
        vsb = ttk.Scrollbar(tw, orient="vertical", command=self._ledger_tree.yview)
        self._ledger_tree.configure(yscrollcommand=vsb.set)
        self._ledger_tree.pack(side="left", fill="both", expand=True, padx=(8,0), pady=(2,4))
        vsb.pack(side="right", fill="y", pady=4)
        self._ledger_tree.tag_configure("ob_row",     background=PILL_BG, foreground=TEXT, font=("Segoe UI",9,"bold"))
        self._ledger_tree.tag_configure("dr_row",     background=ERR_BG,  foreground=ERR_FG)
        self._ledger_tree.tag_configure("cr_row",     background=OK_BG,   foreground=OK_FG)
        self._ledger_tree.tag_configure("manual_row", background=WARN_BG, foreground=WARN_FG)
        self._ledger_tree.tag_configure("total_row",  background=SIDEBAR, foreground="white", font=("Segoe UI",9,"bold"))
        self._ledger_tree.tag_configure("cb_row",     background=HEAD_I,  foreground="white", font=("Segoe UI",9,"bold"))
        self._ledger_tree.bind("<Button-3>", self._ledger_rclick)
        return f

    def _build_summary_view(self, zone):
        f = tk.Frame(zone, bg=CARD)
        ir = tk.Frame(f, bg=BG, padx=10, pady=5); ir.pack(fill="x")
        tk.Label(ir, text="Combined summary of ALL flat accounts  ·  Double-click any row to open its full ledger",
                 bg=BG, fg=TEXT2, font=("Segoe UI", 8)).pack(side="left")
        tw = tk.Frame(f, bg=CARD); tw.pack(fill="both", expand=True)
        cols = (UNIT_LABEL,OWNER_LABEL,"Opening Balance","Total Dr (Bills/Journals)","Total Cr (Payments)","Closing Balance","Status")
        self._summary_tree = ttk.Treeview(tw, columns=cols, show="headings", height=20)
        for c, w, anc in [(UNIT_LABEL,55,"center"),(OWNER_LABEL,185,"w"),("Opening Balance",120,"center"),
                           ("Total Dr (Bills/Journals)",145,"center"),("Total Cr (Payments)",130,"center"),
                           ("Closing Balance",120,"center"),("Status",85,"center")]:
            self._summary_tree.heading(c, text=c); self._summary_tree.column(c, width=w, anchor=anc)
        vsb = ttk.Scrollbar(tw, orient="vertical", command=self._summary_tree.yview)
        self._summary_tree.configure(yscrollcommand=vsb.set)
        self._summary_tree.pack(side="left", fill="both", expand=True, padx=(8,0), pady=(2,4))
        vsb.pack(side="right", fill="y", pady=4)
        self._summary_tree.tag_configure("dr_row",    background=ERR_BG, foreground=ERR_FG)
        self._summary_tree.tag_configure("cr_row",    background=OK_BG,  foreground=OK_FG)
        self._summary_tree.tag_configure("nil_row",   background=ROW_ODD, foreground=TEXT2)
        self._summary_tree.tag_configure("total_row", background=SIDEBAR, foreground="white", font=("Segoe UI",10,"bold"))
        self._summary_tree.bind("<Double-1>", self._summary_dbl_click)
        return f

    def _build_master_view(self, zone):
        f = tk.Frame(zone, bg=CARD)
        ir = tk.Frame(f, bg=BG, padx=10, pady=5); ir.pack(fill="x")
        tk.Label(ir, text="Master Ledger: combined balance of ALL accounts  ·  Double-click any row to drill in",
                 bg=BG, fg=TEXT2, font=("Segoe UI", 8)).pack(side="left")
        tw = tk.Frame(f, bg=CARD); tw.pack(fill="both", expand=True)
        cols = ("Account","Account Type","Opening Bal","Total Dr (Rs.)","Total Cr (Rs.)","Closing Bal","Net Position")
        self._master_tree = ttk.Treeview(tw, columns=cols, show="headings", height=17)
        for c, w, anc in [("Account",200,"w"),("Account Type",95,"center"),("Opening Bal",110,"center"),
                           ("Total Dr (Rs.)",110,"center"),("Total Cr (Rs.)",110,"center"),
                           ("Closing Bal",120,"center"),("Net Position",105,"center")]:
            self._master_tree.heading(c, text=c); self._master_tree.column(c, width=w, anchor=anc)
        vsb = ttk.Scrollbar(tw, orient="vertical", command=self._master_tree.yview)
        self._master_tree.configure(yscrollcommand=vsb.set)
        self._master_tree.pack(side="left", fill="both", expand=True, padx=(8,0), pady=(2,0))
        vsb.pack(side="right", fill="y")
        self._master_tree.tag_configure("section_hdr", background=HEAD_I,  foreground="white", font=("Segoe UI",9,"bold"))
        self._master_tree.tag_configure("exp_hdr",     background=HEAD_E,  foreground="white", font=("Segoe UI",9,"bold"))
        self._master_tree.tag_configure("flat_row",    background=ROW_ODD, foreground=TEXT)
        self._master_tree.tag_configure("exp_row",     background=ERR_BG,  foreground=ERR_FG)
        self._master_tree.tag_configure("subtotal_row",background=PILL_BG, foreground=TEXT, font=("Segoe UI",9,"bold"))
        self._master_tree.tag_configure("total_row",   background=SIDEBAR, foreground="white", font=("Segoe UI",10,"bold"))
        self._master_tree.bind("<Double-1>", self._master_dbl_click)
        self._mchips = tk.Frame(f, bg=CARD); self._mchips.pack(fill="x", padx=8, pady=(4,6))
        return f

    def _switch_view(self, view):
        for frm in (self._ledger_view, self._summary_view, self._master_view): frm.pack_forget()
        if view == "ledger": self._ledger_view.pack(fill="both", expand=True)
        elif view == "summary": self._summary_view.pack(fill="both", expand=True)
        elif view == "master": self._master_view.pack(fill="both", expand=True)
        self._active_view = view

    def _load_account_list(self):
        for i in self._acct_tree.get_children(): self._acct_tree.delete(i)
        try: fy = int(self._fy_var.get())
        except Exception: fy = get_current_fy()
        total_dr_all = 0.0; total_cr_all = 0.0
        self._acct_tree.insert("","end",iid="_hdr_flats",values=(f"── {UNIT_LABEL.upper()} ACCOUNTS ──",""),tags=("flat_hdr",))
        for flat_no, res in RESIDENTS.items():
            acct = get_flat_account_entries(flat_no, fy); cb = acct["closing_balance"]
            total_dr_all += acct["total_dr"]; total_cr_all += acct["total_cr"]
            if abs(cb) < 0.01: bal_str = "Rs.0  (Nil)"; tag = "nil_bal"
            elif cb > 0: bal_str = f"Rs.{cb:,.0f} Dr"; tag = "dr_bal"
            else: bal_str = f"Rs.{abs(cb):,.0f} Cr"; tag = "cr_bal"
            self._acct_tree.insert("","end",iid=f"flat_{flat_no}",
                                    values=(f"  {flat_no}  {res['name'][:16]}", bal_str), tags=(tag,))
        adm_rows = db_fetch_admission_fees()
        adm_total = sum(r[5] for r in adm_rows)
        total_cr_all += adm_total
        adm_tag = "cr_bal" if adm_total > 0 else "nil_bal"
        self._acct_tree.insert("","end",iid="_hdr_adm",
                                values=("── ADMISSION FEES ──",""),tags=("flat_hdr",))
        self._acct_tree.insert("","end",iid="_adm_account",
                                values=("  Admission Fees", f"Rs.{adm_total:,.0f} Cr" if adm_total>0 else "Rs.0  (Nil)"),
                                tags=(adm_tag,))
        exp_accounts = db_get_expenditure_accounts()
        self._acct_tree.insert("","end",iid="_hdr_exp",values=("── EXPENDITURE ──",""),tags=("exp_hdr",))
        for aid, aname in exp_accounts:
            exp_acct = get_expenditure_account_entries(fy, aid); exp_cb = exp_acct["closing_balance"]
            exp_tag = ("dr_bal" if exp_cb > 0 else "cr_bal" if exp_cb < 0 else "nil_bal")
            exp_bal = (f"Rs.{exp_cb:,.0f} Dr" if exp_cb > 0 else f"Rs.{abs(exp_cb):,.0f} Cr" if exp_cb < 0 else "Rs.0  (Nil)")
            self._acct_tree.insert("","end",iid=f"exp_{aid}",
                                    values=(f"  {aname[:18]}", exp_bal), tags=(exp_tag,))
        self._acct_tree.insert("","end",iid="_new_exp_acct",values=("  + New Account",""),tags=("nil_bal",))
        net = total_cr_all - total_dr_all
        self._chip_dr.configure(text=f"{total_dr_all:,.0f}")
        self._chip_cr.configure(text=f"{total_cr_all:,.0f}")
        self._chip_bal.configure(text=f"{'Surplus' if net>=0 else 'Deficit'} {abs(net):,.0f}",
                                  fg=OK_FG if net>=0 else ERR_FG)

    def _on_account_select(self, _=None):
        sel = self._acct_tree.selection()
        if not sel: return
        iid = sel[0]
        try: fy = int(self._fy_var.get())
        except Exception: fy = get_current_fy()
        if iid == "_hdr_flats": self._selected_account = ("summary","ALL_FLATS"); self._load_flat_summary(fy)
        elif iid in ("_hdr_exp", "_hdr_adm"): pass
        elif iid in ("_adm_account",): self._selected_account = ("admission","ADM"); self._load_admission_account()
        elif iid == "_new_exp_acct": NewExpAccountDialog(self, refresh_cb=self._reload_current)
        elif iid.startswith("flat_"):
            flat_no = iid[5:]; self._selected_account = ("flat", flat_no); self._load_flat_account(flat_no, fy)
        elif iid.startswith("exp_"):
            try: aid = int(iid[4:])
            except Exception: return
            accounts = {a[0]: a[1] for a in db_get_expenditure_accounts()}
            aname = accounts.get(aid, f"Account {aid}")
            self._selected_account = ("expenditure", str(aid), aname)
            self._load_expenditure_account(fy, aid, aname)

    def _load_flat_summary(self, fy):
        self._switch_view("summary")
        self._detail_hdr.configure(bg=HEAD_I); self._detail_title.configure(bg=HEAD_I,
            text=f"  All Flat Accounts — Summary  ·  {fy_label(fy)}")
        self._detail_btns.configure(bg=HEAD_I)
        for w in self._detail_btns.winfo_children(): w.destroy()
        _icon_btn(self._detail_btns, "+ Manual Entry", WARNING, WARNING_H,
                  cmd=lambda: self._manual_entry_any_flat(fy)).pack(side="left", padx=3)
        _icon_btn(self._detail_btns, "Export Excel", SUCCESS, SUCCESS_H,
                  cmd=self._export_all_excel).pack(side="left", padx=3)
        _icon_btn(self._detail_btns, "Export All PDFs", ACCENT, ACCENT_H,
                  cmd=self._export_all_pdfs).pack(side="left", padx=3)
        for i in self._summary_tree.get_children(): self._summary_tree.delete(i)
        total_ob = total_dr = total_cr = total_cb = 0.0
        for flat_no, res in RESIDENTS.items():
            acct = get_flat_account_entries(flat_no, fy)
            ob = acct["opening_balance"]; dr = acct["total_dr"]; cr = acct["total_cr"]; cb = acct["closing_balance"]
            total_ob += ob; total_dr += dr; total_cr += cr; total_cb += cb
            ob_s = (f"Rs.{abs(ob):,.0f} {'Dr' if ob>0 else 'Cr'}" if ob!=0 else "Nil")
            cb_s = (f"Rs.{abs(cb):,.0f} {'Dr' if cb>0 else 'Cr'}" if cb!=0 else "Nil")
            status = "Due" if cb > 0 else ("Advance" if cb < 0 else "Clear")
            tag = "dr_row" if cb > 0 else ("cr_row" if cb < 0 else "nil_row")
            self._summary_tree.insert("","end",iid=f"s_{flat_no}",
                                       values=(flat_no, res["name"], ob_s, f"Rs.{dr:,.0f}", f"Rs.{cr:,.0f}", cb_s, status), tags=(tag,))
        tcb_s = (f"Rs.{abs(total_cb):,.0f} {'Dr' if total_cb>0 else 'Cr'}" if total_cb!=0 else "Nil")
        self._summary_tree.insert("","end",iid="s_total",
                                   values=("ALL","TOTAL",f"Rs.{abs(total_ob):,.0f}",
                                           f"Rs.{total_dr:,.0f}",f"Rs.{total_cr:,.0f}",tcb_s,""), tags=("total_row",))
        self._bal_strip.configure(bg=PILL_BG)
        self._bal_label.configure(text=f"  {len(RESIDENTS)} {UNIT_LABEL}s  ·  Total Dr: Rs.{total_dr:,.0f}  ·  Total Cr: Rs.{total_cr:,.0f}  ·  Net: {tcb_s}",
                                   bg=PILL_BG, fg=TEXT)

    def _load_flat_account(self, flat_no, fy):
        self._switch_view("ledger")
        acct = get_flat_account_entries(flat_no, fy); owner = RESIDENTS.get(flat_no,{}).get("name", flat_no)
        self._detail_hdr.configure(bg=HEAD_I); self._detail_title.configure(bg=HEAD_I,
            text=f"  Flat {flat_no}  —  {owner}  ·  {fy_label(fy)}")
        self._detail_btns.configure(bg=HEAD_I)
        for w in self._detail_btns.winfo_children(): w.destroy()
        _icon_btn(self._detail_btns, "+ Manual Entry", WARNING, WARNING_H,
                  cmd=lambda: ManualJournalDialog(self,"flat",flat_no,f"Flat {flat_no}",fy,refresh_cb=self._reload_current)
                  ).pack(side="left", padx=3)
        _icon_btn(self._detail_btns, "Export Excel", SUCCESS, SUCCESS_H,
                  cmd=lambda: self._export_flat_excel(flat_no, fy)).pack(side="left", padx=3)
        _icon_btn(self._detail_btns, "Export PDF", ACCENT, ACCENT_H,
                  cmd=lambda: self._export_flat_pdf(flat_no, fy)).pack(side="left", padx=3)
        cb = acct["closing_balance"]; ob = acct["opening_balance"]
        bal_color = ERR_FG if cb > 0 else OK_FG
        cb_str = f"Rs.{abs(cb):,.0f} {'Dr (owner owes)' if cb>0 else 'Cr (advance)'}"
        ob_str = f"Opening: Rs.{abs(ob):,.0f} {'Dr' if ob>0 else 'Cr'}" if ob!=0 else "Opening: Nil"
        strip_bg = ERR_BG if cb>0 else OK_BG
        self._bal_strip.configure(bg=strip_bg)
        self._bal_label.configure(text=f"  {ob_str}    ·    Closing: {cb_str}    ·    Dr: Rs.{acct['total_dr']:,.0f}    Cr: Rs.{acct['total_cr']:,.0f}",
                                   bg=strip_bg, fg=bal_color)
        self._populate_ledger(acct, "flat")

    def _load_expenditure_account(self, fy, account_id=1, account_name="Society Expenditure"):
        self._switch_view("ledger")
        acct = get_expenditure_account_entries(fy, account_id)
        self._detail_hdr.configure(bg=HEAD_E); self._detail_title.configure(bg=HEAD_E,
            text=f"  {account_name}  ·  {fy_label(fy)}")
        self._detail_btns.configure(bg=HEAD_E)
        for w in self._detail_btns.winfo_children(): w.destroy()
        _icon_btn(self._detail_btns, "+ Manual Entry", WARNING, WARNING_H,
                  cmd=lambda aid=account_id, an=account_name: ManualJournalDialog(
                      self,"expenditure",str(aid),an,fy,refresh_cb=self._reload_current)
                  ).pack(side="left", padx=3)
        _icon_btn(self._detail_btns, "Export Excel", SUCCESS, SUCCESS_H,
                  cmd=lambda aid=account_id, an=account_name: self._export_exp_excel(fy, aid, an)).pack(side="left", padx=3)
        _icon_btn(self._detail_btns, "Export PDF", ACCENT, ACCENT_H,
                  cmd=lambda aid=account_id, an=account_name: self._export_exp_pdf(fy, aid, an)).pack(side="left", padx=3)
        if account_id != 1:
            _icon_btn(self._detail_btns, "Delete Account", DANGER, DANGER_H,
                      cmd=lambda aid=account_id: self._delete_exp_account(aid)).pack(side="left", padx=3)
        cb = acct["closing_balance"]; ob = acct["opening_balance"]
        bal_color = ERR_FG if cb > 0 else OK_FG
        cb_str = f"Rs.{abs(cb):,.0f} {'Dr (net expenditure)' if cb>0 else 'Cr (net surplus)'}"
        ob_str = f"Opening: Rs.{abs(ob):,.0f} {'Dr' if ob>0 else 'Cr'}" if ob!=0 else "Opening: Nil"
        strip_bg = ERR_BG if cb>0 else OK_BG
        self._bal_strip.configure(bg=strip_bg)
        self._bal_label.configure(text=f"  {ob_str}    ·    Closing: {cb_str}    ·    Dr: Rs.{acct['total_dr']:,.0f}    Cr: Rs.{acct['total_cr']:,.0f}",
                                   bg=strip_bg, fg=bal_color)
        self._populate_ledger(acct, "expenditure")

    def _load_admission_account(self):
        self._switch_view("ledger")
        rows = db_fetch_admission_fees()
        total = sum(r[5] for r in rows)
        self._detail_hdr.configure(bg=HEAD_I)
        self._detail_title.configure(bg=HEAD_I, text="  💰  Admission Fees Account")
        self._detail_btns.configure(bg=HEAD_I)
        for w in self._detail_btns.winfo_children(): w.destroy()
        self._bal_strip.configure(bg=OK_BG)
        self._bal_label.configure(bg=OK_BG, fg=OK_FG,
            text=f"  Total Records: {len(rows)}    ·    Total Collected (Cr): Rs.{total:,.0f}")
        for i in self._ledger_tree.get_children(): self._ledger_tree.delete(i)
        running = 0.0
        for r in reversed(rows):
            _, rno, date, flat, owner, amt, mob, _ = r
            running += amt
            self._ledger_tree.insert("","end",
                values=(date, f"Admission Fee  —  {UNIT_LABEL} {flat}  ({owner})",
                        "Receipt", rno, "", f"Rs.{amt:,.0f}", f"Rs.{running:,.0f} Cr"),
                tags=("cr_row",))
        self._ledger_tree.insert("","end",
            values=("","TOTAL","","","",f"Rs.{total:,.0f}",f"Rs.{total:,.0f} Cr"),
            tags=("total_row",))

    def _delete_exp_account(self, aid):
        if messagebox.askyesno("Delete Account", "Permanently delete this account?", parent=self):
            if db_delete_expenditure_account(aid):
                self._selected_account = None; self._reload_current()

    def _open_master_ledger(self):
        try: fy = int(self._fy_var.get())
        except Exception: fy = get_current_fy()
        self._selected_account = ("master","ALL"); self._load_master_ledger(fy)

    def _load_master_ledger(self, fy):
        self._switch_view("master")
        self._detail_hdr.configure(bg=SIDEBAR); self._detail_title.configure(bg=SIDEBAR,
            text=f"  Master Ledger — All Accounts  ·  {fy_label(fy)}")
        self._detail_btns.configure(bg=SIDEBAR)
        for w in self._detail_btns.winfo_children(): w.destroy()
        _icon_btn(self._detail_btns, "Export Excel", SUCCESS, SUCCESS_H,
                  cmd=self._export_all_excel).pack(side="left", padx=3)
        _icon_btn(self._detail_btns, "Export PDF", PURPLE, PURPLE_H,
                  cmd=self._export_all_pdfs).pack(side="left", padx=3)
        for i in self._master_tree.get_children(): self._master_tree.delete(i)
        for w in self._mchips.winfo_children(): w.destroy()
        self._master_tree.insert("","end",iid="_mhdr_flats",values=(f"── {UNIT_LABEL.upper()} ACCOUNTS ──","","","","","",""),tags=("section_hdr",))
        f_ob=f_dr=f_cr=f_cb=0.0
        for flat_no, res in RESIDENTS.items():
            acct = get_flat_account_entries(flat_no, fy)
            ob=acct["opening_balance"]; dr=acct["total_dr"]; cr=acct["total_cr"]; cb=acct["closing_balance"]
            f_ob+=ob; f_dr+=dr; f_cr+=cr; f_cb+=cb
            ob_s = f"Rs.{abs(ob):,.0f} {'Dr' if ob>0 else 'Cr'}" if ob!=0 else "Nil"
            cb_s = f"Rs.{abs(cb):,.0f} {'Dr' if cb>0 else 'Cr'}" if cb!=0 else "Nil"
            net = "Due" if cb>0 else ("Advance" if cb<0 else "Clear")
            self._master_tree.insert("","end",iid=f"m_{flat_no}",
                                      values=(f"  {flat_no}  {res['name'][:24]}","Maintenance",ob_s,
                                              f"Rs.{dr:,.2f}",f"Rs.{cr:,.2f}",cb_s,net), tags=("flat_row",))
        ftcb_s = f"Rs.{abs(f_cb):,.0f} {'Dr' if f_cb>0 else 'Cr'}" if f_cb!=0 else "Nil"
        self._master_tree.insert("","end",iid="_msub_flats",
                                  values=("  Sub-Total (Flat Accounts)","",f"Rs.{abs(f_ob):,.0f}",
                                          f"Rs.{f_dr:,.2f}",f"Rs.{f_cr:,.2f}",ftcb_s,""), tags=("subtotal_row",))
        self._master_tree.insert("","end",iid="_mhdr_exp",values=("── EXPENDITURE ACCOUNTS ──","","","","","",""),tags=("exp_hdr",))
        exp_accounts = db_get_expenditure_accounts(); total_e_dr=total_e_cr=0.0
        for aid, aname in exp_accounts:
            ea=get_expenditure_account_entries(fy, aid)
            e_ob=ea["opening_balance"]; e_dr=ea["total_dr"]; e_cr=ea["total_cr"]; e_cb=ea["closing_balance"]
            total_e_dr+=e_dr; total_e_cr+=e_cr
            eob_s=f"Rs.{abs(e_ob):,.0f} {'Dr' if e_ob>0 else 'Cr'}" if e_ob!=0 else "Nil"
            ecb_s=f"Rs.{abs(e_cb):,.0f} {'Dr' if e_cb>0 else 'Cr'}" if e_cb!=0 else "Nil"
            net_e="Net Expense" if e_cb>0 else "Net Surplus" if e_cb<0 else "Balanced"
            self._master_tree.insert("","end",iid=f"m_exp_{aid}",
                                      values=(f"  {aname[:28]}","Expenditure",eob_s,
                                              f"Rs.{e_dr:,.2f}",f"Rs.{e_cr:,.2f}",ecb_s,net_e), tags=("exp_row",))
        g_dr=f_dr+total_e_dr; g_cr=f_cr+total_e_cr; g_net=f_cr-total_e_dr
        gcb_s=f"Rs.{abs(g_net):,.0f} {'Surplus' if g_net>=0 else 'Deficit'}" if g_net!=0 else "Balanced"
        self._master_tree.insert("","end",iid="_mgrand",
                                  values=("GRAND TOTAL","","",f"Rs.{g_dr:,.2f}",f"Rs.{g_cr:,.2f}",gcb_s,"Net Position"), tags=("total_row",))
        for label, val, bg, fg in [("Income Collected",f"Rs.{f_cr:,.0f}",OK_BG,OK_FG),
                                    ("Total Expenses",f"Rs.{total_e_dr:,.0f}",ERR_BG,ERR_FG),
                                    ("Bills/Dr Journals",f"Rs.{f_dr:,.0f}",WARN_BG,WARN_FG),
                                    ("Net Surplus/Deficit",gcb_s,PILL_BG,ACCENT)]:
            ff = tk.Frame(self._mchips, bg=bg, padx=12, pady=6, highlightbackground=BORDER2, highlightthickness=1)
            ff.pack(side="left", fill="x", expand=True, padx=(0,6))
            tk.Label(ff, text=label, bg=bg, fg=fg, font=("Segoe UI",8,"bold")).pack()
            tk.Label(ff, text=val,   bg=bg, fg=fg, font=("Segoe UI",11,"bold")).pack()
        self._bal_strip.configure(bg=SIDEBAR)
        self._bal_label.configure(text=f"  {fy_label(fy)}  ·  Bills: Rs.{f_dr:,.0f}  ·  Expenses: Rs.{total_e_dr:,.0f}  ·  Income: Rs.{f_cr:,.0f}  ·  Net: {gcb_s}",
                                   bg=SIDEBAR, fg="white")

    def _populate_ledger(self, acct, kind):
        for i in self._ledger_tree.get_children(): self._ledger_tree.delete(i)
        fy = acct["fy_year"]; ob = acct["opening_balance"]
        ob_bal_str = f"Rs.{abs(ob):,.0f} {'Dr' if ob>0 else 'Cr'}" if ob!=0 else "Nil  (Start)"
        self._ledger_tree.insert("","end",iid="_ob",
                                  values=("—",f"Opening Balance — b/f from {fy_label(fy-1)}","b/f","—",
                                          f"{ob:.2f}" if ob>0 else "",f"{abs(ob):.2f}" if ob<0 else "",ob_bal_str), tags=("ob_row",))
        for idx, entry in enumerate(acct["entries"]):
            bal=entry["balance"]
            bal_str=f"Rs.{abs(bal):,.0f} {'Dr' if bal>0 else ('Cr' if bal<0 else 'Nil')}"
            tag=("manual_row" if entry["source"]=="manual" else "dr_row" if entry["type"]=="Dr" else "cr_row")
            self._ledger_tree.insert("","end",iid=f"entry_{idx}_{entry['source_id']}",
                                      values=(entry["date"],entry["narration"],entry.get("vch_type",""),
                                              entry.get("vch_no",""),
                                              f"{entry['dr_amt']:,.2f}" if entry["dr_amt"] else "",
                                              f"{entry['cr_amt']:,.2f}" if entry["cr_amt"] else "",bal_str), tags=(tag,))
        self._ledger_tree.insert("","end",iid="_total",
                                  values=("","TOTAL","","",f"{acct['total_dr']:,.2f}",f"{acct['total_cr']:,.2f}",""), tags=("total_row",))
        cb=acct["closing_balance"]
        cb_str=f"Rs.{abs(cb):,.0f} {'Dr' if cb>0 else ('Cr' if cb<0 else 'Nil')}"
        self._ledger_tree.insert("","end",iid="_cb",
                                  values=("c/f",f"Closing Balance — c/f to {fy_label(fy+1)}","c/f","—",
                                          f"{cb:.2f}" if cb>0 else "",f"{abs(cb):.2f}" if cb<0 else "",cb_str), tags=("cb_row",))

    def _summary_dbl_click(self, event):
        item = self._summary_tree.identify_row(event.y)
        if not item or item == "s_total": return
        flat_no = item[2:]
        if flat_no not in RESIDENTS: return
        try: fy = int(self._fy_var.get())
        except Exception: fy = get_current_fy()
        self._acct_tree.selection_set(f"flat_{flat_no}")
        self._selected_account = ("flat", flat_no); self._load_flat_account(flat_no, fy)

    def _master_dbl_click(self, event):
        item = self._master_tree.identify_row(event.y)
        if not item or item.startswith("_"): return
        try: fy = int(self._fy_var.get())
        except Exception: fy = get_current_fy()
        if item.startswith("m_") and not item.startswith("m_exp_"):
            flat_no = item[2:]
            if flat_no in RESIDENTS:
                self._acct_tree.selection_set(f"flat_{flat_no}")
                self._selected_account = ("flat", flat_no); self._load_flat_account(flat_no, fy)
        elif item.startswith("m_exp_"):
            try: aid = int(item[6:])
            except Exception: return
            accounts = {a[0]: a[1] for a in db_get_expenditure_accounts()}
            aname = accounts.get(aid, f"Account {aid}")
            self._acct_tree.selection_set(f"exp_{aid}")
            self._selected_account = ("expenditure", str(aid), aname)
            self._load_expenditure_account(fy, aid, aname)

    def _manual_entry_any_flat(self, fy):
        dlg = tk.Toplevel(self); dlg.title("Select Flat")
        dlg.geometry("300x175"); dlg.resizable(False,False); dlg.configure(bg=CARD); dlg.grab_set()
        tk.Label(dlg, text="  Select Flat for Manual Entry", bg=HEAD_I, fg="white",
                 font=("Segoe UI",10,"bold"), pady=10).pack(fill="x")
        tk.Label(dlg, text="Choose flat:", bg=CARD, fg=TEXT, font=("Segoe UI",10), pady=8).pack(padx=16, anchor="w")
        flat_var = tk.StringVar(value=list(RESIDENTS.keys())[0])
        ttk.Combobox(dlg, textvariable=flat_var, values=list(RESIDENTS.keys()),
                     state="readonly", width=22).pack(padx=16, pady=4)
        def _open():
            f = flat_var.get(); owner = RESIDENTS.get(f,{}).get("name",f); dlg.destroy()
            ManualJournalDialog(self,"flat",f,f"Flat {f} — {owner}",fy,refresh_cb=self._reload_current)
        bf = tk.Frame(dlg, bg=CARD); bf.pack(pady=10)
        _icon_btn(bf, "Open", ACCENT, ACCENT_H, cmd=_open).pack(side="left", padx=6)
        _icon_btn(bf, "Cancel", DANGER, DANGER_H, cmd=dlg.destroy).pack(side="left", padx=6)

    def _ledger_rclick(self, event):
        item = self._ledger_tree.identify_row(event.y)
        if not item or item.startswith("_") or not item.startswith("entry_"): return
        try: fy = int(self._fy_var.get())
        except Exception: return
        sel = self._acct_tree.selection()
        if not sel: return
        iid = sel[0]
        if iid.startswith("flat_"): acct = get_flat_account_entries(iid[5:], fy)
        elif iid.startswith("exp_"):
            try: aid = int(iid[4:])
            except Exception: return
            acct = get_expenditure_account_entries(fy, aid)
        else: return
        try:
            idx = int(item.split("_")[1]); entry = acct["entries"][idx]
        except Exception: return
        m = tk.Menu(self, tearoff=0)
        if entry["source"] not in ("manual",):
            m.add_command(label="(Cannot delete payment/expenditure entries here)", state="disabled")
        else:
            jid = int(entry["source_id"])
            m.add_command(label=f"Delete: {entry['narration'][:42]}", command=lambda: self._del_manual(jid))
        m.post(event.x_root, event.y_root)

    def _del_manual(self, jid):
        if messagebox.askyesno("Delete", "Delete this journal entry?", parent=self):
            db_delete_manual_journal(jid); self._reload_current()

    def _reload_current(self):
        self._load_account_list()
        if not self._selected_account: return
        atype = self._selected_account[0]
        try: fy = int(self._fy_var.get())
        except Exception: fy = get_current_fy()
        if atype == "flat":
            aid = self._selected_account[1]
            try: self._acct_tree.selection_set(f"flat_{aid}")
            except Exception: pass
            self._load_flat_account(aid, fy)
        elif atype == "expenditure":
            aid_str = self._selected_account[1]
            aname = self._selected_account[2] if len(self._selected_account)>2 else "Expenditure"
            try: self._acct_tree.selection_set(f"exp_{int(aid_str)}")
            except Exception: pass
            self._load_expenditure_account(fy, int(aid_str), aname)
        elif atype == "summary":
            try: self._acct_tree.selection_set("_hdr_flats")
            except Exception: pass
            self._load_flat_summary(fy)
        elif atype == "master": self._load_master_ledger(fy)

    # ── Excel export methods ────────────────────────────────────────────────
    def _export_flat_excel(self, flat_no, fy):
        fp = filedialog.asksaveasfilename(
            parent=self, defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"Ledger_Flat{flat_no}_{fy_label(fy).replace(' ','_')}.xlsx")
        if fp and generate_flat_account_excel(flat_no, fy, fp):
            messagebox.showinfo("Saved", f"Saved:\n{fp}", parent=self)
            if messagebox.askyesno("Open?", "Open in Excel now?", parent=self):
                webbrowser.open(f"file:///{fp}")

    def _export_exp_excel(self, fy, account_id=1, account_name="Society Expenditure"):
        safe = account_name.replace(" ","_").replace("/","-")
        fp = filedialog.asksaveasfilename(
            parent=self, defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"Ledger_Exp_{safe}_{fy_label(fy).replace(' ','_')}.xlsx")
        if fp and generate_expenditure_account_excel(fy, fp, account_id, account_name):
            messagebox.showinfo("Saved", f"Saved:\n{fp}", parent=self)
            if messagebox.askyesno("Open?", "Open in Excel now?", parent=self):
                webbrowser.open(f"file:///{fp}")

    def _export_all_excel(self):
        try: fy = int(self._fy_var.get())
        except Exception: fy = get_current_fy()
        fp = filedialog.asksaveasfilename(
            parent=self, defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=f"MasterLedger_{fy_label(fy).replace(' ','_')}.xlsx")
        if fp and generate_all_ledger_excel(fy, fp):
            messagebox.showinfo("Saved",
                f"Master Excel saved:\n{fp}\n\n"
                f"Sheets: Summary + {len(RESIDENTS)} flat(s) + expenditure account(s)", parent=self)
            if messagebox.askyesno("Open?", "Open in Excel now?", parent=self):
                webbrowser.open(f"file:///{fp}")

    # ── PDF export methods (unchanged) ──────────────────────────────────────
    def _export_flat_pdf(self, flat_no, fy):
        fp = filedialog.asksaveasfilename(parent=self, defaultextension=".pdf",
                                          filetypes=[("PDF files","*.pdf")],
                                          initialfile=f"Account_Flat{flat_no}_{fy_label(fy).replace(' ','_')}.pdf")
        if fp and generate_flat_account_pdf(flat_no, fy, fp):
            messagebox.showinfo("Saved", f"Saved:\n{fp}", parent=self)
            if messagebox.askyesno("Open?", "Open now?", parent=self): webbrowser.open(f"file:///{fp}")

    def _export_exp_pdf(self, fy, account_id=1, account_name="Society Expenditure"):
        safe = account_name.replace(" ","_").replace("/","-")
        fp = filedialog.asksaveasfilename(parent=self, defaultextension=".pdf",
                                          filetypes=[("PDF files","*.pdf")],
                                          initialfile=f"Expenditure_{safe}_{fy_label(fy).replace(' ','_')}.pdf")
        if fp and generate_expenditure_account_pdf(fy, fp, account_id, account_name):
            messagebox.showinfo("Saved", f"Saved:\n{fp}", parent=self)
            if messagebox.askyesno("Open?", "Open now?", parent=self): webbrowser.open(f"file:///{fp}")

    def _export_all_pdfs(self):
        try: fy = int(self._fy_var.get())
        except Exception: fy = get_current_fy()
        folder = filedialog.askdirectory(parent=self, title="Choose folder for PDFs")
        if not folder: return
        count = 0
        for flat_no in RESIDENTS:
            fp = os.path.join(folder, f"Account_Flat{flat_no}_{fy_label(fy).replace(' ','_')}.pdf")
            if generate_flat_account_pdf(flat_no, fy, fp): count += 1
        for aid, aname in db_get_expenditure_accounts():
            safe = aname.replace(" ","_").replace("/","-")
            fp_exp = os.path.join(folder, f"Expenditure_{safe}_{fy_label(fy).replace(' ','_')}.pdf")
            if generate_expenditure_account_pdf(fy, fp_exp, aid, aname): count += 1
        messagebox.showinfo("Done", f"Exported {count} PDFs to:\n{folder}", parent=self)

    def _delete_entire(self):
        if not messagebox.askyesno("Delete Entire Ledger",
                                   "PERMANENTLY DELETE ALL records?\n\nThis CANNOT be undone.",
                                   icon="warning", parent=self): return
        dlg = tk.Toplevel(self); dlg.title("Final Confirmation"); dlg.geometry("360x175")
        dlg.resizable(False,False); dlg.configure(bg=CARD); dlg.grab_set()
        tk.Label(dlg, text="FINAL WARNING", bg=DANGER, fg="white",
                 font=("Segoe UI",12,"bold"), pady=10).pack(fill="x")
        tk.Label(dlg, text="Type  DELETE  to confirm:", bg=CARD, fg=TEXT, font=("Segoe UI",10), pady=10).pack()
        ev = tk.StringVar()
        e = ttk.Entry(dlg, textvariable=ev, width=20, font=("Segoe UI",13,"bold"), justify="center")
        e.pack(pady=4); e.focus()
        def _confirm():
            if ev.get().strip().upper() == "DELETE":
                dlg.destroy(); db_delete_entire_ledger()
                self._selected_account = None; self._reload_current()
                messagebox.showinfo("Done", "All records deleted.", parent=self)
            else: messagebox.showerror("Mismatch", "Type  DELETE  exactly.", parent=dlg)
        bf = tk.Frame(dlg, bg=CARD); bf.pack(pady=10)
        _icon_btn(bf, "Confirm Delete", DANGER, DANGER_H, cmd=_confirm).pack(side="left", padx=8)
        _icon_btn(bf, "Cancel", ACCENT, ACCENT_H, cmd=dlg.destroy).pack(side="left", padx=8)


class ArrearsWindow(tk.Toplevel):
    def __init__(self, parent, prefill_cb=None):
        super().__init__(parent)
        self._prefill_cb = prefill_cb; self._all_arrears = {}
        self.title("Arrears Tracker"); self.geometry("960x620")
        self.configure(bg=BG); self.grab_set()
        self._fee_var = tk.StringVar(value=str(MONTHLY_FEE))
        self._build(); self._load()

    def _build(self):
        hdr = tk.Frame(self, bg=PURPLE, height=52)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        tk.Label(hdr, text="  📅  Arrears Tracker — Outstanding Dues",
                 bg=PURPLE, fg="white", font=("Segoe UI",13,"bold")).pack(side="left", padx=18, pady=16)

        ctrl = _card(self, highlightbackground=BORDER2, highlightthickness=1)
        ctrl.pack(fill="x", padx=12, pady=(8,0))
        cf = tk.Frame(ctrl, bg=CARD, padx=14, pady=10); cf.pack(fill="x")
        _field_label(cf, "Fee / month (Rs.) :").pack(side="left")
        ttk.Entry(cf, textvariable=self._fee_var, width=8).pack(side="left", padx=6)
        _icon_btn(cf, "Recalculate", ACCENT, ACCENT_H, cmd=self._load).pack(side="left")

        chips = tk.Frame(self, bg=BG); chips.pack(fill="x", padx=12, pady=8)
        self._chip_flats  = self._chip(chips, f"{UNIT_LABEL.upper()}S IN ARREARS", ERR_BG, ERR_FG)
        self._chip_months = self._chip(chips, "MONTHS DUE",       WARN_BG, WARN_FG)
        self._chip_owed   = self._chip(chips, "TOTAL OWED (Rs.)", PILL_BG, PURPLE)
        self._chip_clear  = self._chip(chips, "FULLY PAID UP",   OK_BG, OK_FG)

        tf = _card(self, highlightbackground=BORDER2, highlightthickness=1)
        tf.pack(fill="both", expand=True, padx=12, pady=(0,6))
        th = tk.Frame(tf, bg=PURPLE, height=32); th.pack(fill="x"); th.pack_propagate(False)
        tk.Label(th, text="  All Flats — Arrears Summary  (double-click to pre-fill payment form)",
                 bg=PURPLE, fg="white", font=("Segoe UI",9,"bold")).pack(side="left", pady=9)
        cols = (UNIT_LABEL,OWNER_LABEL,"FY Breakdown","Unpaid Months","Amount Owed")
        self._tbl = ttk.Treeview(tf, columns=cols, show="headings", height=12)
        for c, w, anc in [(UNIT_LABEL,55,"center"),(OWNER_LABEL,200,"w"),
                           ("FY Breakdown",370,"w"),("Unpaid Months",95,"center"),("Amount Owed",100,"center")]:
            self._tbl.heading(c, text=c); self._tbl.column(c, width=w, anchor=anc)
        vsb = ttk.Scrollbar(tf, orient="vertical", command=self._tbl.yview)
        self._tbl.configure(yscrollcommand=vsb.set)
        self._tbl.pack(side="left", fill="both", expand=True, padx=(6,0), pady=4)
        vsb.pack(side="right", fill="y", pady=4)
        self._tbl.tag_configure("has_arrears", background=ERR_BG, foreground=ERR_FG)
        self._tbl.tag_configure("no_arrears",  background=OK_BG,  foreground=OK_FG)
        self._tbl.bind("<Double-1>", self._on_dbl)
        self._tbl.bind("<<TreeviewSelect>>", self._on_select)

        self._detail_var = tk.StringVar(value="  Click any row to see breakdown.")
        tk.Label(self, textvariable=self._detail_var, bg=BG, fg=TEXT2,
                 font=("Segoe UI",8), justify="left").pack(anchor="w", padx=14, pady=(0,4))
        bf = tk.Frame(self, bg=BG); bf.pack(fill="x", padx=14, pady=6)
        _icon_btn(bf, "📲  WhatsApp All Defaulters", WA_GRN, WA_GRN_H, cmd=self._bulk_wa).pack(side="left", padx=(0,8))
        _icon_btn(bf, "Close", DANGER, DANGER_H, cmd=self.destroy).pack(side="right")

    def _chip(self, parent, label, bg, fg):
        f = tk.Frame(parent, bg=bg, padx=16, pady=8, highlightbackground=BORDER2, highlightthickness=1)
        f.pack(side="left", padx=(0,8))
        tk.Label(f, text=label, bg=bg, fg=fg, font=("Segoe UI",8,"bold")).pack()
        lbl = tk.Label(f, text="-", bg=bg, fg=fg, font=("Segoe UI",16,"bold")); lbl.pack()
        return lbl

    def _load(self):
        try:
            fee = float(self._fee_var.get())
            if fee <= 0: raise ValueError
        except ValueError:
            messagebox.showerror("Invalid Fee", "Enter a valid positive fee.", parent=self); return
        self._all_arrears = get_all_arrears(fee)
        for i in self._tbl.get_children(): self._tbl.delete(i)
        total_flats_arr=0; total_months=0; total_owed=0.0; total_clear=0
        for fid, res in RESIDENTS.items():
            arr = self._all_arrears[fid]
            if arr["count"] > 0:
                total_flats_arr+=1; total_months+=arr["count"]; total_owed+=arr["total_owed"]
                fy_parts = [f"{fy_label(fy)}: {len(m)} mo" for fy, m in sorted(arr["by_fy"].items())]
                self._tbl.insert("","end",iid=fid,
                                 values=(fid, res["name"], "  |  ".join(fy_parts),
                                         str(arr["count"]), f"Rs.{arr['total_owed']:,.0f}"), tags=("has_arrears",))
            else:
                total_clear+=1
                self._tbl.insert("","end",iid=fid,
                                 values=(fid, res["name"], "All months paid up","0","Rs.0"), tags=("no_arrears",))
        self._chip_flats.configure(text=str(total_flats_arr))
        self._chip_months.configure(text=str(total_months))
        self._chip_owed.configure(text=f"{total_owed:,.0f}")
        self._chip_clear.configure(text=str(total_clear))

    def _on_select(self, _=None):
        sel = self._tbl.selection()
        if not sel: return
        fid = sel[0]; arr = self._all_arrears.get(fid)
        if not arr or arr["count"] == 0:
            self._detail_var.set(f"  Flat {fid} — No arrears."); return
        try: fee = float(self._fee_var.get())
        except Exception: fee = MONTHLY_FEE
        lines = [f"  Flat {fid} — {RESIDENTS[fid]['name']} — {arr['count']} unpaid months = Rs.{arr['total_owed']:,.0f}"]
        for fy, months in sorted(arr["by_fy"].items()):
            lines.append(f"    {fy_label(fy)}: {', '.join(months)} ({len(months)} months = Rs.{len(months)*fee:,.0f})")
        self._detail_var.set("\n".join(lines))

    def _bulk_wa(self):
        defaulters = {fid: res for fid, res in RESIDENTS.items()
                      if self._all_arrears.get(fid, {}).get("count", 0) > 0}
        if not defaulters:
            messagebox.showinfo("No Arrears", "No units with outstanding arrears.", parent=self); return
        try: fee = float(self._fee_var.get())
        except Exception: fee = MONTHLY_FEE
        no_mob = [fid for fid, res in defaulters.items() if not res["mobile"].strip()]
        msg = f"Send arrears reminder to {len(defaulters)} unit(s)?"
        if no_mob: msg += f"\n\n⚠  {len(no_mob)} unit(s) have no mobile: {', '.join(no_mob)}"
        if not messagebox.askyesno("Bulk Arrears Reminder", msg, parent=self): return
        sent = skipped = 0
        for fid, res in defaulters.items():
            mob = res["mobile"].strip()
            if not mob: skipped += 1; continue
            arr = self._all_arrears[fid]
            fy_lines = "\n".join(f"  {fy_label(fy)}: {', '.join(months)}" for fy, months in sorted(arr["by_fy"].items()))
            line = "-" * 30; first = res["name"].split()[0].capitalize()
            text = (f"*{SOCIETY_NAME}*\n{line}\n   *ARREARS REMINDER*\n{line}\n"
                    f"{UNIT_LABEL} No  : *{fid}*\nDear {first},\n\n"
                    f"You have *{arr['count']} month(s)* of maintenance pending:\n{fy_lines}\n\n"
                    f"Total Outstanding: *Rs.{arr['total_owed']:,.0f}*\n\n"
                    f"Kindly clear your dues at the earliest.\n{line}\nThank you.")
            url = "https://wa.me/91" + mob.lstrip("+").lstrip("91") + "?text=" + urllib.parse.quote(text)
            webbrowser.open(url); sent += 1; self.after(400)
        messagebox.showinfo("Done", f"Opened WhatsApp for {sent} unit(s).\nSkipped {skipped} (no mobile).", parent=self)

    def _on_dbl(self, event):
        item = self._tbl.identify_row(event.y)
        if not item: return
        arr = self._all_arrears.get(item)
        if not arr or arr["count"] == 0:
            messagebox.showinfo("No Arrears", f"Flat {item} has no outstanding dues.", parent=self); return
        if not self._prefill_cb:
            messagebox.showinfo("Note", "Open from main window for pre-fill.", parent=self); return
        fy_years = sorted(arr["by_fy"].keys()); chosen_fy = fy_years[0]
        self._prefill_cb(item, chosen_fy, arr["by_fy"][chosen_fy]); self.destroy()


class UnpaidTrackerWindow(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Unpaid Tracker"); self.geometry("650x580")
        self.resizable(False, False); self.configure(bg=BG); self.grab_set()
        self._unpaid_map = {}

        tk.Label(self, text="  🔔  Month-Wise Unpaid Tracker",
                 bg=WARNING, fg="#0A0800", font=("Segoe UI",13,"bold"), pady=14).pack(fill="x")
        ctrl = _card(self, highlightbackground=BORDER2, highlightthickness=1)
        ctrl.pack(fill="x", padx=15, pady=(10,0))
        cf = tk.Frame(ctrl, bg=CARD, padx=16, pady=12); cf.pack(fill="x")
        _field_label(cf, "Month :").pack(side="left")
        self._month_var = tk.StringVar(value=MONTHS[datetime.date.today().month-1])
        ttk.Combobox(cf, textvariable=self._month_var, values=MONTHS, width=12, state="readonly").pack(side="left", padx=6)
        _field_label(cf, "Year :").pack(side="left", padx=(16,0))
        self._year_var = tk.StringVar(value=str(datetime.date.today().year))
        ttk.Spinbox(cf, from_=2020, to=2100, width=7, textvariable=self._year_var).pack(side="left", padx=6)
        _icon_btn(cf, "Check Now", ACCENT, ACCENT_H, cmd=self._check).pack(side="left", padx=10)

        chips = tk.Frame(self, bg=BG); chips.pack(fill="x", padx=15, pady=8)
        self._lbl_paid   = self._chip(chips, "PAID",   OK_BG, OK_FG)
        self._lbl_unpaid = self._chip(chips, "UNPAID", ERR_BG, ERR_FG)
        self._lbl_total  = self._chip(chips, "TOTAL",  PILL_BG, ACCENT)

        self._banner_var = tk.StringVar()
        tk.Label(self, textvariable=self._banner_var, bg=BG, fg=TEXT, font=("Segoe UI",10,"bold")).pack(fill="x", padx=16, pady=(0,4))

        to = _card(self, highlightbackground=BORDER2, highlightthickness=1)
        to.pack(fill="both", expand=True, padx=15)
        cols = ("#",UNIT_LABEL,f"{OWNER_LABEL} Name","Status","Quick Action")
        self._tbl = ttk.Treeview(to, columns=cols, show="headings", height=12)
        for col, w, anc in [("#",32,"center"),(UNIT_LABEL,52,"center"),(f"{OWNER_LABEL} Name",260,"w"),
                             ("Status",100,"center"),("Quick Action",110,"center")]:
            self._tbl.heading(col, text=col); self._tbl.column(col, width=w, anchor=anc)
        vsb = ttk.Scrollbar(to, orient="vertical", command=self._tbl.yview)
        self._tbl.configure(yscrollcommand=vsb.set)
        self._tbl.pack(side="left", fill="both", expand=True, padx=(6,0), pady=4)
        vsb.pack(side="right", fill="y", pady=4)
        self._tbl.tag_configure("paid",   background=OK_BG,  foreground=OK_FG)
        self._tbl.tag_configure("unpaid", background=ERR_BG, foreground=ERR_FG)
        self._tbl.bind("<Double-1>", self._on_dbl)
        tk.Label(self, text="  Double-click UNPAID row to send WhatsApp reminder",
                 bg=BG, fg=MUTED, font=("Segoe UI",8,"italic")).pack(anchor="w", padx=16, pady=(2,6))
        bf = tk.Frame(self, bg=BG); bf.pack(fill="x", padx=14, pady=6)
        _icon_btn(bf, "📲  Send All Reminders", WA_GRN, WA_GRN_H, cmd=self._send_all_reminders).pack(side="left", padx=(0,8))
        _icon_btn(bf, "Close", DANGER, DANGER_H, cmd=self.destroy).pack(side="right")
        self._check()

    def _chip(self, parent, label, bg, fg):
        f = tk.Frame(parent, bg=bg, padx=14, pady=6, highlightbackground=BORDER2, highlightthickness=1)
        f.pack(side="left", padx=(0,8))
        tk.Label(f, text=label, bg=bg, fg=fg, font=("Segoe UI",8,"bold")).pack()
        lbl = tk.Label(f, text="-", bg=bg, fg=fg, font=("Segoe UI",18,"bold")); lbl.pack()
        return lbl

    def _check(self):
        month = self._month_var.get()
        try: year = int(self._year_var.get())
        except Exception: return
        matrix = get_payment_matrix_with_fees(year if month in _FY_APR_DEC else year-1)
        paid_n=unpaid_n=0; self._unpaid_map={}
        for i in self._tbl.get_children(): self._tbl.delete(i)
        for idx, (fid, res) in enumerate(RESIDENTS.items(), start=1):
            if matrix[fid][month] > 0:
                tag="paid"; status="✓ Paid"; action=""; paid_n+=1
            else:
                tag="unpaid"; status="✗ Unpaid"; action="Send Reminder"
                unpaid_n+=1; self._unpaid_map[fid]=res
            self._tbl.insert("","end",iid=fid, values=(idx,fid,res["name"],status,action), tags=(tag,))
        self._lbl_paid.configure(text=str(paid_n))
        self._lbl_unpaid.configure(text=str(unpaid_n))
        self._lbl_total.configure(text=str(len(RESIDENTS)))
        if unpaid_n==0: self._banner_var.set(f"✓  All {len(RESIDENTS)} flats paid for {month} {year}")
        else: self._banner_var.set(f"  {unpaid_n} flat{'s' if unpaid_n>1 else ''} have NOT paid for {month} {year}")

    def _send_all_reminders(self):
        if not self._unpaid_map:
            messagebox.showinfo("All Paid", "No unpaid flats for this month.", parent=self); return
        month = self._month_var.get()
        try: year = int(self._year_var.get())
        except Exception: return
        no_mobile = [fid for fid, res in self._unpaid_map.items() if not res["mobile"].strip()]
        count = len(self._unpaid_map)
        msg = f"Send WhatsApp reminders to {count} unpaid flat(s) for {month} {year}?"
        if no_mobile:
            msg += f"\n\n⚠  {len(no_mobile)} flat(s) have no mobile: {', '.join(no_mobile)}"
        if not messagebox.askyesno("Send All Reminders", msg, parent=self): return
        sent = skipped = 0
        for fid, res in self._unpaid_map.items():
            mob = res["mobile"].strip()
            if not mob: skipped += 1; continue
            text = build_reminder(fid, res["name"], month, year)
            url = "https://wa.me/91" + mob.lstrip("+").lstrip("91") + "?text=" + urllib.parse.quote(text)
            webbrowser.open(url); sent += 1
            self.after(400)
        messagebox.showinfo("Done", f"Opened WhatsApp for {sent} flat(s).\nSkipped {skipped} (no mobile).", parent=self)

    def _on_dbl(self, event):
        item = self._tbl.identify_row(event.y)
        if not item or item not in self._unpaid_map: return
        res = self._unpaid_map[item]; month = self._month_var.get(); year = int(self._year_var.get())
        msg = build_reminder(item, res["name"], month, year); mob = res["mobile"].strip()
        if not mob:
            if messagebox.askyesno("No Mobile", f"No mobile for Flat {item}.\nCopy reminder?", parent=self):
                self.clipboard_clear(); self.clipboard_append(msg)
                messagebox.showinfo("Copied", "Reminder copied.", parent=self)
            return
        WhatsAppDialog(self, mob, msg, title_extra=f"Flat {item}")


class YearlyMatrixWindow(tk.Toplevel):
    MON_ABBR = [m[:3] for m in FY_MONTHS]

    def __init__(self, parent):
        super().__init__(parent)
        self.title("Yearly Payment Matrix"); self.geometry("1250x520")
        self.configure(bg=BG); self.grab_set()
        tk.Label(self, text="  Yearly Payment Matrix  (Financial Year: April → March)",
                 bg=SIDEBAR, fg="white", font=("Segoe UI",13,"bold"), pady=12).pack(fill="x")
        ctrl = _card(self, highlightbackground=BORDER2, highlightthickness=1)
        ctrl.pack(fill="x", padx=12, pady=(10,0))
        cf = tk.Frame(ctrl, bg=CARD, padx=14, pady=10); cf.pack(fill="x")
        _field_label(cf, "FY Start Year :").pack(side="left")
        today = datetime.date.today(); dfy = today.year if today.month>=4 else today.year-1
        self._yr = tk.StringVar(value=str(dfy))
        ttk.Spinbox(cf, from_=2020, to=2100, width=8, textvariable=self._yr).pack(side="left", padx=6)
        _icon_btn(cf, "Load", ACCENT, ACCENT_H, cmd=self._load).pack(side="left")
        self._stats_var = tk.StringVar()
        tk.Label(self, textvariable=self._stats_var, bg=BG, fg=TEXT, font=("Segoe UI",9,"bold")).pack(anchor="w", padx=16, pady=4)
        outer = _card(self, highlightbackground=BORDER2, highlightthickness=1)
        outer.pack(fill="both", expand=True, padx=12, pady=(0,8))
        self._canvas = tk.Canvas(outer, bg=CARD, highlightthickness=0)
        hb = ttk.Scrollbar(outer, orient="horizontal", command=self._canvas.xview)
        vb = ttk.Scrollbar(outer, orient="vertical",   command=self._canvas.yview)
        self._canvas.configure(xscrollcommand=hb.set, yscrollcommand=vb.set)
        hb.pack(side="bottom", fill="x"); vb.pack(side="right", fill="y")
        self._canvas.pack(side="left", fill="both", expand=True)
        self._gf = tk.Frame(self._canvas, bg=CARD)
        self._canvas.create_window((0,0), window=self._gf, anchor="nw")
        self._gf.bind("<Configure>", lambda e: self._canvas.configure(scrollregion=self._canvas.bbox("all")))
        self._load()

    def _load(self):
        try: fy = int(self._yr.get())
        except Exception: return
        for w in self._gf.winfo_children(): w.destroy()
        matrix = get_payment_matrix_with_fees(fy)
        WF=52; WO=190; WM=68; WC=70
        COLS = [UNIT_LABEL,f"{OWNER_LABEL} Name"] + self.MON_ABBR + ["Paid (Rs.)","Unpaid (mo)"]
        WIDTHS = [WF,WO] + [WM]*12 + [WC,WC]
        for c,(col,w) in enumerate(zip(COLS,WIDTHS)):
            bg = SIDEBAR if c < 2 else (HEAD_I if c < 14 else SIDEBAR)
            tk.Label(self._gf, text=col, bg=bg, fg="white",
                     font=("Segoe UI",8,"bold"), width=w//7, height=2, anchor="center").grid(
                         row=0, column=c, sticky="nsew", ipadx=2, ipady=3, padx=1, pady=1)
        grand_paid = 0.0
        for ri, (fid, res) in enumerate(RESIDENTS.items(), start=1):
            rb = ROW_EVEN if ri%2==0 else ROW_ODD
            tk.Label(self._gf, text=fid, bg=SIDEBAR, fg="white",
                     font=("Segoe UI",9,"bold"), width=WF//7, anchor="center").grid(
                         row=ri, column=0, sticky="nsew", padx=1, pady=1, ipady=5)
            tk.Label(self._gf, text=res["name"].title(), bg=rb, fg=TEXT,
                     font=("Segoe UI",8), width=WO//7, anchor="w").grid(
                         row=ri, column=1, sticky="nsew", padx=1, pady=1, ipady=5, ipadx=4)
            pt=0.0; uc=0
            for ci, mon in enumerate(FY_MONTHS):
                fv = matrix[fid][mon]
                if fv > 0:
                    cb2=OK_BG; cf2=OK_FG; sym=f"{fv:,.0f}"; pt+=fv; grand_paid+=fv; fnt=("Segoe UI",8)
                else:
                    cb2=ERR_BG; cf2=ERR_FG; sym="✗"; uc+=1; fnt=("Segoe UI",10,"bold")
                tk.Label(self._gf, text=sym, bg=cb2, fg=cf2, font=fnt, width=WM//7, anchor="center").grid(
                    row=ri, column=2+ci, sticky="nsew", padx=1, pady=1, ipady=5)
            tk.Label(self._gf, text=f"{pt:,.0f}", bg=OK_BG, fg=OK_FG,
                     font=("Segoe UI",8,"bold"), width=WC//7, anchor="center").grid(
                         row=ri, column=14, sticky="nsew", padx=1, pady=1, ipady=5)
            tk.Label(self._gf, text=str(uc), bg=ERR_BG if uc>0 else OK_BG,
                     fg=ERR_FG if uc>0 else OK_FG,
                     font=("Segoe UI",9,"bold"), width=WC//7, anchor="center").grid(
                         row=ri, column=15, sticky="nsew", padx=1, pady=1, ipady=5)
        paid_slots = sum(1 for fid in RESIDENTS for m in FY_MONTHS if matrix[fid][m] > 0)
        nf = len(RESIDENTS); pct = round(paid_slots/(nf*12)*100) if nf else 0
        self._stats_var.set(f"  {fy_label(fy)}   ·   Flats: {nf}   ·   Paid: {paid_slots}/{nf*12} ({pct}%)   ·   Total: Rs.{grand_paid:,.0f}")
        self._canvas.update_idletasks()
        self._canvas.configure(scrollregion=self._canvas.bbox("all"))


class ViewRecordsWindow(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Payment Records"); self.geometry("1020x560")
        self.configure(bg=BG); self.grab_set()

        hdr = tk.Frame(self, bg=SIDEBAR, height=56)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        tk.Label(hdr, text="  📋  Payment Records", bg=SIDEBAR, fg="white",
                 font=("Segoe UI", 13, "bold")).pack(side="left", padx=20, pady=16)
        _icon_btn(hdr, "Yearly Matrix", PURPLE, PURPLE_H,
                  cmd=lambda: YearlyMatrixWindow(self)).pack(side="right", padx=12, pady=12)

        ff = _card(self, highlightbackground=BORDER2, highlightthickness=1)
        ff.pack(fill="x", padx=12, pady=(10,0))
        inner = tk.Frame(ff, bg=CARD, padx=12, pady=10); inner.pack(fill="x")

        _field_label(inner, f"{UNIT_LABEL} :").grid(row=0, column=0, sticky="w", padx=(0,4))
        self._f_flat = tk.StringVar(value="All")
        ttk.Combobox(inner, textvariable=self._f_flat, values=["All"]+list(RESIDENTS.keys()),
                     width=7, state="readonly").grid(row=0, column=1, padx=(0,10))

        _field_label(inner, "FY Year :").grid(row=0, column=2, sticky="w", padx=(0,4))
        self._f_year = tk.StringVar(value="All")
        yr_now = datetime.date.today().year
        ttk.Combobox(inner, textvariable=self._f_year,
                     values=["All"]+[str(y) for y in range(yr_now+1, 2019, -1)],
                     width=8, state="readonly").grid(row=0, column=3, padx=(0,10))

        _icon_btn(inner, "Search", ACCENT, ACCENT_H, cmd=self._search).grid(row=0, column=4, padx=(0,6))
        _icon_btn(inner, "Reset", DANGER, DANGER_H, cmd=self._reset).grid(row=0, column=5)

        row2 = tk.Frame(ff, bg=CARD, padx=12, pady=8); row2.pack(fill="x")
        _field_label(row2, "Owner Name :").pack(side="left", padx=(0,4))
        self._f_owner = tk.StringVar()
        ttk.Entry(row2, textvariable=self._f_owner, width=22).pack(side="left", padx=(0,10))
        _icon_btn(row2, "Export CSV", TEAL, TEAL_H, cmd=self._export_csv).pack(side="right")

        self._stats_var = tk.StringVar()
        tk.Label(self, textvariable=self._stats_var, bg=BG, fg=TEXT, font=("Segoe UI",9,"bold")).pack(anchor="w", padx=14, pady=4)

        to = _card(self, highlightbackground=BORDER2, highlightthickness=1)
        to.pack(fill="both", expand=True, padx=12, pady=(0,6))
        cols = ("Receipt No","Date","Fin. Year",UNIT_LABEL,OWNER_LABEL,"Period","Fee/Month","Total Paid")
        self._tbl = ttk.Treeview(to, columns=cols, show="headings", height=16)
        for col, w, anc in [("Receipt No",130,"center"),("Date",88,"center"),
                             ("Fin. Year",78,"center"),(UNIT_LABEL,48,"center"),
                             (OWNER_LABEL,160,"w"),("Period",126,"w"),
                             ("Fee/Month",78,"center"),("Total Paid",88,"center")]:
            self._tbl.heading(col, text=col); self._tbl.column(col, width=w, anchor=anc)
        vsb = ttk.Scrollbar(to, orient="vertical", command=self._tbl.yview)
        self._tbl.configure(yscrollcommand=vsb.set)
        self._tbl.pack(side="left", fill="both", expand=True, padx=(6,0), pady=4)
        vsb.pack(side="right", fill="y", pady=4)
        self._tbl.bind("<Button-3>", self._rclick)
        self._tbl.bind("<Double-1>", self._on_dbl)
        self._tbl.tag_configure("odd",     background=ROW_ODD,  foreground=TEXT)
        self._tbl.tag_configure("even",    background=ROW_EVEN, foreground=TEXT)
        self._tbl.tag_configure("arrears", background=WARN_BG,  foreground=WARN_FG)
        tk.Label(self, text="  Double-click = export PDF  ·  Right-click = delete",
                 bg=BG, fg=MUTED, font=("Segoe UI",8,"italic")).pack(anchor="w", padx=12, pady=(0,4))
        self._all_rows = []; self._search()

    def _search(self):
        flat = self._f_flat.get(); year = self._f_year.get()
        rows = db_fetch_payments_filtered(flat_filter=None if flat=="All" else flat,
                                          year_filter=None if year=="All" else int(year))
        owner_q = self._f_owner.get().strip().lower()
        if owner_q: rows = [r for r in rows if owner_q in r[6].lower()]
        self._all_rows = rows; self._populate(rows)

    def _reset(self): self._f_flat.set("All"); self._f_year.set("All"); self._f_owner.set(""); self._search()

    def _export_csv(self):
        if not self._all_rows: messagebox.showinfo("No Data", "No records to export.", parent=self); return
        import csv
        fp = filedialog.asksaveasfilename(parent=self, defaultextension=".csv",
                                          filetypes=[("CSV files","*.csv")],
                                          initialfile="PaymentRecords.csv")
        if not fp: return
        with open(fp, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["Receipt No","Date","FY From","FY To",UNIT_LABEL,OWNER_LABEL,"Month From","Month To","Fee/Month","Amount"])
            for r in self._all_rows:
                pid, rno, date, yf, yt, flat, owner, amt, mf, mt, fee = r
                w.writerow([rno, date, yf, yt, flat, owner, mf, mt, fee, amt])
        messagebox.showinfo("Exported", f"CSV saved:\n{fp}", parent=self)
        if messagebox.askyesno("Open?", "Open the file now?", parent=self): webbrowser.open(f"file:///{fp}")

    def _populate(self, rows):
        for i in self._tbl.get_children(): self._tbl.delete(i)
        tot = 0; current_fy = get_current_fy()
        for idx, r in enumerate(rows):
            pid, rno, date, yf, yt, flat, owner, amt, mf, mt, fee = r
            if mf and mt and mf!=mt: period = f"{mf} - {mt} ({months_in_range(mf,mt)} mo)"
            elif mf: period = f"{mf} (1 mo)"
            else: period = "-"
            fee_s = f"Rs.{fee:,.0f}" if fee>0 else "-"
            is_arr = yf < current_fy
            tag = "arrears" if is_arr else ("odd" if idx%2 else "even")
            self._tbl.insert("","end",iid=str(pid),
                             values=(rno, date, f"{yf}-{yt}{'*' if is_arr else ''}",
                                     flat, owner, period, fee_s, f"Rs.{amt:,.0f}"), tags=(tag,))
            tot += amt
        n = len(rows)
        self._stats_var.set(f"  {n} record{'s' if n!=1 else ''}   ·   Total: Rs.{tot:,.0f}")

    def _rclick(self, event):
        item = self._tbl.identify_row(event.y)
        if not item: return
        m = tk.Menu(self, tearoff=0)
        m.add_command(label="Edit Record",    command=lambda: self._edit(int(item)))
        m.add_command(label="Export PDF",     command=lambda: self._export_pdf(int(item)))
        m.add_separator()
        m.add_command(label="Delete Record",  command=lambda: self._delete(int(item)))
        m.post(event.x_root, event.y_root)

    def _edit(self, pid):
        row = next((r for r in self._all_rows if r[0]==pid), None)
        if row: EditPaymentDialog(self, row, refresh_cb=self._search)

    def _delete(self, pid):
        if messagebox.askyesno("Delete","Permanently delete this record?",parent=self):
            db_delete_payment(pid); self._search()

    def _on_dbl(self, event):
        item = self._tbl.identify_row(event.y)
        if item: self._export_pdf(int(item))

    def _export_pdf(self, pid):
        row = next((r for r in self._all_rows if r[0]==pid), None)
        if not row: return
        _, rno, date, yf, yt, flat, owner, amt, mf, mt, fee = row
        d = {"receipt_no":rno,"date":date,"year_from":yf,"year_to":yt,
             "flat_no":flat,"owner_name":owner,"amount":amt,"month_from":mf,"month_to":mt,"monthly_fee":fee}
        fp = filedialog.asksaveasfilename(parent=self, defaultextension=".pdf",
                                          filetypes=[("PDF files","*.pdf")],
                                          initialfile=f"Receipt_{rno}.pdf")
        if fp and generate_receipt_pdf(d, fp):
            if messagebox.askyesno("Saved","Saved. Open now?",parent=self): webbrowser.open(f"file:///{fp}")


class AdmissionFeeWindow(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Admission Fees"); self.geometry("860x560")
        self.configure(bg=BG); self.grab_set()

        hdr = tk.Frame(self, bg=SIDEBAR, height=52)
        hdr.pack(fill="x"); hdr.pack_propagate(False)
        tk.Label(hdr, text="  💰  Admission Fee Tracker", bg=SIDEBAR, fg="white",
                 font=("Segoe UI",13,"bold")).pack(side="left", padx=18, pady=15)
        _icon_btn(hdr, "+ Add Fee", SUCCESS, SUCCESS_H, cmd=self._add_fee_manual).pack(side="right", padx=10, pady=10)

        self._filter = tk.StringVar(value="All")
        flt_f = tk.Frame(hdr, bg=SIDEBAR); flt_f.pack(side="right", padx=6, pady=10)
        for label, val in [("All","All"),("Paid","Paid"),("Unpaid","Unpaid")]:
            tk.Radiobutton(flt_f, text=label, variable=self._filter, value=val,
                           bg=SIDEBAR, fg=TEXT, selectcolor=SIDEBAR, activebackground=SIDEBAR,
                           activeforeground=TEXT, font=("Segoe UI",9),
                           command=self._refresh).pack(side="left", padx=4)

        sf = tk.Frame(self, bg=BG, padx=12, pady=6); sf.pack(fill="x")
        self._paid_lbl   = tk.Label(sf, bg=OK_BG,   fg=OK_FG,   font=("Segoe UI",9,"bold"), padx=12, pady=5, width=18)
        self._unpaid_lbl = tk.Label(sf, bg=ERR_BG,  fg=ERR_FG,  font=("Segoe UI",9,"bold"), padx=12, pady=5, width=18)
        self._total_lbl  = tk.Label(sf, bg=PILL_BG, fg=PILL_FG, font=("Segoe UI",9,"bold"), padx=12, pady=5, width=28)
        self._paid_lbl.pack(side="left", padx=(0,6))
        self._unpaid_lbl.pack(side="left", padx=(0,6))
        self._total_lbl.pack(side="left")

        to = _card(self, highlightbackground=BORDER2, highlightthickness=1)
        to.pack(fill="both", expand=True, padx=12, pady=(0,6))
        cols = (UNIT_LABEL, OWNER_LABEL, "Mobile", "Status", "Receipt No", "Date", "Amount")
        self._tbl = ttk.Treeview(to, columns=cols, show="headings", height=18)
        for col, w, anc in [(UNIT_LABEL,55,"center"),(OWNER_LABEL,160,"w"),
                             ("Mobile",110,"center"),("Status",90,"center"),
                             ("Receipt No",110,"center"),("Date",90,"center"),("Amount",100,"center")]:
            self._tbl.heading(col, text=col); self._tbl.column(col, width=w, anchor=anc)
        vsb = ttk.Scrollbar(to, orient="vertical", command=self._tbl.yview)
        self._tbl.configure(yscrollcommand=vsb.set)
        self._tbl.pack(side="left", fill="both", expand=True, padx=(6,0), pady=4)
        vsb.pack(side="right", fill="y", pady=4)
        self._tbl.tag_configure("paid",   background=OK_BG,  foreground=OK_FG)
        self._tbl.tag_configure("unpaid", background=ERR_BG, foreground=ERR_FG)
        self._tbl.bind("<Double-1>", self._on_dbl)
        self._tbl.bind("<Button-3>", self._rclick)
        tk.Label(self, text="  Double-click any row to record/edit  ·  Right-click paid row = delete",
                 bg=BG, fg=MUTED, font=("Segoe UI",8,"italic")).pack(anchor="w", padx=12, pady=(0,4))
        self._refresh()

    def _add_fee_manual(self):
        flats = db_get_flats()
        if not flats:
            messagebox.showinfo("No Flats", f"No {UNIT_LABEL}s found. Add them via Manage {UNIT_LABEL}s first.", parent=self); return
        paid_map = {r[3] for r in db_fetch_admission_fees()}
        dlg = tk.Toplevel(self); dlg.title(f"Select {UNIT_LABEL}"); dlg.geometry("300x160")
        dlg.configure(bg=CARD); dlg.grab_set(); dlg.resizable(False, False)
        tk.Label(dlg, text=f"  Select {UNIT_LABEL}", bg=SIDEBAR, fg="white",
                 font=("Segoe UI",11,"bold"), pady=10).pack(fill="x")
        fm = tk.Frame(dlg, bg=CARD, padx=20, pady=14); fm.pack(fill="both", expand=True)
        tk.Label(fm, text=f"{UNIT_LABEL} No", bg=CARD, fg=TEXT, font=("Segoe UI",9)).pack(anchor="w")
        flat_var = tk.StringVar()
        all_flats = [r[0] for r in flats]
        cb = ttk.Combobox(fm, textvariable=flat_var, values=all_flats, state="readonly", width=30)
        cb.pack(fill="x", pady=(4,12))
        def _go():
            fn = flat_var.get().strip()
            if not fn: return
            if fn in paid_map:
                if not messagebox.askyesno("Already Paid", f"{UNIT_LABEL} {fn} already has an admission fee record.\nRecord another?", parent=dlg): dlg.destroy(); return
            res = RESIDENTS.get(fn, {}); dlg.destroy()
            AdmissionFeeDialog(self, fn, res.get("name",""), res.get("mobile",""), after_save=self._refresh)
        _icon_btn(fm, "Continue", SUCCESS, SUCCESS_H, cmd=_go).pack(side="left")

    def _refresh(self):
        for i in self._tbl.get_children(): self._tbl.delete(i)
        paid_map = {r[3]: r for r in db_fetch_admission_fees()}
        flats = db_get_flats()
        filt = self._filter.get()
        paid_c = unpaid_c = 0; total_amt = 0.0
        for flat_no, name, mob, _ in flats:
            if flat_no in paid_map:
                paid_c += 1
                r = paid_map[flat_no]; amt = r[5]; total_amt += amt
                if filt in ("All","Paid"):
                    self._tbl.insert("","end", iid=f"p_{r[0]}",
                        values=(flat_no, r[4], r[6], "✔ Paid", r[1], r[2], f"Rs.{amt:,.0f}"), tags=("paid",))
            else:
                unpaid_c += 1
                if filt in ("All","Unpaid"):
                    self._tbl.insert("","end", iid=f"u_{flat_no}",
                        values=(flat_no, name, mob, "✘ Unpaid", "-", "-", "-"), tags=("unpaid",))
        if not flats:
            self._tbl.insert("","end", values=(f"No {UNIT_LABEL}s registered. Add via Manage {UNIT_LABEL}s.","","","","","",""))
        self._paid_lbl.configure(text=f"  ✔ Paid: {paid_c}  ")
        self._unpaid_lbl.configure(text=f"  ✘ Unpaid: {unpaid_c}  ")
        self._total_lbl.configure(text=f"  Total Collected: Rs.{total_amt:,.0f}  ")

    def _on_dbl(self, event):
        item = self._tbl.identify_row(event.y)
        if not item: return
        if item.startswith("u_"):
            flat_no = item[2:]
            res = RESIDENTS.get(flat_no, {})
            AdmissionFeeDialog(self, flat_no, res.get("name",""), res.get("mobile",""), after_save=self._refresh)
        elif item.startswith("p_"):
            messagebox.showinfo("Already Paid", f"This {UNIT_LABEL} has already paid.\nRight-click to delete if needed.", parent=self)

    def _rclick(self, event):
        item = self._tbl.identify_row(event.y)
        if not item or not item.startswith("p_"): return
        aid = int(item[2:])
        row = next((r for r in db_fetch_admission_fees() if r[0]==aid), None)
        m = tk.Menu(self, tearoff=0)
        if row:
            m.add_command(label="Export PDF Receipt", command=lambda: self._export_pdf(row))
            m.add_command(label="Send WhatsApp",      command=lambda: self._send_wa(row))
            m.add_separator()
        m.add_command(label="Delete Record", command=lambda: self._delete(aid))
        m.post(event.x_root, event.y_root)

    def _export_pdf(self, row):
        rid, rno, date, flat_no, owner, amt, mob, paid = row
        data = {"receipt_no": rno, "date": date, "flat_no": flat_no,
                "owner_name": owner, "amount": amt, "mobile": mob}
        fp = filedialog.asksaveasfilename(parent=self, defaultextension=".pdf",
                                          filetypes=[("PDF files","*.pdf")],
                                          initialfile=f"AdmFee_{rno}.pdf")
        if fp and generate_admission_fee_pdf(data, fp):
            messagebox.showinfo("Saved", f"Saved:\n{fp}", parent=self)
            if messagebox.askyesno("Open?", "Open now?", parent=self): webbrowser.open(f"file:///{fp}")

    def _send_wa(self, row):
        rid, rno, date, flat_no, owner, amt, mob, paid = row
        if not mob or not mob.strip():
            messagebox.showerror("No Mobile", f"No mobile number for {UNIT_LABEL} {flat_no}.", parent=self); return
        line = "-" * 30
        msg = (f"*{SOCIETY_NAME}*\n{line}\n   *ADMISSION FEE RECEIPT*\n{line}\n"
               f"Receipt No : {rno}\nDate       : {date}\n{line}\n"
               f"{UNIT_LABEL} No    : *{flat_no}*\n{OWNER_LABEL}      : {owner}\n{line}\n"
               f"Amount     : *Rs. {float(amt):,.0f}*\n{line}\n"
               f"*Thank you for your Admission Fee payment!*")
        WhatsAppDialog(self, mob.strip(), msg, title_extra=f"{UNIT_LABEL} {flat_no}")

    def _delete(self, aid):
        if messagebox.askyesno("Delete", "Remove admission fee record?", parent=self):
            db_delete_admission_fee(aid); self._refresh()


class AddExpenseDialog(tk.Toplevel):
    def __init__(self, parent, refresh_cb=None):
        super().__init__(parent); self._refresh = refresh_cb
        self.title("Add Expenditure"); self.geometry("420x320")
        self.resizable(False, False); self.configure(bg=CARD); self.grab_set()
        tk.Label(self, text="  Add Expenditure", bg=HEAD_E, fg="white",
                 font=("Segoe UI",11,"bold"), pady=12).pack(fill="x")
        frm = tk.Frame(self, bg=CARD, padx=28, pady=14); frm.pack(fill="both", expand=True)
        today_str = datetime.date.today().strftime("%d/%m/%Y")
        self._date_var = tk.StringVar(value=today_str)
        self._cat_var  = tk.StringVar(value=EXPENDITURE_CATEGORIES[0])
        self._desc_var = tk.StringVar()
        self._amt_var  = tk.StringVar()
        accounts = db_get_expenditure_accounts()
        self._acct_map = {a[1]: a[0] for a in accounts}
        self._acct_var = tk.StringVar(value=accounts[0][1] if accounts else "Society Expenditure")

        rows = [("Date (DD/MM/YYYY)", self._date_var, None),
                ("Account",           self._acct_var, [a[1] for a in accounts]),
                ("Category",          self._cat_var,  EXPENDITURE_CATEGORIES),
                ("Description",       self._desc_var, None),
                ("Amount (Rs.)",      self._amt_var,  None)]
        for r, (label, var, choices) in enumerate(rows):
            _field_label(frm, label).grid(row=r, column=0, sticky="w", pady=6, padx=(0,14))
            if choices: ttk.Combobox(frm, textvariable=var, values=choices, width=28, state="readonly").grid(row=r, column=1, sticky="w")
            else: ttk.Entry(frm, textvariable=var, width=30).grid(row=r, column=1, sticky="w")

        bf = tk.Frame(frm, bg=CARD); bf.grid(row=5, column=0, columnspan=2, pady=14)
        _icon_btn(bf, "Save", SUCCESS, SUCCESS_H, cmd=self._save).pack(side="left", padx=(0,8))
        _icon_btn(bf, "Cancel", DANGER, DANGER_H, cmd=self.destroy).pack(side="left")

    def _save(self):
        date=self._date_var.get().strip(); cat=self._cat_var.get().strip()
        desc=self._desc_var.get().strip(); amt_s=self._amt_var.get().strip()
        acct_name=self._acct_var.get().strip(); acct_id=self._acct_map.get(acct_name, 1)
        if not desc:
            messagebox.showerror("Missing","Enter a description.",parent=self); return
        try:
            amt = float(amt_s)
            if amt <= 0: raise ValueError
        except ValueError:
            messagebox.showerror("Invalid","Enter a valid positive amount.",parent=self); return
        db_save_expenditure(date, desc, amt, category=cat, account_id=acct_id)
        if self._refresh: self._refresh()
        self.destroy()


class ReportWindow(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Generate Reports"); self.geometry("480x410")
        self.resizable(False, False); self.configure(bg=BG); self.grab_set()
        tk.Label(self, text="  📄  Generate Reports & Statements",
                 bg=SIDEBAR, fg="white", font=("Segoe UI",12,"bold"), pady=12).pack(fill="x")
        card = _card(self, highlightbackground=BORDER2, highlightthickness=1)
        card.pack(fill="both", expand=True, padx=16, pady=14)
        inner = tk.Frame(card, bg=CARD, padx=24, pady=18); inner.pack(fill="both", expand=True)

        _field_label(inner, "Report Type :").grid(row=0, column=0, sticky="w", pady=6)
        self._rtype = tk.StringVar(value="flat")
        for val, txt, r in [("flat","Individual Flat Account Statement",0),
                             ("exp","Expenditure Account Statement",1),
                             ("all","All Flat Accounts (batch export)",2)]:
            ttk.Radiobutton(inner, text=txt, variable=self._rtype, value=val).grid(row=r, column=1, sticky="w", padx=10, pady=3)

        _field_label(inner, "FY Start Year :").grid(row=3, column=0, sticky="w", pady=8)
        today = datetime.date.today(); dfy = today.year if today.month>=4 else today.year-1
        self._year_var = tk.StringVar(value=str(dfy))
        ttk.Spinbox(inner, from_=2020, to=2100, width=9, textvariable=self._year_var).grid(row=3, column=1, sticky="w", padx=10)

        _field_label(inner, "Flat (individual) :").grid(row=4, column=0, sticky="w", pady=8)
        self._flat_var = tk.StringVar(value=list(RESIDENTS.keys())[0] if RESIDENTS else "")
        ttk.Combobox(inner, textvariable=self._flat_var, values=list(RESIDENTS.keys()), width=10, state="readonly").grid(row=4, column=1, sticky="w", padx=10)

        accounts = db_get_expenditure_accounts(); self._acct_map = {a[1]: a[0] for a in accounts}
        acct_names = [a[1] for a in accounts]
        _field_label(inner, "Exp. Account :").grid(row=5, column=0, sticky="w", pady=8)
        self._exp_acct_var = tk.StringVar(value=acct_names[0] if acct_names else "Society Expenditure")
        ttk.Combobox(inner, textvariable=self._exp_acct_var, values=acct_names, width=22, state="readonly").grid(row=5, column=1, sticky="w", padx=10)

        _h_sep(inner).grid(row=6, column=0, columnspan=2, sticky="ew", pady=10)
        bf = tk.Frame(inner, bg=CARD); bf.grid(row=7, column=0, columnspan=2)
        _icon_btn(bf, "Export PDF", SUCCESS, SUCCESS_H, cmd=self._export).pack(side="left", padx=(0,8))
        _icon_btn(bf, "Close", DANGER, DANGER_H, cmd=self.destroy).pack(side="left")

    def _export(self):
        try: fy = int(self._year_var.get())
        except Exception: return
        rtype = self._rtype.get()
        if rtype == "flat":
            flat = self._flat_var.get()
            fp = filedialog.asksaveasfilename(parent=self, defaultextension=".pdf",
                                              filetypes=[("PDF","*.pdf")],
                                              initialfile=f"Account_Flat{flat}_{fy_label(fy)}.pdf")
            if fp and generate_flat_account_pdf(flat, fy, fp):
                messagebox.showinfo("Saved",f"Saved:\n{fp}",parent=self)
                if messagebox.askyesno("Open?","Open now?",parent=self): webbrowser.open(f"file:///{fp}")
        elif rtype == "exp":
            aname = self._exp_acct_var.get(); aid = self._acct_map.get(aname, 1)
            safe = aname.replace(" ","_").replace("/","-")
            fp = filedialog.asksaveasfilename(parent=self, defaultextension=".pdf",
                                              filetypes=[("PDF","*.pdf")],
                                              initialfile=f"Expenditure_{safe}_{fy_label(fy)}.pdf")
            if fp and generate_expenditure_account_pdf(fy, fp, aid, aname):
                messagebox.showinfo("Saved",f"Saved:\n{fp}",parent=self)
                if messagebox.askyesno("Open?","Open now?",parent=self): webbrowser.open(f"file:///{fp}")
        else:
            folder = filedialog.askdirectory(parent=self, title="Choose folder")
            if not folder: return
            count = 0
            for flat_no in RESIDENTS:
                fp = os.path.join(folder, f"Account_Flat{flat_no}_{fy_label(fy)}.pdf")
                if generate_flat_account_pdf(flat_no, fy, fp): count += 1
            for aid, aname in db_get_expenditure_accounts():
                safe = aname.replace(" ","_").replace("/","-")
                fp_exp = os.path.join(folder, f"Expenditure_{safe}_{fy_label(fy)}.pdf")
                if generate_expenditure_account_pdf(fy, fp_exp, aid, aname): count += 1
            messagebox.showinfo("Done",f"Exported {count} PDFs to:\n{folder}",parent=self)


class EditPaymentDialog(tk.Toplevel):
    def __init__(self, parent, row_data, refresh_cb=None):
        super().__init__(parent)
        self._pid = row_data[0]; self._refresh = refresh_cb
        _, rno, date, yf, yt, flat, owner, amt, mf, mt, fee = row_data
        self.title(f"Edit Payment — {rno}"); self.geometry("480x400")
        self.resizable(False, False); self.configure(bg=CARD); self.grab_set()
        tk.Label(self, text=f"  ✎  Edit Payment — {rno}", bg=HEAD_I, fg="white",
                 font=("Segoe UI",11,"bold"), pady=12).pack(fill="x")
        tk.Label(self, text=f"  {UNIT_LABEL} {flat}  ·  {owner}", bg=BG, fg=TEXT2,
                 font=("Segoe UI",8), pady=5).pack(fill="x")
        frm = tk.Frame(self, bg=CARD, padx=28, pady=16); frm.pack(fill="both", expand=True)
        parts = date.split("/")
        self._dd  = tk.StringVar(value=parts[0] if len(parts)==3 else date[:2])
        self._dm  = tk.StringVar(value=parts[1] if len(parts)==3 else date[3:5])
        self._dy  = tk.StringVar(value=parts[2] if len(parts)==3 else date[6:])
        self._amt = tk.StringVar(value=str(amt))
        self._fee = tk.StringVar(value=str(fee))
        self._late_fee = tk.StringVar(value="0")
        self._mfrom = tk.StringVar(value=mf); self._mto = tk.StringVar(value=mt)
        self._yfrom = tk.StringVar(value=str(yf)); self._yto = tk.StringVar(value=str(yt))
        fields = [
            ("Date (DD / MM / YYYY)", None),
            ("Amount (Rs.)", self._amt),
            ("Monthly Fee (Rs.)", self._fee),
            ("Late Fee (Rs.)", self._late_fee),
            ("Month From", self._mfrom),
            ("Month To", self._mto),
            ("FY From", self._yfrom),
        ]
        r = 0
        _field_label(frm, "Date (DD / MM / YYYY)").grid(row=r, column=0, sticky="w", pady=5, padx=(0,12))
        df = tk.Frame(frm, bg=CARD); df.grid(row=r, column=1, sticky="w")
        for v, lo, hi, w, sep in [(self._dd,1,31,4," / "),(self._dm,1,12,4," / "),(self._dy,2020,2100,7,"")]:
            ttk.Spinbox(df, from_=lo, to=hi, width=w, textvariable=v).pack(side="left")
            if sep: tk.Label(df, text=sep, bg=CARD, fg=TEXT3, font=("Segoe UI",9)).pack(side="left")
        r += 1
        for label, var in [("Amount (Rs.)", self._amt),("Monthly Fee (Rs.)", self._fee),("Late Fee (Rs.)", self._late_fee)]:
            _field_label(frm, label).grid(row=r, column=0, sticky="w", pady=5)
            ttk.Entry(frm, textvariable=var, width=18).grid(row=r, column=1, sticky="w"); r += 1
        _field_label(frm, "Period  (From → To)").grid(row=r, column=0, sticky="w", pady=5)
        pf = tk.Frame(frm, bg=CARD); pf.grid(row=r, column=1, sticky="w")
        ttk.Combobox(pf, textvariable=self._mfrom, values=[""]+FY_MONTHS, width=10).pack(side="left")
        tk.Label(pf, text=" → ", bg=CARD, fg=TEXT3).pack(side="left")
        ttk.Combobox(pf, textvariable=self._mto, values=[""]+FY_MONTHS, width=10).pack(side="left"); r += 1
        _field_label(frm, "Financial Year").grid(row=r, column=0, sticky="w", pady=5)
        fyf = tk.Frame(frm, bg=CARD); fyf.grid(row=r, column=1, sticky="w")
        ttk.Spinbox(fyf, from_=2020, to=2100, width=7, textvariable=self._yfrom).pack(side="left")
        tk.Label(fyf, text=" → ", bg=CARD, fg=TEXT3).pack(side="left")
        ttk.Spinbox(fyf, from_=2021, to=2101, width=7, textvariable=self._yto).pack(side="left"); r += 1
        bf = tk.Frame(frm, bg=CARD); bf.grid(row=r, column=0, columnspan=2, pady=14)
        _icon_btn(bf, "✔  Save Changes", SUCCESS, SUCCESS_H, cmd=self._save).pack(side="left", padx=(0,8))
        _icon_btn(bf, "Cancel", DANGER, DANGER_H, cmd=self.destroy).pack(side="left")

    def _save(self):
        try:
            amt = float(self._amt.get()); assert amt > 0
            fee = float(self._fee.get()); late = float(self._late_fee.get())
            yf = int(self._yfrom.get()); yt = int(self._yto.get())
        except Exception:
            messagebox.showerror("Invalid", "Check numeric fields.", parent=self); return
        date = f"{self._dd.get().zfill(2)}/{self._dm.get().zfill(2)}/{self._dy.get()}"
        data = {"date": date, "year_from": yf, "year_to": yt,
                "flat_no": "", "owner_name": "",
                "amount": amt, "month_from": self._mfrom.get(), "month_to": self._mto.get(),
                "monthly_fee": fee, "late_fee": late}
        con = sqlite3.connect(_db_path())
        con.execute("""UPDATE payments SET date=?,year_from=?,year_to=?,amount=?,
                       month_from=?,month_to=?,monthly_fee=?,late_fee=? WHERE id=?""",
                    (date, yf, yt, amt, self._mfrom.get(), self._mto.get(), fee, late, self._pid))
        con.commit(); con.close()
        messagebox.showinfo("Updated", "Payment record updated.", parent=self)
        if self._refresh: self._refresh()
        self.destroy()


class ReceiptDialog(tk.Toplevel):
    def __init__(self, parent, data, after_save=None):
        super().__init__(parent)
        self._data = data; self._after_save = after_save; self._saved = False
        self.title("Receipt Preview"); self.geometry("480x560")
        self.resizable(False, False); self.configure(bg=CARD)
        self.grab_set(); self.focus_force()
        current_fy = get_current_fy(); is_arrears = data.get("year_from", current_fy) < current_fy
        hdr_bg = HEAD_E if is_arrears else SIDEBAR
        tk.Label(self, text="  Receipt Preview", bg=hdr_bg, fg="white",
                 font=("Segoe UI",12,"bold"), pady=12).pack(fill="x")
        if is_arrears:
            tk.Label(self, text="  ★  ARREARS SETTLEMENT", bg=WARN_BG, fg=WARN_FG,
                     font=("Segoe UI",9,"bold"), pady=6).pack(fill="x")
        msg = build_receipt(data)
        frm = tk.Frame(self, bg=CARD, padx=16, pady=10); frm.pack(fill="both", expand=True)
        txt = tk.Text(frm, font=("Consolas",10), bg=BG, fg=TEXT, relief="flat",
                      wrap="word", padx=12, pady=10, height=18, cursor="arrow",
                      highlightthickness=1, highlightbackground=BORDER2)
        txt.insert("1.0", msg); txt.configure(state="disabled"); txt.pack(fill="both", expand=True)
        _h_sep(self).pack(fill="x")
        bf = tk.Frame(self, bg=CARD, padx=16, pady=10); bf.pack(fill="x")
        _icon_btn(bf, "Copy",       ACCENT, ACCENT_H, cmd=lambda: self._copy(msg)).pack(side="left", padx=(0,4))
        _icon_btn(bf, "Save Record",SUCCESS, SUCCESS_H, cmd=self._save).pack(side="left", padx=(0,4))
        _icon_btn(bf, "PDF",        PURPLE, PURPLE_H, cmd=self._pdf).pack(side="left", padx=(0,4))
        _icon_btn(bf, "WhatsApp",   WA_GRN, WA_GRN_H, cmd=lambda: self._wa(msg)).pack(side="left")
        _icon_btn(bf, "Close",      DANGER, DANGER_H, cmd=self.destroy).pack(side="right")

    def _copy(self, msg):
        self.clipboard_clear(); self.clipboard_append(msg)
        messagebox.showinfo("Copied", "Receipt copied.", parent=self)

    def _ensure_saved(self):
        if not self._saved:
            if db_save_payment(self._data):
                self._saved = True
                if self._after_save: self._after_save()
                return True
            return False
        return True

    def _save(self):
        if self._saved:
            messagebox.showinfo("Already Saved","Already saved.",parent=self); return
        if self._ensure_saved():
            messagebox.showinfo("Saved!","Payment record saved.",parent=self); self.destroy()

    def _pdf(self):
        if not self._ensure_saved(): return
        fp = filedialog.asksaveasfilename(parent=self, defaultextension=".pdf",
                                          filetypes=[("PDF files","*.pdf")],
                                          initialfile=f"Receipt_{self._data['receipt_no']}.pdf")
        if fp and generate_receipt_pdf(self._data, fp):
            messagebox.showinfo("Saved",f"Saved:\n{fp}",parent=self)
            if messagebox.askyesno("Open?","Open now?",parent=self): webbrowser.open(f"file:///{fp}")

    def _wa(self, msg):
        mob = self._data.get("mobile","").strip()
        if not mob:
            messagebox.showerror("No Mobile","No mobile number.",parent=self); return
        if not self._ensure_saved(): return
        WhatsAppDialog(self, mob, msg, title_extra=f"Flat {self._data['flat_no']}")


class AdmissionFeeDialog(tk.Toplevel):
    def __init__(self, parent, flat_no, owner_name, mobile, after_save=None):
        super().__init__(parent)
        self.title("Record Admission Fee"); self.geometry("420x320")
        self.resizable(False, False); self.configure(bg=CARD)
        self.grab_set(); self.focus_force()
        self._after_save = after_save
        self._flat_no = flat_no
        tk.Label(self, text=f"  💰  Admission Fee — {UNIT_LABEL} {flat_no}", bg=SIDEBAR, fg="white",
                 font=("Segoe UI", 11, "bold"), pady=12).pack(fill="x")
        _h_sep(self).pack(side="bottom", fill="x")
        bf = tk.Frame(self, bg=CARD, padx=16, pady=10); bf.pack(side="bottom", fill="x")
        _icon_btn(bf, "Save", SUCCESS, SUCCESS_H, cmd=self._save).pack(side="left", padx=(0,8))
        _icon_btn(bf, "Cancel", DANGER, DANGER_H, cmd=self.destroy).pack(side="left")
        fm = tk.Frame(self, bg=CARD, padx=24, pady=16); fm.pack(fill="both", expand=True)
        self._owner = tk.StringVar(value=owner_name)
        self._mobile = tk.StringVar(value=mobile)
        self._amount = tk.StringVar()
        self._date = tk.StringVar(value=datetime.date.today().strftime("%d/%m/%Y"))
        for r, (label, var) in enumerate([("Owner Name", self._owner),
                                           ("Mobile",     self._mobile),
                                           ("Amount (Rs.)", self._amount),
                                           ("Date (DD/MM/YYYY)", self._date)]):
            tk.Label(fm, text=label, bg=CARD, fg=TEXT2, font=("Segoe UI",9)).grid(
                row=r, column=0, sticky="w", pady=7, padx=(0,14))
            ttk.Entry(fm, textvariable=var, width=28).grid(row=r, column=1, sticky="w")

    def _save(self):
        owner = self._owner.get().strip(); mob = self._mobile.get().strip()
        amt_s = self._amount.get().strip(); date = self._date.get().strip()
        if not owner or not amt_s or not date:
            messagebox.showerror("Missing", "Fill all fields.", parent=self); return
        try:
            amt = float(amt_s)
            if amt <= 0: raise ValueError
        except ValueError:
            messagebox.showerror("Invalid", "Amount must be positive.", parent=self); return
        data = {"receipt_no": _build_admission_receipt_no(), "date": date,
                "flat_no": self._flat_no, "owner_name": owner,
                "amount": amt, "mobile": mob, "paid": 1}
        if db_save_admission_fee(data):
            messagebox.showinfo("Saved", "Admission fee recorded.", parent=self)
            if self._after_save: self._after_save()
            self.destroy()


def _setup_styles(root):
    s = ttk.Style(root)
    s.theme_use("clam")
    s.configure(".", background=BG, foreground=TEXT, font=("Segoe UI", 10))
    s.configure("TFrame", background=BG)
    s.configure("TLabel", background=BG, foreground=TEXT, font=("Segoe UI", 10))
    s.configure("TEntry", fieldbackground="#111828", foreground=TEXT, insertcolor=ACCENT2,
                bordercolor=BORDER2, lightcolor=BORDER2, darkcolor=BORDER2,
                relief="flat", padding=10, font=("Segoe UI", 10))
    s.map("TEntry", fieldbackground=[("focus", "#111828")], bordercolor=[("focus", ACCENT)])
    s.configure("TCombobox", fieldbackground="#111828", background="#111828", foreground=TEXT,
                selectbackground="#111828", selectforeground=TEXT,
                bordercolor=BORDER2, arrowcolor=ACCENT2, padding=9, font=("Segoe UI", 10))
    s.map("TCombobox", fieldbackground=[("readonly", "#111828")], foreground=[("readonly", TEXT)],
          selectbackground=[("readonly", "#111828")], bordercolor=[("focus", ACCENT)])
    s.configure("TSpinbox", fieldbackground="#111828", foreground=TEXT, insertcolor=ACCENT2,
                bordercolor=BORDER2, arrowcolor=ACCENT2, padding=9, font=("Segoe UI", 10))
    s.map("TSpinbox", fieldbackground=[("focus", "#111828")], bordercolor=[("focus", ACCENT)])
    s.configure("TRadiobutton", background=CARD, foreground=TEXT, font=("Segoe UI", 10))
    s.map("TRadiobutton", background=[("active", CARD)], foreground=[("active", ACCENT2)])
    s.configure("Treeview", background=CARD, fieldbackground=CARD, foreground=TEXT,
                font=("Segoe UI", 10), rowheight=36, borderwidth=0)
    s.configure("Treeview.Heading", background="#0A0D1A", foreground=TEXT2,
                font=("Segoe UI", 9, "bold"), relief="flat", padding=10)
    s.map("Treeview", background=[("selected", ACCENT)], foreground=[("selected", "white")])
    s.map("Treeview.Heading", background=[("active", SIDEBAR_H)])
    s.configure("TScrollbar", troughcolor=BG, background=BORDER, borderwidth=0, arrowsize=12)
    s.map("TScrollbar", background=[("active", ACCENT)])


class SettingsWindow(tk.Toplevel):
    """Society-wide settings — name, address, monthly fee, FY start, expense categories."""

    def __init__(self, parent, on_close=None):
        super().__init__(parent)
        self._on_close = on_close
        self.title("Society Settings")
        self.geometry("560x560"); self.resizable(False, False)
        self.configure(bg=BG); self.grab_set()
        self.protocol("WM_DELETE_WINDOW", self._close)
        self._build()

    def _close(self):
        if self._on_close: self._on_close()
        self.destroy()

    def _entry(self, parent, var, width=38):
        e = tk.Entry(parent, textvariable=var, bg=CARD, fg=TEXT, insertbackground=TEXT,
                     relief="flat", font=("Segoe UI", 10), width=width,
                     highlightbackground=BORDER2, highlightthickness=1, highlightcolor=ACCENT)
        return e

    def _build(self):
        import json as _json
        # ── Header ─────────────────────────────────────────────────────────────
        hdr = tk.Frame(self, bg=PURPLE, height=56); hdr.pack(fill="x"); hdr.pack_propagate(False)
        tk.Label(hdr, text="  ⚙  Society Settings", bg=PURPLE, fg="white",
                 font=("Segoe UI", 13, "bold")).pack(side="left", padx=20, pady=16)

        body = tk.Frame(self, bg=BG, padx=28, pady=20); body.pack(fill="both", expand=True)

        # ── Society name ───────────────────────────────────────────────────────
        _field_label(body, "Society / Apartment Name").pack(anchor="w", pady=(0, 4))
        self._name_var = tk.StringVar(value=SOCIETY_NAME)
        self._entry(body, self._name_var).pack(fill="x", pady=(0, 14))

        # ── Address ────────────────────────────────────────────────────────────
        _field_label(body, "Address (shown on receipts & PDFs)").pack(anchor="w", pady=(0, 4))
        self._addr_var = tk.StringVar(value=SOCIETY_ADDRESS)
        self._entry(body, self._addr_var).pack(fill="x", pady=(0, 14))

        # ── Monthly fee & FY start ─────────────────────────────────────────────
        row2 = tk.Frame(body, bg=BG); row2.pack(fill="x", pady=(0, 14))
        lf = tk.Frame(row2, bg=BG); lf.pack(side="left", fill="x", expand=True, padx=(0, 12))
        _field_label(lf, "Default Monthly Fee (Rs.)").pack(anchor="w", pady=(0, 4))
        self._fee_var = tk.StringVar(value=str(int(MONTHLY_FEE)))
        self._entry(lf, self._fee_var, width=16).pack(anchor="w")

        rf = tk.Frame(row2, bg=BG); rf.pack(side="left", fill="x", expand=True)
        _field_label(rf, "Society Start Financial Year").pack(anchor="w", pady=(0, 4))
        self._fy_var = tk.StringVar(value=str(SOCIETY_START_FY))
        self._entry(rf, self._fy_var, width=10).pack(anchor="w")

        # ── Unit / Member label ────────────────────────────────────────────────
        row3 = tk.Frame(body, bg=BG); row3.pack(fill="x", pady=(0, 14))
        _field_label(row3, "Unit Label (Flat / Member / Plot / Shop …)").pack(anchor="w", pady=(0, 4))
        self._unit_var = tk.StringVar(value=UNIT_LABEL)
        ttk.Combobox(row3, textvariable=self._unit_var,
                     values=["Flat", "Member", "Plot", "Unit", "Shop", "Office", "Villa", "Site"],
                     width=16).pack(anchor="w")

        # ── Expenditure categories ─────────────────────────────────────────────
        tk.Frame(body, bg=BORDER, height=1).pack(fill="x", pady=(4, 12))
        tk.Label(body, text="EXPENDITURE CATEGORIES", bg=BG, fg=TEXT3,
                 font=("Segoe UI", 7, "bold")).pack(anchor="w", pady=(0, 6))
        tk.Label(body, text="One category per line:", bg=BG, fg=TEXT2,
                 font=("Segoe UI", 9)).pack(anchor="w", pady=(0, 4))

        txt_frame = tk.Frame(body, bg=BORDER, padx=1, pady=1); txt_frame.pack(fill="x")
        self._cats_text = tk.Text(txt_frame, bg=CARD, fg=TEXT, insertbackground=TEXT,
                                  relief="flat", font=("Segoe UI", 10), height=6, wrap="none",
                                  highlightthickness=0)
        self._cats_text.pack(fill="x")
        self._cats_text.insert("1.0", "\n".join(EXPENDITURE_CATEGORIES))

        # ── Buttons ────────────────────────────────────────────────────────────
        btn_row = tk.Frame(body, bg=BG); btn_row.pack(fill="x", pady=(16, 0))
        _icon_btn(btn_row, "✔  Save Settings", SUCCESS, SUCCESS_H, cmd=self._save).pack(side="left")
        _icon_btn(btn_row, "✕  Cancel", DANGER, DANGER_H, cmd=self._close).pack(side="left", padx=10)

    def _save(self):
        import json as _json
        name = self._name_var.get().strip()
        addr = self._addr_var.get().strip()
        if not name:
            messagebox.showerror("Missing", "Society name cannot be empty.", parent=self); return
        try:
            fee = float(self._fee_var.get())
            if fee <= 0: raise ValueError
        except Exception:
            messagebox.showerror("Invalid", "Monthly fee must be a positive number.", parent=self); return
        try:
            fy = int(self._fy_var.get())
            if fy < 2000 or fy > 2100: raise ValueError
        except Exception:
            messagebox.showerror("Invalid", "Start financial year must be a 4-digit year.", parent=self); return
        cats_raw = self._cats_text.get("1.0", "end").strip()
        cats = [c.strip() for c in cats_raw.splitlines() if c.strip()]
        if not cats:
            messagebox.showerror("Missing", "At least one expenditure category is required.", parent=self); return

        db_set_setting("society_name",    name)
        db_set_setting("society_address", addr)
        db_set_setting("monthly_fee",     str(fee))
        db_set_setting("society_start_fy",str(fy))
        db_set_setting("unit_label",      self._unit_var.get().strip() or "Flat")
        db_set_setting("exp_categories",  _json.dumps(cats))
        existing = {n for _, n in db_get_expenditure_accounts()}
        con = sqlite3.connect(_db_path())
        for cat in cats:
            if cat not in existing:
                con.execute("INSERT OR IGNORE INTO expenditure_accounts (name) VALUES (?)", (cat,))
        con.commit(); con.close()
        load_settings()
        messagebox.showinfo("Saved", "Settings saved!\nRestart the app for all labels to refresh.", parent=self)
        self._close()


class SocietyLauncherWindow(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Society Manager — Select Account")
        self.geometry("600x540"); self.resizable(False, False)
        self.configure(bg=BG)
        self._choice = None
        _setup_styles(self)
        self._build()
        self._load()

    def _build(self):
        hdr = tk.Frame(self, bg=ACCENT, height=64); hdr.pack(fill="x"); hdr.pack_propagate(False)
        left_stripe = tk.Frame(hdr, bg=ACCENT_H, width=6); left_stripe.pack(side="left", fill="y")
        hdr_inner = tk.Frame(hdr, bg=ACCENT); hdr_inner.pack(side="left", padx=18, pady=18)
        tk.Label(hdr_inner, text="🌐  Society Account Manager", bg=ACCENT, fg="white",
                 font=("Segoe UI", 15, "bold")).pack(side="left")

        info = tk.Frame(self, bg="#040509", height=32); info.pack(fill="x"); info.pack_propagate(False)
        tk.Label(info, text="  Select a registered society to open, or create a new one",
                 bg="#040509", fg=TEXT3, font=("Segoe UI", 8, "italic")).pack(side="left", padx=16, pady=9)

        sep = tk.Frame(self, bg=BORDER, height=1); sep.pack(side="bottom", fill="x")
        bf = tk.Frame(self, bg="#040509", padx=18, pady=14); bf.pack(side="bottom", fill="x")
        _icon_btn(bf, "▶  Open Selected", ACCENT, ACCENT_H, cmd=self._open_selected).pack(side="left", padx=(0,8))
        _icon_btn(bf, "+ New Society",    SUCCESS, SUCCESS_H, cmd=self._new_society).pack(side="left", padx=(0,8))
        _icon_btn(bf, "📂 Open .db File", TEAL,   TEAL_H,   cmd=self._open_existing).pack(side="left", padx=(0,8))
        _icon_btn(bf, "✕ Remove",         DANGER,  DANGER_H,  cmd=self._remove_selected).pack(side="left")
        _icon_btn(bf, "✎ Rename",         WARNING, WARNING_H, cmd=self._rename_selected).pack(side="right")

        body = tk.Frame(self, bg=BG, padx=20, pady=16); body.pack(fill="both", expand=True)

        lf = tk.Frame(body, bg=BG); lf.pack(fill="both", expand=True)
        tk.Label(lf, text="Registered Societies", bg=BG, fg=TEXT2,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(0, 8))

        tbl_frame = tk.Frame(lf, bg=CARD, highlightbackground=BORDER, highlightthickness=1)
        tbl_frame.pack(fill="both", expand=True)

        cols = ("Society Name", "Database File")
        self._tree = ttk.Treeview(tbl_frame, columns=cols, show="headings", height=10, selectmode="browse")
        self._tree.heading("Society Name", text="Society Name")
        self._tree.heading("Database File", text="Database File")
        self._tree.column("Society Name", width=200, anchor="w")
        self._tree.column("Database File", width=310, anchor="w")
        vsb = ttk.Scrollbar(tbl_frame, orient="vertical", command=self._tree.yview)
        self._tree.configure(yscrollcommand=vsb.set)
        self._tree.pack(side="left", fill="both", expand=True, padx=(6,0), pady=6)
        vsb.pack(side="left", fill="y", pady=6)
        self._tree.tag_configure("odd",  background=ROW_ODD,  foreground=TEXT)
        self._tree.tag_configure("even", background=ROW_EVEN, foreground=TEXT)
        self._tree.bind("<Double-1>", lambda e: self._open_selected())

    def _load(self):
        for i in self._tree.get_children(): self._tree.delete(i)
        for idx, s in enumerate(society_list()):
            tag = "odd" if idx % 2 else "even"
            self._tree.insert("", "end", iid=s["db"],
                              values=(s["name"], s["db"]), tags=(tag,))

    def _selected_db(self):
        sel = self._tree.selection()
        return sel[0] if sel else None

    def _open_selected(self):
        db = self._selected_db()
        if not db:
            messagebox.showwarning("Select", "Please select a society first.", parent=self); return
        society_switch(db); self._choice = db; self.destroy()

    def _new_society(self):
        dlg = _SimpleSocietyDialog(self, mode="new")
        self.wait_window(dlg)
        if dlg.result:
            name, path, unit_label = dlg.result
            society_add(name, path)
            society_switch(path)
            init_db()
            db_set_setting("society_name", name)
            db_set_setting("unit_label", unit_label or "Flat")
            self._choice = path
            self.destroy()

    def _open_existing(self):
        path = filedialog.askopenfilename(parent=self, title="Open Society Database",
                                          filetypes=[("SQLite DB", "*.db"), ("All Files", "*.*")])
        if not path: return
        existing = {r["db"]: r["name"] for r in society_list()}
        if path in existing:
            society_switch(path); self._choice = path; self.destroy(); return
        name = os.path.splitext(os.path.basename(path))[0]
        dlg = _SimpleSocietyDialog(self, mode="name", default_name=name)
        self.wait_window(dlg)
        if dlg.result:
            name, _, _u = dlg.result
            society_add(name, path)
        society_switch(path); self._choice = path; self.destroy()

    def _remove_selected(self):
        db = self._selected_db()
        if not db: return
        if messagebox.askyesno("Remove", "Remove this society from the list?\n(The .db file will not be deleted.)",
                               parent=self, icon="warning"):
            society_remove(db); self._load()

    def _rename_selected(self):
        db = self._selected_db()
        if not db: return
        reg = {r["db"]: r["name"] for r in society_list()}
        dlg = _SimpleSocietyDialog(self, mode="name", default_name=reg.get(db, ""))
        self.wait_window(dlg)
        if dlg.result:
            society_rename(db, dlg.result[0]); self._load()


class _SimpleSocietyDialog(tk.Toplevel):
    def __init__(self, parent, mode="new", default_name=""):
        super().__init__(parent)
        self.result = None; self._mode = mode
        self.title("New Society" if mode == "new" else "Society Name")
        self.geometry("420x370" if mode == "new" else "380x150")
        self.resizable(False, False); self.configure(bg=CARD); self.grab_set()

        tk.Label(self, text="  " + ("Create New Society" if mode == "new" else "Enter Society Name"),
                 bg=ACCENT, fg="white", font=("Segoe UI", 11, "bold"), pady=10).pack(fill="x")

        frm = tk.Frame(self, bg=CARD, padx=24, pady=16); frm.pack(fill="both", expand=True)
        self._name_var = tk.StringVar(value=default_name)
        tk.Label(frm, text="Society Name:", bg=CARD, fg=TEXT2, font=("Segoe UI", 9)).grid(row=0, column=0, sticky="w", pady=8)
        ttk.Entry(frm, textvariable=self._name_var, width=30).grid(row=0, column=1, padx=10)

        self._path_var = tk.StringVar(); self._unit_var = tk.StringVar(value="Flat")
        if mode == "new":
            tk.Label(frm, text="Save .db to:", bg=CARD, fg=TEXT2, font=("Segoe UI", 9)).grid(row=1, column=0, sticky="w", pady=8)
            pf = tk.Frame(frm, bg=CARD); pf.grid(row=1, column=1, padx=10, sticky="w")
            ttk.Entry(pf, textvariable=self._path_var, width=22).pack(side="left")
            _icon_btn(pf, "Browse", TEAL, TEAL_H, cmd=self._browse).pack(side="left", padx=4)
            tk.Label(frm, text="Unit Type:", bg=CARD, fg=TEXT2, font=("Segoe UI", 9)).grid(row=2, column=0, sticky="w", pady=8)
            ttk.Combobox(frm, textvariable=self._unit_var,
                         values=["Flat","Member","Plot","Unit","Shop","Office","Villa","Site"],
                         width=14, state="readonly").grid(row=2, column=1, padx=10, sticky="w")
            guide = tk.Frame(frm, bg="#0D1117", padx=10, pady=8)
            guide.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(4,0))
            tips = (
                "Flat → Apartment / Housing society\n"
                "Member → Club / Association / RWA\n"
                "Plot → Plotted layout / Gated community\n"
                "Shop / Office → Commercial complex\n"
                "Villa / Site → Villa community / Layout"
            )
            tk.Label(guide, text=tips, bg="#0D1117", fg="#475569",
                     font=("Segoe UI", 8), justify="left").pack(anchor="w")

        bf = tk.Frame(frm, bg=CARD)
        bf.grid(row=4 if mode == "new" else 1, column=0, columnspan=2, pady=12)
        _icon_btn(bf, "OK",     SUCCESS, SUCCESS_H, cmd=self._ok).pack(side="left", padx=(0, 8))
        _icon_btn(bf, "Cancel", DANGER,  DANGER_H,  cmd=self.destroy).pack(side="left")

    def _browse(self):
        path = filedialog.asksaveasfilename(parent=self, defaultextension=".db",
                                            filetypes=[("SQLite DB", "*.db")],
                                            title="Save Society Database As")
        if path: self._path_var.set(path)

    def _ok(self):
        name = self._name_var.get().strip()
        if not name:
            messagebox.showerror("Missing", "Society name is required.", parent=self); return
        if self._mode == "new":
            path = self._path_var.get().strip()
            if not path:
                messagebox.showerror("Missing", "Choose where to save the database.", parent=self); return
            self.result = (name, path, self._unit_var.get())
        else:
            self.result = (name, None, None)
        self.destroy()


class DashboardWindow(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Dashboard"); self.geometry("1100x740")
        self.configure(bg=BG); self.resizable(True, True)
        today = datetime.date.today()
        dfy = today.year if today.month >= 4 else today.year - 1
        self._fy_var = tk.StringVar(value=str(dfy))
        hdr = tk.Frame(self, bg=ACCENT, height=50); hdr.pack(fill="x"); hdr.pack_propagate(False)
        tk.Label(hdr, text="  📊  Dashboard", bg=ACCENT, fg="white",
                 font=("Segoe UI",13,"bold")).pack(side="left", padx=18, pady=14)
        fy_f = tk.Frame(hdr, bg=ACCENT); fy_f.pack(side="right", padx=20)
        tk.Label(fy_f, text="FY:", bg=ACCENT, fg="white", font=("Segoe UI",9)).pack(side="left", padx=(0,4))
        ttk.Combobox(fy_f, textvariable=self._fy_var,
                     values=[str(y) for y in range(max(SOCIETY_START_FY, 2020), dfy+2)],
                     width=6, state="readonly").pack(side="left")
        _icon_btn(fy_f, "Refresh", SUCCESS, SUCCESS_H, cmd=self._load).pack(side="left", padx=8)
        self._fy_var.trace_add("write", lambda *_: self._load())
        canvas = tk.Canvas(self, bg=BG, highlightthickness=0)
        vscroll = ttk.Scrollbar(self, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vscroll.set)
        vscroll.pack(side="right", fill="y"); canvas.pack(fill="both", expand=True)
        self._body = tk.Frame(canvas, bg=BG)
        canvas.create_window((0,0), window=self._body, anchor="nw", tags="body")
        self._body.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig("body", width=e.width))
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))
        self._load()

    def _load(self):
        for w in self._body.winfo_children(): w.destroy()
        try: fy = int(self._fy_var.get())
        except Exception: fy = get_current_fy()
        self._build_dashboard(fy)

    def _build_dashboard(self, fy):
        body = self._body
        n_flats = len(RESIDENTS)
        expected = n_flats * 12 * MONTHLY_FEE
        total_collected = 0.0; total_dues = 0.0; defaulters = []
        for flat_no, res in RESIDENTS.items():
            acct = get_flat_account_entries(flat_no, fy)
            total_collected += acct["total_cr"]
            cb = acct["closing_balance"]
            if cb > 0: total_dues += cb; defaulters.append((flat_no, res["name"], cb))
        total_exp = 0.0; exp_by_account = []
        for aid, aname in db_get_expenditure_accounts():
            ea = get_expenditure_account_entries(fy, aid)
            net = ea["total_dr"] - ea["total_cr"]
            total_exp += net; exp_by_account.append((aname, net))
        adm_total = sum(r[5] for r in db_fetch_admission_fees())
        net_position = total_collected + adm_total - total_exp
        collection_pct = (total_collected / expected * 100) if expected > 0 else 0

        cards_frame = tk.Frame(body, bg=BG); cards_frame.pack(fill="x", padx=20, pady=(16,8))
        def stat_card(label, value, color, sub=""):
            card = tk.Frame(cards_frame, bg=CARD, highlightbackground=color, highlightthickness=2, padx=16, pady=14)
            card.pack(side="left", fill="both", expand=True, padx=6)
            tk.Label(card, text=label, bg=CARD, fg=TEXT2, font=("Segoe UI",8,"bold")).pack(anchor="w")
            tk.Label(card, text=value, bg=CARD, fg=color, font=("Segoe UI",15,"bold")).pack(anchor="w", pady=(4,0))
            if sub: tk.Label(card, text=sub, bg=CARD, fg=TEXT3, font=("Segoe UI",8)).pack(anchor="w")

        stat_card("TOTAL EXPECTED",    f"Rs.{expected:,.0f}",       ACCENT,  f"{n_flats} units × 12mo × Rs.{MONTHLY_FEE:,.0f}")
        stat_card("COLLECTED",         f"Rs.{total_collected:,.0f}", SUCCESS, f"{collection_pct:.1f}% of expected")
        stat_card("OUTSTANDING DUES",  f"Rs.{total_dues:,.0f}",     DANGER,  f"{len(defaulters)} units pending")
        stat_card("TOTAL EXPENDITURE", f"Rs.{total_exp:,.0f}",      WARNING, f"{len(exp_by_account)} expense accounts")
        stat_card("NET POSITION",      f"Rs.{abs(net_position):,.0f}",
                  SUCCESS if net_position >= 0 else DANGER,
                  "Surplus" if net_position >= 0 else "Deficit")

        prog_frame = tk.Frame(body, bg=CARD, highlightbackground=BORDER2, highlightthickness=1)
        prog_frame.pack(fill="x", padx=20, pady=(4,8))
        ph = tk.Frame(prog_frame, bg=HEAD_I, height=36); ph.pack(fill="x"); ph.pack_propagate(False)
        tk.Label(ph, text=f"  Monthly Collection Progress — {fy_label(fy)}", bg=HEAD_I, fg="white",
                 font=("Segoe UI",10,"bold")).pack(side="left", padx=14, pady=10)
        matrix = get_payment_matrix_with_fees(fy)
        prog_body = tk.Frame(prog_frame, bg=CARD, padx=16, pady=12); prog_body.pack(fill="x")
        for month in FY_MONTHS:
            paid_count = sum(1 for f in RESIDENTS if matrix.get(f, {}).get(month, 0) > 0)
            pct = paid_count / n_flats if n_flats > 0 else 0
            row = tk.Frame(prog_body, bg=CARD); row.pack(fill="x", pady=2)
            tk.Label(row, text=month[:3], bg=CARD, fg=TEXT2, font=("Segoe UI",8), width=4).pack(side="left")
            bar_bg = tk.Frame(row, bg=BORDER, height=14); bar_bg.pack(side="left", padx=6, fill="x", expand=True)
            bar_bg.update_idletasks()
            bar_color = SUCCESS if pct >= 0.9 else (WARNING if pct >= 0.5 else DANGER)
            bar_fill = tk.Frame(bar_bg, bg=bar_color, height=14)
            bar_fill.place(relx=0, rely=0, relwidth=pct, relheight=1.0)
            tk.Label(row, text=f"{paid_count}/{n_flats} paid  ({pct*100:.0f}%)", bg=CARD, fg=TEXT3,
                     font=("Segoe UI",8), width=18).pack(side="left", padx=4)

        bot_frame = tk.Frame(body, bg=BG); bot_frame.pack(fill="both", expand=True, padx=20, pady=(0,20))

        left_panel = tk.Frame(bot_frame, bg=CARD, highlightbackground=BORDER2, highlightthickness=1)
        left_panel.pack(side="left", fill="both", expand=True, padx=(0,8))
        lh = tk.Frame(left_panel, bg=HEAD_E, height=36); lh.pack(fill="x"); lh.pack_propagate(False)
        tk.Label(lh, text="  ⚠  Top Defaulters", bg=HEAD_E, fg="white",
                 font=("Segoe UI",10,"bold")).pack(side="left", padx=14, pady=10)
        defaulters.sort(key=lambda x: -x[2])
        if defaulters:
            for i, (fno, name, due) in enumerate(defaulters[:12]):
                r = tk.Frame(left_panel, bg=ROW_ODD if i%2==0 else ROW_EVEN, padx=12, pady=7)
                r.pack(fill="x")
                tk.Label(r, text=f"{UNIT_LABEL} {fno}", bg=r["bg"], fg=DANGER,
                         font=("Segoe UI",9,"bold"), width=8).pack(side="left")
                tk.Label(r, text=name, bg=r["bg"], fg=TEXT, font=("Segoe UI",9)).pack(side="left", padx=6)
                tk.Label(r, text=f"Rs.{due:,.0f}", bg=r["bg"], fg=WARNING,
                         font=("Segoe UI",9,"bold")).pack(side="right")
        else:
            tk.Label(left_panel, text="✅  All units clear!", bg=CARD, fg=SUCCESS,
                     font=("Segoe UI",11,"bold"), pady=24).pack()

        right_panel = tk.Frame(bot_frame, bg=CARD, highlightbackground=BORDER2, highlightthickness=1)
        right_panel.pack(side="left", fill="both", expand=True, padx=(8,0))
        rh = tk.Frame(right_panel, bg=SIDEBAR, height=36); rh.pack(fill="x"); rh.pack_propagate(False)
        tk.Label(rh, text="  💸  Expenditure Breakdown", bg=SIDEBAR, fg="white",
                 font=("Segoe UI",10,"bold")).pack(side="left", padx=14, pady=10)
        for i, (aname, net) in enumerate(sorted(exp_by_account, key=lambda x: -x[1])):
            r = tk.Frame(right_panel, bg=ROW_ODD if i%2==0 else ROW_EVEN, padx=12, pady=7)
            r.pack(fill="x")
            tk.Label(r, text=aname, bg=r["bg"], fg=TEXT, font=("Segoe UI",9)).pack(side="left")
            tk.Label(r, text=f"Rs.{net:,.0f}", bg=r["bg"],
                     fg=WARNING if net > 0 else SUCCESS, font=("Segoe UI",9,"bold")).pack(side="right")
        tk.Frame(right_panel, bg=BORDER2, height=1).pack(fill="x", padx=12)
        r = tk.Frame(right_panel, bg=CARD, padx=12, pady=8); r.pack(fill="x")
        tk.Label(r, text="TOTAL EXPENDITURE", bg=CARD, fg=TEXT2, font=("Segoe UI",9,"bold")).pack(side="left")
        tk.Label(r, text=f"Rs.{total_exp:,.0f}", bg=CARD, fg=WARNING,
                 font=("Segoe UI",11,"bold")).pack(side="right")
        tk.Frame(right_panel, bg=BORDER2, height=1).pack(fill="x", padx=12)
        r2 = tk.Frame(right_panel, bg=CARD, padx=12, pady=8); r2.pack(fill="x")
        tk.Label(r2, text="NET POSITION", bg=CARD, fg=TEXT2, font=("Segoe UI",9,"bold")).pack(side="left")
        net_color = SUCCESS if net_position >= 0 else DANGER
        tk.Label(r2, text=f"Rs.{abs(net_position):,.0f} ({'Surplus' if net_position>=0 else 'Deficit'})",
                 bg=CARD, fg=net_color, font=("Segoe UI",11,"bold")).pack(side="right")


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(f"{SOCIETY_NAME} — Maintenance Manager")
        self.geometry("1100x820"); self.minsize(920, 700)
        self.resizable(True, True); self.configure(bg=BG)
        _setup_styles(self); init_db(); load_settings(); load_residents()
        self._arr_data = None
        self._auto_backup()
        self._build()
        

    def _switch_society(self):
        self.withdraw()
        launcher = SocietyLauncherWindow()
        launcher.mainloop()
        if launcher._choice:
            global SOCIETY_NAME, SOCIETY_ADDRESS, MONTHLY_FEE, SOCIETY_START_FY, EXPENDITURE_CATEGORIES, RESIDENTS
            init_db(); load_settings(); load_residents()
            for w in self.winfo_children(): w.destroy()
            self.title(f"{SOCIETY_NAME} — Maintenance Manager")
            self._arr_data = None
            self._build()
            self.deiconify()
        else:
            self.deiconify()

    def _auto_backup(self):
        try:
            import shutil
            db = _db_path()
            if not os.path.exists(db): return
            backup_dir = os.path.join(os.path.dirname(db), "backups")
            os.makedirs(backup_dir, exist_ok=True)
            stamp = datetime.date.today().strftime("%Y%m%d")
            base = os.path.splitext(os.path.basename(db))[0]
            backup_path = os.path.join(backup_dir, f"{base}_{stamp}.db")
            if not os.path.exists(backup_path):
                shutil.copy2(db, backup_path)
                old_backups = sorted(
                    [f for f in os.listdir(backup_dir) if f.startswith(base) and f.endswith(".db")])
                for old in old_backups[:-30]:
                    try: os.remove(os.path.join(backup_dir, old))
                    except Exception: pass
        except Exception: pass


    def _build(self):
        root_frame = tk.Frame(self, bg=BG); root_frame.pack(fill="both", expand=True)

        sb = tk.Frame(root_frame, bg=SIDEBAR, width=224)
        sb.pack(side="left", fill="y"); sb.pack_propagate(False)

        sh = tk.Frame(sb, bg="#040509", height=92)
        sh.pack(fill="x"); sh.pack_propagate(False)
        icon_ring = tk.Frame(sh, bg="#0F0A2A", width=48, height=48)
        icon_ring.pack_propagate(False)
        icon_ring.place(relx=0.5, rely=0.5, anchor="center")
        tk.Label(icon_ring, text="🏠", bg="#0F0A2A", fg=ACCENT2, font=("Segoe UI", 24)).place(relx=0.5, rely=0.5, anchor="center")

        name_frame = tk.Frame(sb, bg=SIDEBAR); name_frame.pack(fill="x", padx=14, pady=(10,2))
        tk.Label(name_frame, text=SOCIETY_NAME, bg=SIDEBAR, fg=TEXT,
                 font=("Segoe UI", 10, "bold"), wraplength=196, justify="center").pack()
        tk.Label(sb, text="Maintenance Manager", bg=SIDEBAR, fg=TEXT3,
                 font=("Segoe UI", 7)).pack(pady=(0,12))

        tk.Frame(sb, bg=BORDER, height=1).pack(fill="x", padx=18, pady=2)
        tk.Label(sb, text="NAVIGATION", bg=SIDEBAR, fg=TEXT3,
                 font=("Segoe UI", 7, "bold"), anchor="w").pack(anchor="w", padx=18, pady=(10,6))

        def nav(text, icon, cmd): _sidebar_btn(sb, text, icon, cmd)
        nav("New Payment",     "💳", self._scroll_to_form)
        nav("Dashboard",       "📊", lambda: DashboardWindow(self))
        nav("Admission Fees",  "💰", lambda: AdmissionFeeWindow(self))
        nav("Records",         "📋", lambda: ViewRecordsWindow(self))
        nav("Unpaid Tracker",  "🔔", lambda: UnpaidTrackerWindow(self))
        nav("Arrears",         "📅", lambda: ArrearsWindow(self, prefill_cb=self._arrears_prefill_callback))
        nav("Accounts Ledger", "📒", lambda: LedgerWindow(self))
        nav("Reports",         "📄", lambda: ReportWindow(self))

        tk.Frame(sb, bg=BORDER, height=1).pack(fill="x", padx=18, pady=8)
        tk.Label(sb, text="SETTINGS", bg=SIDEBAR, fg=TEXT3,
                 font=("Segoe UI", 7, "bold"), anchor="w").pack(anchor="w", padx=18, pady=(0,6))
        nav(f"Manage {UNIT_LABEL}s", "🏠", self._open_flats_manager)
        nav("Settings",        "⚙", self._open_settings)
        nav("Switch Society",  "🌐", self._switch_society)

        tk.Frame(sb, bg=BORDER, height=1).pack(fill="x", padx=18, pady=8)

        today = datetime.date.today()
        dfy = today.year if today.month >= 4 else today.year - 1
        info_f = tk.Frame(sb, bg=SIDEBAR, padx=18); info_f.pack(fill="x")
        tk.Label(info_f, text=today.strftime("%d %b %Y"), bg=SIDEBAR, fg=TEXT3,
                 font=("Segoe UI", 8, "bold")).pack(anchor="w")
        tk.Label(info_f, text=fy_label(dfy), bg=SIDEBAR, fg=TEXT3,
                 font=("Segoe UI", 7)).pack(anchor="w", pady=(2,0))

        main = tk.Frame(root_frame, bg=BG); main.pack(side="left", fill="both", expand=True)

        topbar = tk.Frame(main, bg=CARD, height=62,
                           highlightbackground=BORDER, highlightthickness=1)
        topbar.pack(fill="x"); topbar.pack_propagate(False)

        accent_bar = tk.Frame(topbar, bg=ACCENT, width=4)
        accent_bar.pack(side="left", fill="y")

        top_left = tk.Frame(topbar, bg=CARD); top_left.pack(side="left", padx=20, pady=16)
        tk.Label(top_left, text=SOCIETY_NAME, bg=CARD, fg=TEXT,
                 font=("Segoe UI", 13, "bold")).pack(anchor="w")
        tk.Label(top_left, text=SOCIETY_ADDRESS or "Society Maintenance Manager", bg=CARD, fg=TEXT3,
                 font=("Segoe UI", 8)).pack(anchor="w")

        self._rno_var = tk.StringVar(value="Receipt No. auto-generated")
        rno_pill = tk.Frame(topbar, bg=PILL_BG, padx=14, pady=7,
                             highlightbackground=ACCENT, highlightthickness=1)
        rno_pill.pack(side="right", padx=20, pady=14)
        tk.Label(rno_pill, textvariable=self._rno_var, bg=PILL_BG, fg=PILL_FG,
                 font=("Segoe UI", 9, "bold")).pack()

        canvas = tk.Canvas(main, bg=BG, highlightthickness=0)
        vscroll = ttk.Scrollbar(main, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vscroll.set)
        vscroll.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        self._scroll_canvas = canvas

        content = tk.Frame(canvas, bg=BG); self._content_frame = content
        canvas.create_window((0, 0), window=content, anchor="nw", tags="content")
        content.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig("content", width=e.width))
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

        self._build_form(content)
        self._build_recent(content)

    def _open_flats_manager(self):
        FlatsManagerWindow(self, on_close=self._on_flats_changed)

    def _open_settings(self):
        SettingsWindow(self, on_close=self._on_settings_changed)

    def _on_settings_changed(self):
        load_settings(); load_residents()
        self.title(f"{SOCIETY_NAME} — Maintenance Manager")
        for w in self._content_frame.winfo_children(): w.destroy()
        self._build_form(self._content_frame)
        self._build_recent(self._content_frame)

    def _on_flats_changed(self):
        load_residents()
        for w in self._content_frame.winfo_children(): w.destroy()
        self._build_form(self._content_frame)
        self._build_recent(self._content_frame)

    def _scroll_to_form(self):
        self._scroll_canvas.yview_moveto(0)

    def _build_form(self, parent):
        fc = tk.Frame(parent, bg=CARD, highlightbackground=BORDER, highlightthickness=1)
        fc.pack(fill="x", padx=20, pady=(18, 8))

        fh = tk.Frame(fc, bg=ACCENT, height=50)
        fh.pack(fill="x"); fh.pack_propagate(False)
        tk.Label(fh, text="  💳  New Payment Entry", bg=ACCENT, fg="white",
                 font=("Segoe UI", 12, "bold")).pack(side="left", padx=20, pady=15)
        tk.Label(fh, text="Fill in the details and generate a receipt", bg=ACCENT, fg=ACCENT2,
                 font=("Segoe UI", 8)).pack(side="left", pady=15)

        body = tk.Frame(fc, bg=CARD, padx=24, pady=18); body.pack(fill="both", expand=True)

        left  = tk.Frame(body, bg=CARD); left.pack(side="left", fill="both", expand=True, padx=(0,20))
        tk.Frame(body, bg=BORDER, width=1).pack(side="left", fill="y")
        right = tk.Frame(body, bg=CARD); right.pack(side="left", fill="both", expand=True, padx=(20,0))

        _section_label(left, "FINANCIAL YEAR").pack(anchor="w", pady=(0,5))
        fy_row = tk.Frame(left, bg=CARD); fy_row.pack(fill="x", pady=(0,14))
        today = datetime.date.today(); dfy = today.year if today.month >= 4 else today.year - 1
        self._yfrom = ttk.Spinbox(fy_row, from_=2020, to=2100, width=7); self._yfrom.set(dfy)
        self._yfrom.pack(side="left")
        tk.Label(fy_row, text=" → ", bg=CARD, fg=TEXT3, font=("Segoe UI",9)).pack(side="left")
        self._yto = ttk.Spinbox(fy_row, from_=2021, to=2101, width=7); self._yto.set(dfy+1)
        self._yto.pack(side="left")
        self._fy_hint_lbl = tk.Label(fy_row, text="", bg=CARD, fg=WARNING, font=("Segoe UI",8,"bold"))
        self._fy_hint_lbl.pack(side="left", padx=8)
        self._yfrom.bind("<FocusOut>", self._sync_yr)
        self._yfrom.bind("<KeyRelease>", self._sync_yr)

        _section_label(left, "PAYMENT DATE").pack(anchor="w", pady=(2,5))
        df = tk.Frame(left, bg=CARD); df.pack(fill="x", pady=(0,14))
        td = datetime.date.today()
        self._dd = tk.StringVar(value=str(td.day).zfill(2))
        self._dm = tk.StringVar(value=str(td.month).zfill(2))
        self._dy = tk.StringVar(value=str(td.year))
        for var, lo, hi, w, sep in [(self._dd,1,31,4,"  /  "),(self._dm,1,12,4,"  /  "),(self._dy,2020,2100,7,"")]:
            ttk.Spinbox(df, from_=lo, to=hi, width=w, textvariable=var).pack(side="left")
            if sep: tk.Label(df, text=sep, bg=CARD, fg=TEXT3, font=("Segoe UI",9)).pack(side="left")

        _section_label(left, UNIT_LABEL.upper()).pack(anchor="w", pady=(2,5))
        self._flat = tk.StringVar()
        flat_keys = list(RESIDENTS.keys())
        flat_cb = ttk.Combobox(left, textvariable=self._flat, values=flat_keys,
                                width=14, state="readonly", font=("Segoe UI",11,"bold"))
        flat_cb.pack(anchor="w", pady=(0,4))
        flat_cb.bind("<<ComboboxSelected>>", self._on_flat)

        self._arr_slot = tk.Frame(left, bg=CARD)
        self._arr_slot.pack(fill="x", pady=(0,8))
        self._arr_frame = tk.Frame(self._arr_slot, bg=WARN_BG, highlightbackground=WARNING, highlightthickness=0)
        arr_inner = tk.Frame(self._arr_frame, bg=WARN_BG); arr_inner.pack(fill="x", padx=10, pady=8)
        self._arr_lbl = tk.Label(arr_inner, text="", bg=WARN_BG, fg=WARN_FG,
                                  font=("Segoe UI",9,"bold"), justify="left", wraplength=300)
        self._arr_lbl.pack(side="left")
        _icon_btn(arr_inner, "Pre-fill", WARNING, WARNING_H, cmd=self._do_prefill_arrears).pack(side="right")

        _section_label(left, OWNER_LABEL.upper()).pack(anchor="w", pady=(2,5))
        self._owner = tk.StringVar()
        ttk.Entry(left, textvariable=self._owner, width=32, state="readonly").pack(anchor="w", pady=(0,12))

        _section_label(left, "MOBILE").pack(anchor="w", pady=(2,5))
        mob_f = tk.Frame(left, bg=CARD); mob_f.pack(fill="x", pady=(0,12))
        self._mobile = tk.StringVar()
        ttk.Entry(mob_f, textvariable=self._mobile, width=18).pack(side="left")
        tk.Label(mob_f, text=" editable", bg=CARD, fg=TEXT3, font=("Segoe UI",8,"italic")).pack(side="left")

        _section_label(right, "MONTHLY FEE (Rs.)").pack(anchor="w", pady=(0,5))
        fee_f = tk.Frame(right, bg=CARD); fee_f.pack(fill="x", pady=(0,14))
        self._fee = tk.StringVar(value=str(MONTHLY_FEE))
        ttk.Entry(fee_f, textvariable=self._fee, width=12).pack(side="left")
        self._fee_hint = tk.Label(fee_f, text="", bg=CARD, fg=ACCENT, font=("Segoe UI",8,"italic"))
        self._fee_hint.pack(side="left", padx=8)
        self._fee.trace_add("write", self._auto_calc)

        _section_label(right, "AMOUNT PAID (Rs.)").pack(anchor="w", pady=(2,5))
        amt_f = tk.Frame(right, bg=CARD); amt_f.pack(fill="x", pady=(0,14))
        self._amt = tk.StringVar(); self._amt.trace_add("write", self._auto_calc)
        self._amt_entry = ttk.Entry(amt_f, textvariable=self._amt, width=16, font=("Segoe UI",13,"bold"))
        self._amt_entry.pack(side="left")
        self._amt_hint = tk.Label(amt_f, text="", bg=CARD, fg=SUCCESS, font=("Segoe UI",9,"bold"))
        self._amt_hint.pack(side="left", padx=10)

        _section_label(right, "LATE FEE (Rs.)  — optional").pack(anchor="w", pady=(2,5))
        lf_f = tk.Frame(right, bg=CARD); lf_f.pack(fill="x", pady=(0,10))
        self._late_fee = tk.StringVar(value="0")
        ttk.Entry(lf_f, textvariable=self._late_fee, width=12).pack(side="left")
        self._lf_hint = tk.Label(lf_f, text="auto-suggest →", bg=CARD, fg=TEXT3, font=("Segoe UI",8,"italic"))
        self._lf_hint.pack(side="left", padx=6)
        _icon_btn(lf_f, "Suggest", TEAL, TEAL_H, cmd=self._suggest_late_fee).pack(side="left", padx=2)

        _section_label(right, "PAYMENT PERIOD").pack(anchor="w", pady=(2,5))
        pf = tk.Frame(right, bg=CARD); pf.pack(fill="x", pady=(0,12))
        self._mfrom = tk.StringVar(); self._mto = tk.StringVar()
        self._mfrom_cb = ttk.Combobox(pf, textvariable=self._mfrom, values=[""]+FY_MONTHS, width=11)
        self._mfrom_cb.pack(side="left")
        self._mfrom_cb.bind("<<ComboboxSelected>>", self._on_mfrom_select)
        tk.Label(pf, text=" → ", bg=CARD, fg=TEXT3, font=("Segoe UI",9)).pack(side="left")
        self._mto_cb = ttk.Combobox(pf, textvariable=self._mto, values=[""]+FY_MONTHS, width=11)
        self._mto_cb.pack(side="left")
        self._mto_cb.bind("<<ComboboxSelected>>", lambda *_: self._update_rno_preview())
        tk.Label(pf, text=" auto", bg=CARD, fg=TEXT3, font=("Segoe UI",8)).pack(side="left")

        self._xfy_banner = tk.Label(right, text="", bg=WARN_BG, fg=WARN_FG,
                                     font=("Segoe UI",9,"bold"), anchor="w", padx=10, pady=6)
        self._xfy_banner.pack_forget()

        _h_sep(right, BORDER2).pack(fill="x", pady=(8,14))

        btn_grid = tk.Frame(right, bg=CARD); btn_grid.pack(fill="x")
        _icon_btn(btn_grid, "Clear", BORDER2, BORDER, fg=TEXT2, cmd=self._clear).grid(row=0, column=0, padx=(0,6), pady=4, sticky="ew")
        _icon_btn(btn_grid, "Preview Receipt", ACCENT, ACCENT_H, cmd=self._generate).grid(row=0, column=1, padx=(0,0), pady=4, sticky="ew")
        _icon_btn(btn_grid, "Save + PDF", PURPLE, PURPLE_H, cmd=self._save_pdf).grid(row=1, column=0, padx=(0,6), pady=4, sticky="ew")
        _icon_btn(btn_grid, "Send WhatsApp", WA_GRN, WA_GRN_H, cmd=self._send_wa).grid(row=1, column=1, padx=0, pady=4, sticky="ew")
        btn_grid.columnconfigure(0, weight=1); btn_grid.columnconfigure(1, weight=1)

    def _build_recent(self, parent):
        rc = tk.Frame(parent, bg=CARD, highlightbackground=BORDER, highlightthickness=1)
        rc.pack(fill="x", padx=20, pady=(0,20))

        rh = tk.Frame(rc, bg="#0A0D1A", height=46)
        rh.pack(fill="x"); rh.pack_propagate(False)
        tk.Label(rh, text="  🕘  Recent Payments", bg="#0A0D1A", fg=TEXT,
                 font=("Segoe UI", 10, "bold")).pack(side="left", padx=18, pady=13)
        tk.Label(rh, text="latest 10 entries", bg="#0A0D1A", fg=TEXT3,
                 font=("Segoe UI", 8)).pack(side="left")

        cols = ("Receipt No","Date",UNIT_LABEL,OWNER_LABEL,"Period","Fee/Mo","Total")
        self._tree = ttk.Treeview(rc, columns=cols, show="headings", height=6)
        for c, w in zip(cols, (132,88,44,165,112,72,82)):
            self._tree.heading(c, text=c)
            self._tree.column(c, width=w, anchor="w" if c in (OWNER_LABEL,"Period") else "center")
        vsb = ttk.Scrollbar(rc, orient="vertical", command=self._tree.yview)
        self._tree.configure(yscrollcommand=vsb.set)
        self._tree.pack(side="left", fill="both", expand=True, padx=(8,0), pady=8)
        vsb.pack(side="right", fill="y", pady=8)
        self._tree.tag_configure("arrears", background=WARN_BG, foreground=WARN_FG)
        self._tree.tag_configure("normal",  background=ROW_ODD,  foreground=TEXT)
        self._tree.tag_configure("alt",     background=ROW_EVEN, foreground=TEXT)
        self._refresh_tree()

    def _sync_yr(self, _=None):
        try:
            yf = int(self._yfrom.get()); self._yto.set(yf + 1)
            current_fy = get_current_fy()
            self._fy_hint_lbl.configure(text="← Past FY (arrears)" if yf < current_fy else "")
        except Exception: pass
        self._update_month_dropdowns()


    def _on_flat(self, _=None):
        f = self._flat.get()
        if f in RESIDENTS:
            self._owner.set(RESIDENTS[f]["name"]); self._mobile.set(RESIDENTS[f]["mobile"])
        self._update_arrears_banner()
        self._update_rno_preview()
        self._update_month_dropdowns()

    def _on_mfrom_select(self, _=None):
        self._auto_calc(); self._update_rno_preview()

    def _update_month_dropdowns(self):
        flat = self._flat.get()
        try: yf = int(self._yfrom.get())
        except Exception: yf = get_current_fy()

        if flat and flat in RESIDENTS:
            matrix = get_payment_matrix_with_fees(yf)
            unpaid = [""] + [m for m in FY_MONTHS if matrix.get(flat, {}).get(m, 0) == 0]
        else:
            unpaid = [""] + FY_MONTHS  # no flat selected, show all

    # Update both comboboxes
        for cb_var, cb_widget in [(self._mfrom, "_mfrom_cb"), (self._mto, "_mto_cb")]:
            widget = getattr(self, cb_widget, None)
            if widget:
                current_val = cb_var.get()
                widget["values"] = unpaid
            # Clear selection if previously selected month is now paid
                if current_val and current_val not in unpaid:
                    cb_var.set("")



    def _update_arrears_banner(self):
        flat = self._flat.get()
        if not flat:
            self._arr_frame.pack_forget(); self._arr_data = None; return
        try: fee = float(self._fee.get())
        except Exception: fee = MONTHLY_FEE
        arr = get_arrears_for_flat(flat, fee)
        self._arr_data = arr
        if arr["count"] == 0:
            self._arr_frame.pack_forget(); return
        fy_parts = [f"{fy_label(fy)}: {len(m)} mo" for fy, m in sorted(arr["by_fy"].items())]
        self._arr_lbl.configure(
            text=(f"⚠  Flat {flat} has {arr['count']} month{'s' if arr['count']>1 else ''} "
                  f"unpaid  (Rs.{arr['total_owed']:,.0f})   [{' + '.join(fy_parts)}]"))
        self._arr_frame.configure(highlightthickness=1)
        self._arr_frame.pack(fill="x")

    def _arrears_prefill_callback(self, flat_no, fy_year, months_list):
        self._flat.set(flat_no)
        if flat_no in RESIDENTS:
            self._owner.set(RESIDENTS[flat_no]["name"])
            self._mobile.set(RESIDENTS[flat_no]["mobile"])
        self._update_arrears_banner()
        arr = self._arr_data
        if arr and len(arr.get("by_fy", {})) > 1:
            fy_years = sorted(arr["by_fy"].keys())
            chosen_fy = self._ask_arrears_fy(flat_no, fy_years, arr)
            if chosen_fy is None:
                return
            months_list = arr["by_fy"][chosen_fy]
            fy_year = chosen_fy
        use_current = self._ask_fy_booking_preference(flat_no, fy_year)
        if use_current is None:
            return
        self._prefill_for_arrears(flat_no, fy_year, months_list, use_current_fy=use_current)

    def _do_prefill_arrears(self):
        flat = self._flat.get()
        if not flat or not self._arr_data or self._arr_data["count"] == 0:
            return
        arr = self._arr_data
        fy_years = sorted(arr["by_fy"].keys())
        chosen_fy = fy_years[0] if len(fy_years) == 1 else self._ask_arrears_fy(flat, fy_years, arr)
        if chosen_fy is None:
            return
        use_current = self._ask_fy_booking_preference(flat, chosen_fy)
        if use_current is None:
            return
        self._prefill_for_arrears(flat, chosen_fy, arr["by_fy"][chosen_fy], use_current_fy=use_current)

    def _prefill_for_arrears(self, flat_no, fy_year, months_list, use_current_fy=False):
        self._flat.set(flat_no)
        if flat_no in RESIDENTS:
            self._owner.set(RESIDENTS[flat_no]["name"])
            self._mobile.set(RESIDENTS[flat_no]["mobile"])

        current_fy = get_current_fy()
        booking_fy = current_fy if use_current_fy else fy_year

        self._yfrom.set(booking_fy)
        self._yto.set(booking_fy + 1)
        self._sync_yr()

        try:
            fee = float(self._fee.get())
        except Exception:
            fee = MONTHLY_FEE
        total_amt = round(fee * len(months_list))
        self._amt.set(str(total_amt))

        period_note = ""
        if use_current_fy:
            self._mfrom.set("")
            self._mto.set("")
            period_note = f"Booked under Current FY ({fy_label(current_fy)}) as lump-sum\nNo period set — months remain unpaid in past-FY matrix."
        else:
            months_set = set(months_list)
            contiguous = True
            if len(months_list) > 1:
                start_idx = FY_MONTHS.index(months_list[0])
                end_idx   = FY_MONTHS.index(months_list[-1])
                for i in range(start_idx, end_idx + 1):
                    if FY_MONTHS[i] not in months_set:
                        contiguous = False
                        break

            if contiguous:
                self._mfrom.set(months_list[0])
                self._mto.set(months_list[-1])
                period_note = f"Booked under Past FY ({fy_label(fy_year)}) — months marked paid."
            else:
                self._mfrom.set("")
                self._mto.set("")
                gap_months = [m for m in FY_MONTHS
                              if FY_MONTHS.index(months_list[0]) <= FY_MONTHS.index(m) <= FY_MONTHS.index(months_list[-1])
                              and m not in months_set]
                period_note = (
                    f"⚠  Non-contiguous months detected.\n"
                    f"Unpaid: {', '.join(months_list)}\n"
                    f"Already paid in range: {', '.join(gap_months)}\n\n"
                    f"Period fields left blank to avoid overwriting paid months.\n"
                    f"Set Month From / To manually after reviewing."
                )

        self._auto_calc()
        self._update_rno_preview()
        self._update_arrears_banner()

        n = len(months_list)
        messagebox.showinfo(
            "Pre-filled",
            f"Form ready for Flat {flat_no}\n"
            f"Arrears period: {months_list[0]} → {months_list[-1]}  ({n} month{'s' if n != 1 else ''})\n"
            f"Amount: Rs.{total_amt:,.0f}\n\n"
            f"{period_note}"
        )

    def _ask_arrears_fy(self, flat_no, fy_years, arr):
        dlg = tk.Toplevel(self); dlg.title(f"Select FY — Flat {flat_no}")
        dlg_h = 160 + len(fy_years) * 38
        dlg.geometry(f"380x{dlg_h}"); dlg.resizable(False, False); dlg.configure(bg=CARD); dlg.grab_set()
        tk.Label(dlg, text=f"  Select FY to settle", bg=HEAD_I, fg="white",
                 font=("Segoe UI",10,"bold"), pady=12).pack(fill="x")
        tk.Label(dlg, text="Select financial year:", bg=CARD, fg=TEXT, font=("Segoe UI",10), pady=8).pack(padx=20, anchor="w")
        chosen = tk.StringVar(value=str(fy_years[0]))
        try: fee = float(self._fee.get())
        except Exception: fee = MONTHLY_FEE
        for fy in fy_years:
            months = arr["by_fy"][fy]; n = len(months)
            ttk.Radiobutton(dlg, text=f"{fy_label(fy)}: {n} month{'s' if n != 1 else ''} = Rs.{n*fee:,.0f}",
                            variable=chosen, value=str(fy)).pack(anchor="w", padx=30, pady=5)
        result = [None]
        def _ok():
            try: result[0] = int(chosen.get())
            except Exception: pass
            dlg.destroy()
        tk.Frame(dlg, bg=BORDER, height=1).pack(fill="x", pady=(8,0))
        bf = tk.Frame(dlg, bg=CARD); bf.pack(pady=14)
        _icon_btn(bf, "OK", ACCENT, ACCENT_H, cmd=_ok).pack(side="left", padx=6)
        _icon_btn(bf, "Cancel", DANGER, DANGER_H, cmd=dlg.destroy).pack(side="left", padx=6)
        dlg.wait_window(); return result[0]

    def _ask_fy_booking_preference(self, flat_no, arrears_fy):
        current_fy = get_current_fy()
        if arrears_fy >= current_fy:
            return False

        dlg = tk.Toplevel(self)
        dlg.title("Arrears Booking")
        dlg.geometry("440x330")
        dlg.resizable(False, False)
        dlg.configure(bg=CARD)
        dlg.grab_set()

        tk.Label(dlg, text=f"  How to record this arrears payment?",
                 bg=HEAD_I, fg="white", font=("Segoe UI", 10, "bold"), pady=12).pack(fill="x")
        tk.Label(dlg, text=f"  Flat {flat_no}  ·  Arrears from {fy_label(arrears_fy)}",
                 bg=BG, fg=TEXT2, font=("Segoe UI", 8), pady=5).pack(fill="x")

        result = [None]
        def _ok():
            result[0] = (choice.get() == "current")
            dlg.destroy()

        tk.Frame(dlg, bg=BORDER, height=1).pack(side="bottom", fill="x")
        bf = tk.Frame(dlg, bg=CARD, padx=20, pady=12); bf.pack(side="bottom", fill="x")
        _icon_btn(bf, "Continue", SUCCESS, SUCCESS_H, cmd=_ok).pack(side="left", padx=(0, 8))
        _icon_btn(bf, "Cancel", DANGER, DANGER_H, cmd=dlg.destroy).pack(side="left")

        body = tk.Frame(dlg, bg=CARD, padx=24, pady=16); body.pack(fill="both", expand=True)
        choice = tk.StringVar(value="past")

        opt1 = tk.Frame(body, bg=BG, highlightbackground=BORDER2, highlightthickness=1)
        opt1.pack(fill="x", pady=4)
        inn1 = tk.Frame(opt1, bg=BG, padx=14, pady=10); inn1.pack(fill="x")
        ttk.Radiobutton(inn1, text=f"Book under Past FY  ({fy_label(arrears_fy)})",
                        variable=choice, value="past").pack(anchor="w")
        tk.Label(inn1, text="Months will be marked as paid in the original year.",
                 bg=BG, fg=TEXT3, font=("Segoe UI", 8, "italic")).pack(anchor="w", padx=20)

        opt2 = tk.Frame(body, bg=BG, highlightbackground=BORDER2, highlightthickness=1)
        opt2.pack(fill="x", pady=4)
        inn2 = tk.Frame(opt2, bg=BG, padx=14, pady=10); inn2.pack(fill="x")
        ttk.Radiobutton(inn2, text=f"Book under Current FY  ({fy_label(current_fy)})",
                        variable=choice, value="current").pack(anchor="w")
        tk.Label(inn2, text="Recorded as a lump-sum receipt this year. Past months stay unpaid in matrix.",
                 bg=BG, fg=TEXT3, font=("Segoe UI", 8, "italic")).pack(anchor="w", padx=20)

        dlg.wait_window()
        return result[0]

    def _auto_calc(self, *_):
        try: amt = float(self._amt.get())
        except Exception:
            self._amt_hint.configure(text=""); self._fee_hint.configure(text="")
            self._xfy_banner.pack_forget(); self._update_rno_preview(); return
        try:
            fee = float(self._fee.get())
            if fee <= 0: raise ValueError
        except Exception:
            self._amt_hint.configure(text=""); self._fee_hint.configure(text="")
            self._xfy_banner.pack_forget(); self._update_rno_preview(); return
        n_months = calc_months_from_amount(amt, fee)
        self._fee_hint.configure(text=f"({n_months} mo @ Rs.{fee:,.0f})")
        remainder = amt - fee * n_months
        self._amt_hint.configure(
            text=f"= {n_months} mo" + (f" +Rs.{remainder:,.0f}" if remainder > 0 else ""),
            fg=WARNING if remainder > 0 else SUCCESS)
        mf = self._mfrom.get().strip()
        if mf and mf in FY_MONTHS and n_months >= 1:
            this_n, next_n, overflows, this_end, next_end = check_fy_overflow(mf, n_months)
            if overflows:
                self._mto.set("March")
                try: yf = int(self._yfrom.get())
                except Exception: yf = get_current_fy()
                self._xfy_banner.configure(
                    text=f"⚠  Cross-FY: {mf}–Mar ({this_n}mo) + Apr–{next_end} ({next_n}mo) → 2 records")
                self._xfy_banner.pack(fill="x", pady=(0, 8))
            else:
                self._mto.set(advance_fy_months(mf, n_months)); self._xfy_banner.pack_forget()
        else:
            if mf and mf in FY_MONTHS: self._mto.set(advance_fy_months(mf, n_months))
            self._xfy_banner.pack_forget()
        self._update_rno_preview()

    def _update_rno_preview(self):
        try: yf = int(self._yfrom.get())
        except Exception:
            td = datetime.date.today(); yf = td.year if td.month >= 4 else td.year - 1
        mf = self._mfrom.get().strip(); mt = self._mto.get().strip()
        self._rno_var.set(f"Preview: {_build_receipt_no(mf, mt, yf)}")

    def _collect(self):
        flat = self._flat.get().strip(); amt_s = self._amt.get().strip()
        if not flat:
            messagebox.showerror("Missing", f"Please select a {UNIT_LABEL} No."); return None
        if not amt_s:
            messagebox.showerror("Missing", "Please enter an Amount."); return None
        try:
            amt = float(amt_s)
            if amt <= 0: raise ValueError
        except ValueError:
            messagebox.showerror("Invalid", "Amount must be positive."); return None
        try:
            fee = float(self._fee.get())
            if fee < 0: fee = 0.0
        except Exception: fee = 0.0
        date = f"{self._dd.get().zfill(2)}/{self._dm.get().zfill(2)}/{self._dy.get()}"
        mf = self._mfrom.get().strip(); mt = self._mto.get().strip()
        try: yf = int(self._yfrom.get())
        except Exception:
            td = datetime.date.today(); yf = td.year if td.month >= 4 else td.year - 1
        try: yt = int(self._yto.get())
        except Exception: yt = yf + 1
        try: late = float(self._late_fee.get()); late = max(0.0, late)
        except Exception: late = 0.0
        return {"receipt_no": _build_receipt_no(mf, mt, yf),
                "date": date, "year_from": yf, "year_to": yt,
                "flat_no": flat, "owner_name": self._owner.get(),
                "mobile": self._mobile.get().strip(),
                "amount": amt, "month_from": mf, "month_to": mt, "monthly_fee": fee, "late_fee": late}

    def _check_overflow(self, base):
        mf = base.get("month_from", ""); fee = float(base.get("monthly_fee", 0) or 0)
        amt = float(base.get("amount", 0))
        if not mf or mf not in FY_MONTHS or fee <= 0:
            return False, 0, 0, "", "", None, None
        n = calc_months_from_amount(amt, fee)
        this_n, next_n, overflows, this_end, next_end = check_fy_overflow(mf, n)
        if not overflows: return False, 0, 0, "", "", None, None
        rec_a, rec_b = build_split_records(base, this_n, next_n, this_end, next_end)
        return True, this_n, next_n, this_end, next_end, rec_a, rec_b

    def _confirm_past_fy(self, d):
        current_fy = get_current_fy(); yf = d.get("year_from", current_fy)
        if yf >= current_fy: return True
        return messagebox.askyesno("Arrears Payment",
            f"Recording payment for past FY:\n  {fy_label(yf)}\n\nContinue?", icon="question")

    def _generate(self):
        d = self._collect()
        if not d: return
        if not self._confirm_past_fy(d): return
        overflows, this_n, next_n, this_end, next_end, rec_a, rec_b = self._check_overflow(d)
        if overflows:
            dlg = CrossFYSplitDialog(self, rec_a, rec_b, this_n, next_n)
            self.wait_window(dlg)
            if not dlg.confirmed: return
            ReceiptDialog(self, rec_a); ReceiptDialog(self, rec_b, after_save=self._on_saved)
        else:
            ReceiptDialog(self, d, after_save=self._on_saved)

    def _save_pdf(self):
        d = self._collect()
        if not d: return
        if not self._confirm_past_fy(d): return
        overflows, this_n, next_n, this_end, next_end, rec_a, rec_b = self._check_overflow(d)
        if overflows:
            dlg = CrossFYSplitDialog(self, rec_a, rec_b, this_n, next_n)
            self.wait_window(dlg)
            if not dlg.confirmed: return
            for rec in (rec_a, rec_b):
                if not db_save_payment(rec): return
            self._on_saved()
            for rec in (rec_a, rec_b):
                fp = filedialog.asksaveasfilename(defaultextension=".pdf",
                                                  filetypes=[("PDF", "*.pdf")],
                                                  initialfile=f"Receipt_{rec['receipt_no']}.pdf")
                if fp and generate_receipt_pdf(rec, fp):
                    if messagebox.askyesno("Open?", f"Open {rec['receipt_no']}?"):
                        webbrowser.open(f"file:///{fp}")
        else:
            if not db_save_payment(d): return
            self._on_saved()
            fp = filedialog.asksaveasfilename(defaultextension=".pdf",
                                              filetypes=[("PDF", "*.pdf")],
                                              initialfile=f"Receipt_{d['receipt_no']}.pdf")
            if fp and generate_receipt_pdf(d, fp):
                messagebox.showinfo("PDF Saved", f"Saved:\n{fp}")
                if messagebox.askyesno("Open?", "Open now?"): webbrowser.open(f"file:///{fp}")

    def _send_wa(self):
        d = self._collect()
        if not d: return
        mob = d["mobile"]
        if not mob:
            messagebox.showerror("No Mobile", "No mobile number. Enter it first."); return
        if not self._confirm_past_fy(d): return
        overflows, this_n, next_n, this_end, next_end, rec_a, rec_b = self._check_overflow(d)
        if overflows:
            dlg = CrossFYSplitDialog(self, rec_a, rec_b, this_n, next_n)
            self.wait_window(dlg)
            if not dlg.confirmed: return
            for rec in (rec_a, rec_b):
                if not db_save_payment(rec): return
            self._on_saved()
            combined = build_receipt(rec_a) + "\n\n" + "-" * 30 + "\n" + build_receipt(rec_b)
            WhatsAppDialog(self, mob, combined, title_extra=f"Flat {d['flat_no']} (2 records)")
        else:
            if db_save_payment(d):
                self._on_saved()
                WhatsAppDialog(self, mob, build_receipt(d), title_extra=f"Flat {d['flat_no']}")

    def _suggest_late_fee(self):
        flat = self._flat.get()
        if not flat:
            messagebox.showinfo("Select Flat", "Select a flat first.", parent=self); return
        try: fee = float(self._fee.get())
        except Exception: fee = MONTHLY_FEE
        arr = get_arrears_for_flat(flat, fee)
        if arr["count"] == 0:
            self._lf_hint.configure(text="No arrears — no late fee", fg=SUCCESS); return
        rate = max(50.0, fee * 0.02)
        suggested = round(arr["count"] * rate)
        self._late_fee.set(str(int(suggested)))
        self._lf_hint.configure(text=f"Rs.{rate:.0f}/mo × {arr['count']} mo", fg=TEAL)

    def _on_saved(self): self._refresh_tree(); self._clear()

    def _clear(self):
        for v in (self._flat, self._owner, self._mobile, self._amt, self._mfrom, self._mto):
            v.set("")
        self._fee.set(str(MONTHLY_FEE)); self._late_fee.set("0")
        self._amt_hint.configure(text=""); self._fee_hint.configure(text="")
        self._lf_hint.configure(text="auto-suggest →", fg=TEXT3)
        self._fy_hint_lbl.configure(text="")
        self._rno_var.set("Receipt No. auto-generated")
        self._xfy_banner.pack_forget()
        self._arr_frame.pack_forget()
        self._arr_data = None
        today = datetime.date.today(); dfy = today.year if today.month >= 4 else today.year - 1
        self._yfrom.set(dfy); self._yto.set(dfy + 1)

    def _refresh_tree(self):
        for i in self._tree.get_children(): self._tree.delete(i)
        current_fy = get_current_fy()
        for idx, r in enumerate(db_fetch_payments()[:10]):
            _, rno, date, yf, yt, flat, owner, amt, mf, mt, fee = r
            if mf and mt and mf != mt: period = f"{mf[:3]}-{mt[:3]} ({months_in_range(mf, mt)}mo)"
            elif mf: period = f"{mf[:3]} (1mo)"
            else: period = "-"
            fee_s = f"Rs.{fee:,.0f}" if fee > 0 else "-"
            is_arr = yf < current_fy
            if is_arr: period = "★ " + period
            tag = "arrears" if is_arr else ("alt" if idx % 2 else "normal")
            self._tree.insert("", "end", values=(rno, date, flat, owner, period, fee_s, f"Rs.{amt:,.0f}"),
                              tags=(tag,))


if __name__ == "__main__":
    launcher = SocietyLauncherWindow()
    launcher.mainloop()
    if not launcher._choice:
        raise SystemExit
    app = App()
    app.mainloop()
